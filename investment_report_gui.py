"""
Investment Report GUI - Interfaz gráfica para el generador de reportes
Dark theme, pestañas múltiples, análisis técnico + Volumen MA
Compatible con N26 (~120 acciones)
"""
import sys
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except AttributeError:
        pass

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import queue
import os
from datetime import datetime
import pandas as pd
import numpy as np
from scoring_service import (
    compute_buy_score,
    compute_bond_score,
    compute_dip_score,
    compute_bigdip_score,
    compute_confidence,
)

# ─────────────────────────────────────────────
#  COLORES Y TEMA
# ─────────────────────────────────────────────

from gui.colors import COLORS  # noqa: E402  (import after sys.platform block)

# ─────────────────────────────────────────────
#  ESTILOS TTK
# ─────────────────────────────────────────────

def apply_styles(root):
    style = ttk.Style(root)
    style.theme_use('clam')

    style.configure('.',
        background=COLORS['bg'],
        foreground=COLORS['text'],
        fieldbackground=COLORS['surface2'],
        bordercolor=COLORS['border'],
        darkcolor=COLORS['surface'],
        lightcolor=COLORS['surface2'],
        troughcolor=COLORS['surface2'],
        selectbackground=COLORS['primary_dark'],
        selectforeground=COLORS['text'],
        font=('Segoe UI', 10),
    )

    style.configure('TNotebook',
        background=COLORS['bg'],
        borderwidth=0,
        tabmargins=0,
    )
    style.configure('TNotebook.Tab',
        background=COLORS['surface'],
        foreground=COLORS['text_dim'],
        padding=[20, 10],
        font=('Segoe UI', 10, 'bold'),
        borderwidth=0,
    )
    style.map('TNotebook.Tab',
        background=[('selected', COLORS['surface2']), ('active', COLORS['surface3'])],
        foreground=[('selected', COLORS['primary']), ('active', COLORS['text'])],
    )

    style.configure('Treeview',
        background=COLORS['surface'],
        foreground=COLORS['text'],
        fieldbackground=COLORS['surface'],
        rowheight=28,
        borderwidth=0,
        font=('Consolas', 9),
    )
    style.configure('Treeview.Heading',
        background=COLORS['surface2'],
        foreground=COLORS['text_dim'],
        relief='flat',
        font=('Segoe UI', 9, 'bold'),
        borderwidth=1,
    )
    style.map('Treeview',
        background=[('selected', COLORS['primary_dark'])],
        foreground=[('selected', COLORS['text'])],
    )
    style.map('Treeview.Heading',
        background=[('active', COLORS['surface3'])],
    )

    style.configure('TProgressbar',
        troughcolor=COLORS['surface2'],
        background=COLORS['primary'],
        borderwidth=0,
        relief='flat',
    )

    style.configure('TScrollbar',
        background=COLORS['surface2'],
        troughcolor=COLORS['surface'],
        borderwidth=0,
        arrowcolor=COLORS['text_dim'],
    )

    style.configure('TCombobox',
        fieldbackground=COLORS['surface2'],
        background=COLORS['surface2'],
        foreground=COLORS['text'],
        selectbackground=COLORS['primary_dark'],
        arrowcolor=COLORS['text_dim'],
        bordercolor=COLORS['border'],
    )
    style.map('TCombobox',
        fieldbackground=[('readonly', COLORS['surface2'])],
        selectbackground=[('readonly', COLORS['primary_dark'])],
    )


# ─────────────────────────────────────────────
#  WIDGETS PERSONALIZADOS
# ─────────────────────────────────────────────

class StyledButton(tk.Frame):
    def __init__(self, parent, text, command, color=None, width=None, **kwargs):
        bg = color or COLORS['primary']
        super().__init__(parent, bg=bg, cursor='hand2', **kwargs)
        self._bg = bg
        self._hover = self._darken(bg)
        self._command = command

        padx = 20
        pady = 10
        lkw = dict(
            text=text,
            bg=bg,
            fg=COLORS['text'],
            font=('Segoe UI', 10, 'bold'),
            cursor='hand2',
            padx=padx,
            pady=pady,
        )
        if width:
            lkw['width'] = width
        self._label = tk.Label(self, **lkw)
        self._label.pack(fill=tk.BOTH, expand=True)

        for w in (self, self._label):
            w.bind('<Button-1>', self._click)
            w.bind('<Enter>', self._enter)
            w.bind('<Leave>', self._leave)

    def _darken(self, hex_color):
        r = int(hex_color[1:3], 16)
        g = int(hex_color[3:5], 16)
        b = int(hex_color[5:7], 16)
        r = max(0, r - 30)
        g = max(0, g - 30)
        b = max(0, b - 30)
        return f'#{r:02x}{g:02x}{b:02x}'

    def _click(self, e):
        if self._command:
            self._command()

    def _enter(self, e):
        self.config(bg=self._hover)
        self._label.config(bg=self._hover)

    def _leave(self, e):
        self.config(bg=self._bg)
        self._label.config(bg=self._bg)

    def set_enabled(self, enabled):
        if enabled:
            self.config(bg=self._bg)
            self._label.config(bg=self._bg, fg=COLORS['text'], cursor='hand2')
            self.config(cursor='hand2')
        else:
            self.config(bg=COLORS['surface3'])
            self._label.config(bg=COLORS['surface3'], fg=COLORS['text_dim'], cursor='arrow')
            self.config(cursor='arrow')


class MetricCard(tk.Frame):
    def __init__(self, parent, label, value, value_color=None, **kwargs):
        super().__init__(parent, bg=COLORS['surface'], padx=16, pady=12, **kwargs)
        tk.Label(self, text=label, bg=COLORS['surface'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(anchor='w')
        self._val_label = tk.Label(self, text=value, bg=COLORS['surface'],
                                   fg=value_color or COLORS['text'],
                                   font=('Segoe UI', 20, 'bold'))
        self._val_label.pack(anchor='w', pady=(2, 0))

    def update_value(self, value, color=None):
        self._val_label.config(text=value)
        if color:
            self._val_label.config(fg=color)


# ─────────────────────────────────────────────
#  VENTANA PRINCIPAL
# ─────────────────────────────────────────────

from gui.tabs.journal_tab import JournalTabMixin
from gui.tabs.alerts_mixin import AlertsMixin
from gui.tabs.portfolio_risk_tab import PortfolioRiskTabMixin
from gui.tabs.ml_tab import MLTabMixin
from gui.data_controller import DataController


class InvestmentReportGUI(JournalTabMixin, AlertsMixin, PortfolioRiskTabMixin, MLTabMixin):
    def __init__(self, root):
        self.root = root
        self.root.title("Investment Report - N26 Portfolio Analyzer")
        self.root.geometry("1280x820")
        self.root.minsize(1024, 700)
        self.root.configure(bg=COLORS['bg'])

        apply_styles(root)

        # ── DataController (4.2) ─────────────────────────────────────────────
        self.ctrl = DataController()

        # Estado interno
        self.df_results = None
        self.portfolio  = []
        self.running    = False
        self.q          = queue.Queue()
        self._vix_data  = None
        self._spy_data  = None
        self._last_analysis_time = None
        self._scheduler_timer  = None
        self._earnings_cache   = {}
        self._notif_enabled    = None

        # Claude IA (opcional — se activa si el SDK o CLI están disponibles)
        try:
            from claude_analyzer import ClaudeAnalyzer, CLAUDE_AVAILABLE
            self._claude_analyzer = ClaudeAnalyzer()
            self._claude_available = CLAUDE_AVAILABLE
            self._claude_mode_label = self._claude_analyzer.mode_label()
        except Exception:
            self._claude_analyzer  = None
            self._claude_available = False
            self._claude_mode_label = 'No disponible'
        self._ia_results = []
        self._ia_running  = False
        self._data_updating = False
        self._chart_canvas = None
        self._chart_fig    = None

        # Objetos de análisis (se importan lazy para no bloquear la GUI)
        self.db = None
        self.analyzer = None

        # Configuración (editable desde Settings)
        self.cfg = {
            'interval': '1d',
            'portfolio_n': 12,
            'max_corr': 0.70,
            'force_download': True,
        }

        self._build_ui()

        # ── Threading (4.3): event_generate desde workers + fallback 500ms ──
        # Los workers llaman self._q_notify() que dispara '<<QueueUpdate>>'.
        # root.after(500, ...) actúa como red de seguridad si el evento se pierde.
        self.root.bind('<<QueueUpdate>>', lambda _e: self._poll_queue())
        self._start_poll_fallback()

        # Cargar resultados del último análisis (si existen)
        self.root.after(200, self._load_last_results)
        # Verificar frescura de datos en la BD
        self.root.after(500, self._check_data_freshness)

    def _q_notify(self) -> None:
        """Notifica al hilo principal que hay mensajes en la cola.
        Llamado desde workers (hilo secundario) — thread-safe en tkinter."""
        try:
            self.root.event_generate('<<QueueUpdate>>', when='tail')
        except Exception:
            pass  # root destruido al cerrar la app

    def _start_poll_fallback(self) -> None:
        """Fallback de 500 ms para mensajes cuyo evento se haya perdido."""
        self._poll_queue()
        self.root.after(500, self._start_poll_fallback)

    # ─── BUILD UI ─────────────────────────────

    def _build_ui(self):
        # Header
        self._build_header()

        # Notebook (tabs)
        self.nb = ttk.Notebook(self.root)
        self.nb.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)

        # ── Tabs de DECISIÓN (nivel principal — amigables) ────────────────
        self.tab_dashboard  = tk.Frame(self.nb, bg=COLORS['bg'])
        self.tab_topbuys    = tk.Frame(self.nb, bg=COLORS['bg'])
        self.tab_dip        = tk.Frame(self.nb, bg=COLORS['bg'])
        self.tab_pred       = tk.Frame(self.nb, bg=COLORS['bg'])
        self.tab_portfolio  = tk.Frame(self.nb, bg=COLORS['bg'])
        self.tab_journal         = tk.Frame(self.nb, bg=COLORS['bg'])
        self.tab_portfolio_risk  = tk.Frame(self.nb, bg=COLORS['bg'])
        self.tab_ml              = tk.Frame(self.nb, bg=COLORS['bg'])

        # ── Super-tab ANÁLISIS con sub-notebook ───────────────────────────
        self.tab_analysis_container = tk.Frame(self.nb, bg=COLORS['bg'])
        self.nb_analysis = ttk.Notebook(self.tab_analysis_container)
        self.nb_analysis.pack(fill=tk.BOTH, expand=True)

        # Frames que viven dentro del sub-notebook de análisis
        self.tab_rankings   = tk.Frame(self.nb_analysis, bg=COLORS['bg'])
        self.tab_volume     = tk.Frame(self.nb_analysis, bg=COLORS['bg'])
        self.tab_candles    = tk.Frame(self.nb_analysis, bg=COLORS['bg'])
        self.tab_multitf    = tk.Frame(self.nb_analysis, bg=COLORS['bg'])
        self.tab_rs         = tk.Frame(self.nb_analysis, bg=COLORS['bg'])
        self.tab_risk       = tk.Frame(self.nb_analysis, bg=COLORS['bg'])
        self.tab_backtest   = tk.Frame(self.nb_analysis, bg=COLORS['bg'])
        self.tab_fund       = tk.Frame(self.nb_analysis, bg=COLORS['bg'])
        self.tab_sector     = tk.Frame(self.nb_analysis, bg=COLORS['bg'])

        self.tab_chart      = tk.Frame(self.nb, bg=COLORS['bg'])
        self.tab_ia         = tk.Frame(self.nb, bg=COLORS['bg'])
        self.tab_settings   = tk.Frame(self.nb, bg=COLORS['bg'])

        # ── Registrar tabs principales ────────────────────────────────────
        self.nb.add(self.tab_dashboard,          text='  🏠 Dashboard  ')
        self.nb.add(self.tab_topbuys,            text='  ✅ ¿Qué Comprar?  ')
        self.nb.add(self.tab_dip,                text='  📉 Comprar en Dip  ')
        self.nb.add(self.tab_pred,               text='  🔮 Predicción  ')
        self.nb.add(self.tab_portfolio,          text='  💼 Portfolio  ')
        self.nb.add(self.tab_journal,            text='  📓 Mis Inversiones  ')
        self.nb.add(self.tab_portfolio_risk,     text='  📈 Riesgo Portafolio  ')
        self.nb.add(self.tab_ml,                 text='  Modelo ML  ')
        self.nb.add(self.tab_analysis_container, text='  🔬 Análisis Técnico  ')
        self.nb.add(self.tab_chart,              text='  📊 Gráfico  ')
        self.nb.add(self.tab_ia,                 text='  🤖 Análisis IA  ')
        self.nb.add(self.tab_settings,           text='  ⚙️ Configuración  ')

        # ── Registrar tabs de análisis (sub-notebook) ─────────────────────
        self.nb_analysis.add(self.tab_rankings,  text='  Rankings  ')
        self.nb_analysis.add(self.tab_volume,    text='  Volumen MA  ')
        self.nb_analysis.add(self.tab_candles,   text='  Patrones Vela  ')
        self.nb_analysis.add(self.tab_multitf,   text='  Multi-Timeframe  ')
        self.nb_analysis.add(self.tab_rs,        text='  Fuerza Relativa  ')
        self.nb_analysis.add(self.tab_risk,      text='  Riesgo  ')
        self.nb_analysis.add(self.tab_backtest,  text='  Backtesting  ')
        self.nb_analysis.add(self.tab_fund,      text='  Fundamentales  ')
        self.nb_analysis.add(self.tab_sector,    text='  Rotación Sectorial  ')

        self._build_dashboard()
        self._build_topbuys()
        self._build_dip_tab()
        self._build_prediction_tab()
        self._build_rankings()
        self._build_volume()
        self._build_candles()
        self._build_multitf()
        self._build_rs()
        self._build_risk()
        self._build_backtest()
        self._build_fundamentals()
        self._build_portfolio()
        self._build_sector_rotation()
        self._build_journal()
        self._build_portfolio_risk()
        self._build_ml_tab()
        self._build_chart_tab()
        self._build_ia_tab()
        self._build_settings()

        # Status bar
        self._build_statusbar()

    def _build_header(self):
        hdr = tk.Frame(self.root, bg=COLORS['surface'], height=64)
        hdr.pack(fill=tk.X)
        hdr.pack_propagate(False)

        # Logo
        tk.Label(hdr, text='◈', font=('Segoe UI', 28),
                 bg=COLORS['surface'], fg=COLORS['primary']).pack(side=tk.LEFT, padx=(20, 10))

        title_f = tk.Frame(hdr, bg=COLORS['surface'])
        title_f.pack(side=tk.LEFT)
        tk.Label(title_f, text='INVESTMENT REPORT', font=('Segoe UI', 16, 'bold'),
                 bg=COLORS['surface'], fg=COLORS['text']).pack(anchor='w')
        tk.Label(title_f, text='N26 Portfolio Analyzer · ~120 acciones · Análisis técnico + Volumen MA',
                 font=('Segoe UI', 9), bg=COLORS['surface'], fg=COLORS['text_dim']).pack(anchor='w')

        # Status badge
        self.hdr_status = tk.Label(hdr, text='● LISTO', font=('Segoe UI', 10, 'bold'),
                                   bg=COLORS['surface'], fg=COLORS['text_dim'],
                                   padx=16, pady=8)
        self.hdr_status.pack(side=tk.RIGHT, padx=20)

        # Última actualización
        self.hdr_last_update = tk.Label(hdr, text='', font=('Segoe UI', 9),
                                        bg=COLORS['surface'], fg=COLORS['text_dim'],
                                        padx=8, pady=8)
        self.hdr_last_update.pack(side=tk.RIGHT, padx=(0, 4))

        # DB freshness label
        self.hdr_db_fresh = tk.Label(hdr, text='', font=('Segoe UI', 9),
                                     bg=COLORS['surface'], fg=COLORS['text_dim'],
                                     padx=8, pady=8, cursor='hand2')
        self.hdr_db_fresh.pack(side=tk.RIGHT, padx=(0, 2))
        self.hdr_db_fresh.bind('<Button-1>', lambda e: self._run_data_update())

        # Botón actualizar datos
        self.btn_update_data = tk.Button(
            hdr, text='⬇ Actualizar Datos',
            bg=COLORS['surface2'], fg=COLORS['primary'],
            font=('Segoe UI', 9, 'bold'), relief='flat',
            padx=12, pady=6, cursor='hand2',
            command=self._run_data_update)
        self.btn_update_data.pack(side=tk.RIGHT, padx=(0, 8))

    # ─── PERSISTENCIA DE RESULTADOS ──────────

    def _save_last_results(self):
        """Guarda df_results + metadata en disco para precarga en próxima sesión."""
        import json
        if self.df_results is None or self.df_results.empty:
            return
        base = os.path.dirname(os.path.abspath(__file__))
        try:
            parquet_path = os.path.join(base, 'last_analysis.parquet')
            self.df_results.to_parquet(parquet_path, index=False)
            spy_meta = None
            if self._spy_data:
                spy_meta = {k: v for k, v in self._spy_data.items()
                            if not hasattr(v, 'to_json')}
            meta = {
                'timestamp': datetime.now().isoformat(),
                'vix_data': self._vix_data,
                'spy_data': spy_meta,
                'portfolio': self.portfolio,
            }
            with open(os.path.join(base, 'last_analysis_meta.json'), 'w') as f:
                json.dump(meta, f, indent=2)
            self._last_analysis_time = datetime.now()
            self._update_last_update_label()
            self._log('ok', 'Resultados guardados en disco (precarga activa para próxima sesión).')
        except Exception as e:
            self._log('warn', f'No se pudieron guardar resultados: {e}')

    def _load_last_results(self):
        """Carga resultados previos desde disco al iniciar la app.

        Intenta parquet primero; si no existe, carga el pkl legado y lo migra.
        """
        import json
        base = os.path.dirname(os.path.abspath(__file__))
        parquet_path = os.path.join(base, 'last_analysis.parquet')
        pkl_path     = os.path.join(base, 'last_analysis.pkl')
        meta_path    = os.path.join(base, 'last_analysis_meta.json')

        if not os.path.exists(meta_path):
            return

        try:
            if os.path.exists(parquet_path):
                import pandas as pd
                self.df_results = pd.read_parquet(parquet_path)
            elif os.path.exists(pkl_path):
                import pickle
                with open(pkl_path, 'rb') as f:
                    self.df_results = pickle.load(f)
                # Migrate: re-save as parquet, remove legacy pkl
                try:
                    self.df_results.to_parquet(parquet_path, index=False)
                    os.remove(pkl_path)
                except Exception:
                    pass
            else:
                return

            with open(meta_path, 'r') as f:
                meta = json.load(f)
            self._vix_data = meta.get('vix_data')
            self._spy_data = meta.get('spy_data')
            self.portfolio = meta.get('portfolio', [])
            ts_str = meta.get('timestamp', '')
            self._last_analysis_time = datetime.fromisoformat(ts_str) if ts_str else None
            self._populate_all_tables(self.df_results)
            self.btn_export.set_enabled(True)
            if self._last_analysis_time:
                dt_str = self._last_analysis_time.strftime('%d/%m %H:%M')
                self.hdr_status.config(text=f'● DATOS PREVIOS ({dt_str})', fg=COLORS['warning'])
            self._update_last_update_label()
            self._log('ok', f'Datos previos cargados ({ts_str[:16] if ts_str else "?"}). '
                             'Pulsa "Ejecutar" para actualizar.')
        except Exception as e:
            self._log('warn', f'No se pudieron cargar datos previos: {e}')

    def _update_last_update_label(self):
        """Actualiza el label de última actualización en el header (verde/amarillo/naranja)."""
        if self._last_analysis_time is None:
            self.hdr_last_update.config(text='')
            return
        now = datetime.now()
        diff = (now - self._last_analysis_time).total_seconds()
        dt_str = self._last_analysis_time.strftime('%d/%m/%Y %H:%M')
        if diff < 3600:
            color = COLORS['success']
        elif diff < 86400:
            color = COLORS['warning']
        else:
            color = COLORS['danger']
        self.hdr_last_update.config(text=f'Última actualización: {dt_str}', fg=color)

    # ─── TAB 1: DASHBOARD ────────────────────

    def _build_dashboard(self):
        p = self.tab_dashboard

        # Top: botones de acción
        action_bar = tk.Frame(p, bg=COLORS['bg'])
        action_bar.pack(fill=tk.X, padx=24, pady=(20, 0))

        self.btn_run = StyledButton(action_bar, '▶  Ejecutar Análisis Completo',
                                    self._start_analysis, color=COLORS['success'])
        self.btn_run.pack(side=tk.LEFT, padx=(0, 10))

        self.btn_stop = StyledButton(action_bar, '■  Detener',
                                     self._stop_analysis, color=COLORS['surface3'])
        self.btn_stop.pack(side=tk.LEFT, padx=(0, 10))

        self.btn_export = StyledButton(action_bar, '↓  Exportar Reporte',
                                       self._export_report, color=COLORS['primary'])
        self.btn_export.pack(side=tk.LEFT)
        self.btn_export.set_enabled(False)

        # Progress
        prog_f = tk.Frame(p, bg=COLORS['bg'])
        prog_f.pack(fill=tk.X, padx=24, pady=(14, 0))

        self.progress_label = tk.Label(prog_f, text='Listo para analizar',
                                        bg=COLORS['bg'], fg=COLORS['text_dim'],
                                        font=('Segoe UI', 9))
        self.progress_label.pack(anchor='w')

        self.progress_bar = ttk.Progressbar(prog_f, mode='determinate', length=400)
        self.progress_bar.pack(fill=tk.X, pady=(4, 0))

        # Métricas resumen
        metrics_f = tk.Frame(p, bg=COLORS['bg'])
        metrics_f.pack(fill=tk.X, padx=24, pady=16)
        for i in range(5):
            metrics_f.columnconfigure(i, weight=1)

        for i in range(6):
            metrics_f.columnconfigure(i, weight=1)

        self.m_total    = MetricCard(metrics_f, 'ACCIONES',        '—')
        self.m_buy      = MetricCard(metrics_f, 'COMPRAR (Vol MA)', '—', COLORS['buy'])
        self.m_sell     = MetricCard(metrics_f, 'VENDER (Vol MA)',  '—', COLORS['sell'])
        self.m_topscore = MetricCard(metrics_f, 'MEJOR SCORE',      '—', COLORS['primary'])
        self.m_confident= MetricCard(metrics_f, 'ALTA CONFIANZA',  '—', COLORS['accent'])
        self.m_vix      = MetricCard(metrics_f, 'VIX (MERCADO)',   '—')
        self.m_target   = MetricCard(metrics_f, 'RET.ANUAL TOP10', '—')

        metrics_f.columnconfigure(6, weight=1)
        for i, card in enumerate([self.m_total, self.m_buy, self.m_sell,
                                   self.m_topscore, self.m_confident, self.m_vix, self.m_target]):
            card.grid(row=0, column=i, sticky='nsew', padx=4)

        # Log de consola
        log_f = tk.Frame(p, bg=COLORS['bg'])
        log_f.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 16))

        tk.Label(log_f, text='LOG DEL PROCESO', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9, 'bold')).pack(anchor='w', pady=(0, 6))

        self.log_box = scrolledtext.ScrolledText(
            log_f, bg=COLORS['surface'], fg=COLORS['text_dim'],
            font=('Consolas', 9), relief='flat', bd=0,
            insertbackground=COLORS['text'], state='disabled',
            wrap=tk.WORD,
        )
        self.log_box.pack(fill=tk.BOTH, expand=True)

        # Tags de color para log
        self.log_box.tag_config('ok',   foreground=COLORS['success'])
        self.log_box.tag_config('err',  foreground=COLORS['danger'])
        self.log_box.tag_config('warn', foreground=COLORS['warning'])
        self.log_box.tag_config('info', foreground=COLORS['primary'])
        self.log_box.tag_config('dim',  foreground=COLORS['text_dim'])

    # ─── TAB 2: TOP COMPRAS (3×10) ───────────

    def _build_topbuys(self):
        p = self.tab_topbuys

        # ── Cabecera global ──
        top = tk.Frame(p, bg=COLORS['bg'])
        top.pack(fill=tk.X, padx=24, pady=(16, 0))
        tk.Label(top, text='Top 30 — 10 por cada método',
                 bg=COLORS['bg'], fg=COLORS['text'],
                 font=('Segoe UI', 13, 'bold')).pack(side=tk.LEFT)

        # ── Controles globales ──
        ctrl = tk.Frame(p, bg=COLORS['bg'])
        ctrl.pack(fill=tk.X, padx=24, pady=(8, 6))

        self.topbuy_etfs = tk.BooleanVar(value=True)
        tk.Label(ctrl, text='Incluir ETFs:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)
        tk.Checkbutton(ctrl, variable=self.topbuy_etfs,
                       bg=COLORS['bg'], fg=COLORS['text'], selectcolor=COLORS['surface2'],
                       activebackground=COLORS['bg'],
                       command=self._populate_topbuys).pack(side=tk.LEFT, padx=(4, 20))

        self.topbuy_only_buy = tk.BooleanVar(value=True)
        tk.Label(ctrl, text='Solo Vol COMPRAR (tablas 1 y 2):', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)
        tk.Checkbutton(ctrl, variable=self.topbuy_only_buy,
                       bg=COLORS['bg'], fg=COLORS['text'], selectcolor=COLORS['surface2'],
                       activebackground=COLORS['bg'],
                       command=self._populate_topbuys).pack(side=tk.LEFT, padx=(4, 20))

        # Filtro por sector ──────────────────────────────────────────────
        tk.Label(ctrl, text='Sector:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)
        self.topbuy_sector = tk.StringVar(value='Todos')
        self._sector_combo = ttk.Combobox(
            ctrl, textvariable=self.topbuy_sector,
            state='readonly', width=18, font=('Segoe UI', 9)
        )
        # Sectores se poblarán cuando lleguen los datos; por ahora valores fijos
        self._sector_combo['values'] = (
            'Todos', 'Technology', 'Financials', 'Healthcare',
            'Consumer Discretionary', 'Consumer Staples',
            'Energy', 'Communications', 'Industrials',
            'Materials', 'Real Estate', 'Utilities',
            'ETF — Indices', 'ETF — Sectorial',
            'ETF — Tematico', 'ETF — Commodities', 'ETF — Bonos',
        )
        self._sector_combo.pack(side=tk.LEFT, padx=(4, 16))
        self._sector_combo.bind('<<ComboboxSelected>>', lambda e: self._populate_topbuys())

        tk.Button(ctrl, text='📅 Cargar Earnings',
                  command=self._fetch_earnings_async,
                  bg=COLORS['surface2'], fg=COLORS['text'],
                  font=('Segoe UI', 9), relief='flat', padx=8, pady=3,
                  cursor='hand2').pack(side=tk.LEFT, padx=(0, 12))

        tk.Button(ctrl,
                  text='🎯 Ver Precio Objetivo',
                  bg=COLORS['success'], fg='white',
                  font=('Segoe UI', 9, 'bold'), relief='flat', cursor='hand2',
                  command=self._show_target_any_tree).pack(side=tk.LEFT)
        tk.Label(ctrl, text='  (selecciona una fila de cualquier tabla)',
                 bg=COLORS['bg'], fg=COLORS['text_dim'],
                 font=('Segoe UI', 9)).pack(side=tk.LEFT)

        tk.Button(ctrl,
                  text='💰 Distribuir Capital',
                  bg='#4a3b00', fg='#f0c04a',
                  font=('Segoe UI', 9, 'bold'), relief='flat', cursor='hand2',
                  command=self._show_capital_distribution).pack(side=tk.RIGHT, padx=(12, 0))

        # ── PanedWindow vertical para las 3 tablas redimensionables ──
        pane = ttk.PanedWindow(p, orient=tk.VERTICAL)
        pane.pack(fill=tk.BOTH, expand=True, padx=24, pady=(4, 0))

        # ─── Sección 1: Score Compuesto ─────────────────────────────────
        f1 = tk.Frame(pane, bg=COLORS['bg'])
        pane.add(f1, weight=1)

        hdr1 = tk.Frame(f1, bg='#0d2a1a', padx=12, pady=5)
        hdr1.pack(fill=tk.X)
        tk.Label(hdr1, text='🟢  TOP 10 — Score Compuesto',
                 bg='#0d2a1a', fg=COLORS['success'],
                 font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT)
        self._hdr1_sector_lbl = tk.Label(hdr1, text='',
                 bg='#0d2a1a', fg=COLORS['warning'],
                 font=('Segoe UI', 9, 'bold'))
        self._hdr1_sector_lbl.pack(side=tk.LEFT, padx=(8, 0))
        tk.Label(hdr1,
                 text='  Combina: Vol MA · Confianza técnica · RS vs SPY · Riesgo · Backtesting · Fundamentales',
                 bg='#0d2a1a', fg='#4a8f5a',
                 font=('Segoe UI', 8)).pack(side=tk.LEFT)

        cols1 = ('r', 'dec', 'sym', 'sect', 'score', 'ann_ret', 'rs252', 'price', 'stop', 'vol', 'reasons')
        hdrs1 = ('#', 'Decisión', 'Símbolo', 'Sector', 'Score', 'Ret.Anual', 'RS 252d',
                 'Precio', 'Stop', 'Vol MA', 'Por qué')
        self.topbuy_tree, sf1 = self._make_treeview(
            f1, cols1, hdrs1,
            col_widths=[30, 90, 68, 130, 60, 72, 72, 68, 68, 78, 310],
            height=10)
        sf1.pack(fill=tk.BOTH, expand=True)
        self.topbuy_tree.bind('<Double-1>',
                              lambda e: self._detail_from_tree(self.topbuy_tree, 2))

        # ─── Sección 2: Caída Atractiva ──────────────────────────────────
        f2 = tk.Frame(pane, bg=COLORS['bg'])
        pane.add(f2, weight=1)

        hdr2 = tk.Frame(f2, bg='#1a1a00', padx=12, pady=5)
        hdr2.pack(fill=tk.X)
        tk.Label(hdr2, text='📉  TOP 10 — Caída Atractiva',
                 bg='#1a1a00', fg=COLORS['warning'],
                 font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT)
        tk.Label(hdr2,
                 text='  Calidad alta · Tendencia positiva · Caída reciente inusual → entrada más barata',
                 bg='#1a1a00', fg='#8a7a00',
                 font=('Segoe UI', 8)).pack(side=tk.LEFT)

        cols2 = ('r', 'dec', 'sym', 'sect', 'dsc', 'c5', 'c20', 'rsi', 'inusual', 'price', 'stop')
        hdrs2 = ('#', 'Decisión', 'Símbolo', 'Sector', 'DipScore',
                 'Caída 5d', 'Tend.20d', 'RSI', 'Inusual×', 'Precio', 'Stop')
        self.topbuy_tree2, sf2 = self._make_treeview(
            f2, cols2, hdrs2,
            col_widths=[30, 100, 68, 130, 68, 72, 72, 52, 70, 68, 68],
            height=10)
        sf2.pack(fill=tk.BOTH, expand=True)
        self.topbuy_tree2.bind('<Double-1>',
                               lambda e: self._detail_from_tree(self.topbuy_tree2, 2))

        # ─── Sección 3: Predicción ───────────────────────────────────────
        f3 = tk.Frame(pane, bg=COLORS['bg'])
        pane.add(f3, weight=1)

        hdr3 = tk.Frame(f3, bg='#0d1a2d', padx=12, pady=5)
        hdr3.pack(fill=tk.X)
        tk.Label(hdr3, text='🔮  TOP 10 — Predicción 10 días',
                 bg='#0d1a2d', fg=COLORS['primary'],
                 font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT)
        tk.Label(hdr3,
                 text='  Regresión de tendencia + reconocimiento de patrones históricos similares',
                 bg='#0d1a2d', fg='#2a4a7a',
                 font=('Segoe UI', 8)).pack(side=tk.LEFT)

        cols3 = ('sym', 'sect', 'sig', 'trend', 'r2', 'wr', 'avg', 'npat', 'price', 'proj')
        hdrs3 = ('Símbolo', 'Sector', 'Señal', 'Tendencia 10d', 'R²',
                 'WR Patrones', 'Ret.Medio', 'N Pat.', 'Precio', 'Precio ~10d')
        self.topbuy_tree3, sf3 = self._make_treeview(
            f3, cols3, hdrs3,
            col_widths=[68, 130, 100, 100, 52, 88, 80, 55, 68, 100],
            height=10)
        sf3.pack(fill=tk.BOTH, expand=True)
        self.topbuy_tree3.bind('<Double-1>',
                               lambda e: self._detail_from_tree(self.topbuy_tree3, 0))

        # ─── Sección 4: Big Dip / Turnaround ────────────────────────────
        f4 = tk.Frame(pane, bg=COLORS['bg'])
        pane.add(f4, weight=1)

        hdr4 = tk.Frame(f4, bg='#2a0d1a', padx=12, pady=5)
        hdr4.pack(fill=tk.X)
        tk.Label(hdr4, text='💎  TOP 10 — Big Dip · Turnaround',
                 bg='#2a0d1a', fg='#e87aaa',
                 font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT)
        tk.Label(hdr4,
                 text='  Hundidas -22%+ desde máx 52w · RSI saliendo de sobrevendido · últimos días en verde · dinero volviendo',
                 bg='#2a0d1a', fg='#7a3a5a',
                 font=('Segoe UI', 8)).pack(side=tk.LEFT)

        cols4 = ('r', 'dec', 'sym', 'sect', 'bds',
                 'desde_max', 'c5', 'c20', 'rsi', 'vol', 'pred', 'price')
        hdrs4 = ('#', 'Señal', 'Símbolo', 'Sector', 'BigDipScore',
                 'Bajo Máx 52w', 'Cambio 5d', 'Cambio 20d', 'RSI', 'Vol MA', 'Predicción', 'Precio')
        self.topbuy_tree4, sf4 = self._make_treeview(
            f4, cols4, hdrs4,
            col_widths=[30, 110, 68, 130, 88,
                        90, 72, 72, 52, 78, 100, 68],
            height=10)
        sf4.pack(fill=tk.BOTH, expand=True)
        self.topbuy_tree4.bind('<Double-1>',
                               lambda e: self._detail_from_tree(self.topbuy_tree4, 2))

        # ── Resumen ──
        self.topbuy_summary = tk.Label(p, text='',
                                        bg=COLORS['surface'], fg=COLORS['text_dim'],
                                        font=('Segoe UI', 9), anchor='w',
                                        padx=16, pady=6)
        self.topbuy_summary.pack(fill=tk.X, padx=24, pady=(4, 8))

    # Bond ETFs — scoring propio, no de equity
    _BOND_ETFS = {'TLT', 'AGG', 'HYG', 'TIP'}

    def _compute_buy_score(self, row):
        return compute_buy_score(row)

    def _compute_bond_score(self, row):
        return compute_bond_score(row)

    def _build_buy_reasons(self, row):
        """Construye string legible con los motivos de la recomendación."""
        parts = []

        # Vol MA
        vol = row.get('vol_signal', '')
        ratio = row.get('vol_ratio')
        if vol == 'COMPRAR' and ratio is not None:
            parts.append(f"Vol↑ {ratio:.2f}x")
        elif vol == 'VENDER':
            parts.append("Vol↓")

        # Multi-TF
        tfa = row.get('tf_alignment', '')
        if tfa and tfa != 'N/A':
            if 'ALCISTA' in tfa:
                pts = row.get('tf_conf_pts', 0) or 0
                parts.append(f"TF+ ({pts:+d}pts)")
            elif 'BAJISTA' in tfa:
                parts.append("TF-")

        # RS vs SPY — prioriza rs_252d (mediano plazo)
        rs = row.get('rs_signal', '')
        rs_252 = row.get('rs_252d')
        rs_20  = row.get('rs_20d')
        if rs_252 is not None and rs == 'LIDER':
            parts.append(f"LIDER SPY +{rs_252:.1f}%/año")
        elif rs_252 is not None and rs == 'REZAGADO':
            parts.append(f"Rezagado SPY {rs_252:.1f}%/año")
        elif rs == 'LIDER' and rs_20 is not None:
            parts.append(f"LIDER SPY +{rs_20:.1f}%")

        # Retorno anualizado vs objetivo 15%
        ann_ret = row.get('ann_return')
        if ann_ret is not None and ann_ret > 15:
            parts.append(f"RetAnual {ann_ret:.0f}%")

        # Riesgo / Sharpe
        sharpe = row.get('sharpe')
        risk = row.get('risk_signal', '')
        if sharpe is not None:
            parts.append(f"Sharpe {sharpe:.2f}")
        if 'BAJO' in risk:
            parts.append("Riesgo Bajo")
        elif 'ALTO' in risk:
            parts.append("Riesgo Alto")

        # Backtesting
        wr = row.get('bt_win_rate_vol')
        if wr is not None:
            parts.append(f"WR {wr*100:.0f}%")

        # Fundamentales
        fund = row.get('fund_signal', '')
        if fund == 'FAVORABLE':
            parts.append("Fund OK")
        elif fund == 'DESFAVORABLE':
            parts.append("Fund Mal")

        return "  |  ".join(parts) if parts else "—"

    def _fetch_earnings_async(self):
        """Descarga fechas de earnings en background para los símbolos visibles."""
        if self.df_results is None or self.df_results.empty:
            messagebox.showinfo('Sin datos', 'Ejecuta el análisis primero.')
            return

        symbols = self.df_results['symbol'].tolist()
        self._log('info', f'Cargando earnings para {len(symbols)} símbolos...')

        def _fetch():
            import yfinance as yf
            from datetime import datetime, timedelta
            cache = {}
            today = datetime.now().date()
            warning_days = 14
            for sym in symbols:
                try:
                    cal = yf.Ticker(sym).calendar
                    if cal is not None and not cal.empty:
                        # calendar es un DataFrame con fechas como índice
                        dates = [pd.to_datetime(v).date()
                                 for v in cal.values.flatten()
                                 if pd.notna(v)]
                        future = [d for d in dates if d >= today]
                        if future:
                            next_e = min(future)
                            days_away = (next_e - today).days
                            label = next_e.strftime('%d/%m')
                            if days_away <= warning_days:
                                cache[sym] = f'⚠️ {label} ({days_away}d)'
                            else:
                                cache[sym] = label
                        else:
                            cache[sym] = '—'
                    else:
                        cache[sym] = '—'
                except Exception:
                    cache[sym] = '—'

            self._earnings_cache.update(cache)
            self.q.put(('earnings_done', len([v for v in cache.values() if v != '—'])))

        import threading
        threading.Thread(target=_fetch, daemon=True).start()

    def _show_target_any_tree(self):
        """Busca la selección en cualquiera de las 4 tablas y muestra precio objetivo."""
        trees = [
            (self.topbuy_tree,  2),   # sym_col
            (self.topbuy_tree2, 2),
            (self.topbuy_tree3, 0),
            (self.topbuy_tree4, 2),
        ]
        for tree, sym_col in trees:
            sel = tree.selection()
            if sel:
                sym = str(tree.item(sel[0], 'values')[sym_col]).strip()
                if self.df_results is not None:
                    rows = self.df_results[self.df_results['symbol'] == sym]
                    if not rows.empty:
                        self._show_price_target_popup(rows.iloc[0].to_dict())
                return
        messagebox.showinfo('Selección', 'Selecciona una fila en cualquiera de las tablas.')

    def _populate_topbuys(self):
        if self.df_results is None or self.df_results.empty:
            return

        df_all = self.df_results.copy()

        # Filtro ETFs común
        if not self.topbuy_etfs.get():
            try:
                from stock_universe import is_etf
                df_all = df_all[~df_all['symbol'].apply(is_etf)]
            except Exception:
                pass

        # Filtro por sector (nuevo) — aplica a TODAS las tablas
        sector_sel = getattr(self, 'topbuy_sector', None)
        sector_sel = sector_sel.get() if sector_sel else 'Todos'
        if sector_sel and sector_sel != 'Todos' and 'sector' in df_all.columns:
            df_all = df_all[df_all['sector'] == sector_sel]
        else:
            # Sin filtro de sector: excluir bond ETFs del ranking general.
            # Son instrumentos de preservación de capital/renta, no de crecimiento.
            # El scoring de equity (RS vs SPY, ret >15%) no aplica a bonos.
            # Para verlos: seleccionar "ETF — Bonos" en el filtro de sector.
            _BOND_ETFS = {'TLT', 'AGG', 'HYG', 'TIP'}
            if 'symbol' in df_all.columns:
                df_all = df_all[~df_all['symbol'].isin(_BOND_ETFS)]

        # ─── Tabla 1: Score Compuesto ────────────────────────────────────
        df1 = df_all.copy()
        if self.topbuy_only_buy.get():
            df1 = df1[df1['vol_signal'] == 'COMPRAR']
        df1['_buy_score'] = df1.apply(self._compute_buy_score, axis=1)
        df1 = df1.sort_values('_buy_score', ascending=False, na_position='last').head(10)

        t1 = self.topbuy_tree
        t1.delete(*t1.get_children())
        for i, (_, row) in enumerate(df1.iterrows(), 1):
            score = row['_buy_score']
            price = row.get('price')
            atr_pct = row.get('atr_pct')
            ann_ret = row.get('ann_return')
            rs_252  = row.get('rs_252d')
            conf  = row.get('confidence', '—')
            high_conf = conf in ('MUY ALTA', 'ALTA')
            if score >= 70 and high_conf: dec = '🟢 INVERTIR'
            elif score >= 70:             dec = '🟢 INVERTIR*'
            elif score >= 50:             dec = '🟡 VIGILAR'
            else:                         dec = '🔴 EVITAR'
            stop_str = '—'
            if price and atr_pct:
                stop_str = f"${price - 2 * price * atr_pct / 100:.2f}"
            tag = 'buy' if score >= 70 else 'neutral' if score >= 50 else 'sell'
            t1.insert('', 'end', tags=(tag,), values=(
                i, dec,
                row.get('symbol', ''), row.get('sector', ''),
                f"{score:.1f}",
                f"{ann_ret:+.1f}%" if ann_ret is not None else '—',
                f"{rs_252:+.1f}%" if rs_252 is not None else '—',
                f"${price:.2f}" if price else '—',
                stop_str,
                row.get('vol_signal', '—'),
                self._build_buy_reasons(row),
            ))

        # ─── Tabla 2: Caída Atractiva ────────────────────────────────────
        df2 = df_all.copy()
        if self.topbuy_only_buy.get():
            df2 = df2[df2['vol_signal'] == 'COMPRAR']
        df2 = df2[df2['change_20d'].fillna(-999) > 0]   # uptrend
        df2['_buy_score'] = df2.apply(self._compute_buy_score, axis=1)
        df2 = df2[df2['_buy_score'] >= 45]
        df2['_dip_score'] = df2.apply(self._compute_dip_score, axis=1)
        df2 = df2.sort_values('_dip_score', ascending=False, na_position='last').head(10)

        t2 = self.topbuy_tree2
        t2.delete(*t2.get_children())
        for i, (_, row) in enumerate(df2.iterrows(), 1):
            dsc   = row['_dip_score']
            price = row.get('price')
            atr_pct = row.get('atr_pct')
            c5    = row.get('change_5d')
            c20   = row.get('change_20d')
            rsi   = row.get('rsi')
            inusual = row.get('dip_atr_multiple')
            if dsc >= 65:   dec = '🟢 COMPRAR DIP'
            elif dsc >= 50: dec = '🟡 VIGILAR'
            else:           dec = '🔴 ESPERAR'
            stop_str = '—'
            if price and atr_pct:
                stop_str = f"${price - 2 * price * atr_pct / 100:.2f}"
            tag = 'buy' if dsc >= 65 else 'neutral' if dsc >= 50 else 'sell'
            t2.insert('', 'end', tags=(tag,), values=(
                i, dec,
                row.get('symbol', ''), row.get('sector', ''),
                f"{dsc:.1f}",
                f"{c5:+.1f}%"  if c5  is not None else '—',
                f"{c20:+.1f}%" if c20 is not None else '—',
                f"{rsi:.0f}"   if rsi is not None else '—',
                f"{inusual:.1f}×" if inusual is not None else '—',
                f"${price:.2f}" if price else '—',
                stop_str,
            ))

        # ─── Tabla 3: Predicción ─────────────────────────────────────────
        df3 = df_all.copy()
        # Ordenar: primero señales SUBE FUERTE/SUBE, luego por WR patrones
        sig_order = {'SUBE FUERTE': 0, 'SUBE': 1, 'NEUTRO': 2, 'BAJA': 3, 'BAJA FUERTE': 4}
        df3['_sig_ord'] = df3['pred_signal'].map(sig_order).fillna(2)
        df3 = df3.sort_values(['_sig_ord', 'pred_pat_wr'],
                              ascending=[True, False], na_position='last').head(10)

        t3 = self.topbuy_tree3
        t3.delete(*t3.get_children())
        for _, row in df3.iterrows():
            price   = row.get('price')
            trend10 = row.get('pred_trend_10d')
            r2      = row.get('pred_r2')
            pat_wr  = row.get('pred_pat_wr')
            pat_avg = row.get('pred_pat_avg')
            pat_n   = row.get('pred_pat_n', 0)
            signal  = row.get('pred_signal', 'NEUTRO')
            proj_str = '—'
            if price and trend10 is not None:
                proj_str = f"${price*(1+trend10/100):.2f} ({trend10:+.1f}%)"
            tag = ('buy'  if signal in ('SUBE', 'SUBE FUERTE') else
                   'sell' if signal in ('BAJA', 'BAJA FUERTE') else 'neutral')
            t3.insert('', 'end', tags=(tag,), values=(
                row.get('symbol', ''),
                row.get('sector', ''),
                signal,
                f"{trend10:+.1f}%" if trend10 is not None else '—',
                f"{r2:.2f}"        if r2       is not None else '—',
                f"{pat_wr:.0f}%"   if pat_wr   is not None else '—',
                f"{pat_avg:+.2f}%" if pat_avg  is not None else '—',
                str(int(pat_n))    if pat_n    else '—',
                f"${price:.2f}"    if price    else '—',
                proj_str,
            ))

        # ─── Tabla 4: Big Dip / Turnaround ──────────────────────────────
        df4 = df_all.copy()
        # Calcular score y filtrar: solo los que realmente están hundidos (retorna 0 si no cumple)
        df4['_bigdip_score'] = df4.apply(self._compute_bigdip_score, axis=1)
        df4 = df4[df4['_bigdip_score'] > 0]
        df4 = df4.sort_values('_bigdip_score', ascending=False, na_position='last').head(10)

        t4 = self.topbuy_tree4
        t4.delete(*t4.get_children())
        for i, (_, row) in enumerate(df4.iterrows(), 1):
            bds     = row['_bigdip_score']
            price   = row.get('price')
            res_52w = row.get('resistance_52w') or 0
            c5      = row.get('change_5d')
            c20     = row.get('change_20d')
            rsi     = row.get('rsi')
            vol     = row.get('vol_signal', '—')
            pred    = row.get('pred_signal', '—')

            desde_max = ((res_52w - price) / res_52w * 100
                         if price and res_52w > price else None)

            if   bds >= 65: dec = '🚀 TURNAROUND'
            elif bds >= 45: dec = '👀 VIGILAR'
            else:           dec = '⏳ AÚN PRONTO'

            tag = 'buy' if bds >= 65 else 'neutral' if bds >= 45 else 'sell'
            t4.insert('', 'end', tags=(tag,), values=(
                i, dec,
                row.get('symbol', ''),
                row.get('sector', ''),
                f"{bds:.1f}",
                f"−{desde_max:.1f}%" if desde_max is not None else '—',
                f"{c5:+.1f}%"  if c5  is not None else '—',
                f"{c20:+.1f}%" if c20 is not None else '—',
                f"{rsi:.0f}"   if rsi is not None else '—',
                vol,
                pred,
                f"${price:.2f}" if price else '—',
            ))

        # ── Resumen combinado accionable (los 4 métodos) ──────────────────
        top_score   = df1['symbol'].head(3).tolist()  if not df1.empty else []
        top_dip     = df2['symbol'].head(3).tolist()  if not df2.empty else []
        top_pred    = (df3[df3['pred_signal'].isin(['SUBE', 'SUBE FUERTE'])]
                       ['symbol'].head(3).tolist())
        top_bigdip  = (df4[df4['_bigdip_score'] >= 65]['symbol'].head(3).tolist()
                       if not df4.empty else [])

        from collections import Counter
        all_syms  = top_score + top_dip + top_pred + top_bigdip
        counts    = Counter(all_syms)
        confirmed = [s for s, c in counts.most_common() if c >= 2]

        if confirmed:
            txt = f"⭐ En ≥2 métodos: {', '.join(confirmed[:5])}   |   "
        else:
            txt = ''
        parts = []
        if top_score:  parts.append(f"🟢 {', '.join(top_score[:2])}")
        if top_dip:    parts.append(f"📉 {', '.join(top_dip[:2])}")
        if top_pred:   parts.append(f"🔮 {', '.join(top_pred[:2])}")
        if top_bigdip: parts.append(f"💎 {', '.join(top_bigdip[:2])}")
        self.topbuy_summary.config(text=f"  {txt}{'   '.join(parts)}")

        # Actualizar label de sector activo en cabecera tabla 1
        if hasattr(self, '_hdr1_sector_lbl'):
            if sector_sel and sector_sel != 'Todos':
                n_total = len(self.df_results[self.df_results['sector'] == sector_sel])
                self._hdr1_sector_lbl.config(
                    text=f'[Sector: {sector_sel}  —  {n_total} acciones]'
                )
            else:
                self._hdr1_sector_lbl.config(text='')

    # ─── TAB 3: CAÍDA ATRACTIVA (BUY THE DIP) ──

    # ─────────────────────────────────────────────
    #  MEJORA 1: DISTRIBUCIÓN DE CAPITAL (Kelly%)
    # ─────────────────────────────────────────────

    def _show_capital_distribution(self):
        """
        Popup donde el usuario ingresa un monto total (€) y recibe
        una distribución óptima entre las mejores acciones usando Kelly%.
        Si no hay Kelly% disponible, usa el Score Compuesto normalizado.
        """
        if self.df_results is None or self.df_results.empty:
            messagebox.showinfo('Sin datos', 'Ejecuta el análisis primero.')
            return

        # ── Ventana de entrada ──────────────────────────────────────
        ask_win = tk.Toplevel(self.root)
        ask_win.title('💰 Distribución de Capital')
        ask_win.geometry('480x220')
        ask_win.configure(bg=COLORS['bg'])
        ask_win.resizable(False, False)
        ask_win.grab_set()

        tk.Label(ask_win, text='💰  Distribución de Capital',
                 bg=COLORS['bg'], fg=COLORS['primary'],
                 font=('Segoe UI', 13, 'bold')).pack(padx=24, pady=(20, 4), anchor='w')
        tk.Label(ask_win,
                 text='Calcula cómo repartir tu capital entre las mejores oportunidades\n'
                      'usando el % de Kelly de cada acción (riesgo/retorno óptimo).',
                 bg=COLORS['bg'], fg=COLORS['text_dim'],
                 font=('Segoe UI', 9), justify='left').pack(padx=24, anchor='w')

        form = tk.Frame(ask_win, bg=COLORS['bg'])
        form.pack(padx=24, pady=14)

        tk.Label(form, text='Capital total (€):', bg=COLORS['bg'],
                 fg=COLORS['text'], font=('Segoe UI', 10)).grid(row=0, column=0, sticky='w')
        capital_var = tk.StringVar(value='3000')
        tk.Entry(form, textvariable=capital_var, width=12,
                 bg=COLORS['surface2'], fg=COLORS['text'],
                 insertbackground=COLORS['text'],
                 relief='flat', bd=4, font=('Segoe UI', 10)).grid(row=0, column=1, padx=(8, 20))

        tk.Label(form, text='Top N acciones:', bg=COLORS['bg'],
                 fg=COLORS['text'], font=('Segoe UI', 10)).grid(row=0, column=2, sticky='w')
        topn_var = tk.StringVar(value='5')
        ttk.Combobox(form, textvariable=topn_var, width=4, state='readonly',
                     values=['3', '5', '7', '10'],
                     font=('Segoe UI', 10)).grid(row=0, column=3, padx=(8, 0))

        only_buy_var = tk.BooleanVar(value=True)
        tk.Checkbutton(ask_win, text='Solo Vol MA = COMPRAR',
                       variable=only_buy_var,
                       bg=COLORS['bg'], fg=COLORS['text'],
                       selectcolor=COLORS['surface2'],
                       activebackground=COLORS['bg'],
                       font=('Segoe UI', 9)).pack(padx=24, anchor='w')

        def compute():
            try:
                capital = float(capital_var.get().replace(',', '.').replace('€', ''))
                top_n   = int(topn_var.get())
            except ValueError:
                messagebox.showerror('Error', 'Capital inválido.')
                return

            df = self.df_results.copy()
            if only_buy_var.get():
                df = df[df['vol_signal'] == 'COMPRAR']
            df['_buy_score'] = df.apply(self._compute_buy_score, axis=1)
            df = df.sort_values('_buy_score', ascending=False).head(top_n)

            if df.empty:
                messagebox.showinfo('Sin candidatos',
                                    'No hay acciones que cumplan los filtros.')
                return
            ask_win.destroy()
            self._show_capital_result(df, capital)

        tk.Button(ask_win, text='Calcular distribución →',
                  command=compute,
                  bg=COLORS['warning'], fg='black',
                  font=('Segoe UI', 10, 'bold'), relief='flat',
                  padx=16, pady=6, cursor='hand2').pack(pady=(0, 16))

    def _show_capital_result(self, df, capital):
        """Muestra la tabla de distribución de capital."""
        # ── Calcular pesos ──────────────────────────────────────────
        # Preferir Kelly% si existe, si no usar buy_score normalizado
        weights = []
        for _, row in df.iterrows():
            kelly = row.get('kelly_pct')
            if kelly and kelly > 0:
                w = min(kelly, 25)   # Limitar Kelly al 25% por posición
            else:
                w = max(row.get('_buy_score', 10), 1)
            weights.append(w)

        total_w = sum(weights)
        amounts  = [capital * w / total_w for w in weights]

        # ── Ventana resultado ───────────────────────────────────────
        win = tk.Toplevel(self.root)
        win.title(f'💰 Distribución de €{capital:,.0f}')
        win.geometry('760x420')
        win.configure(bg=COLORS['bg'])
        win.resizable(True, True)

        # Cabecera
        hdr = tk.Frame(win, bg=COLORS['surface'], padx=20, pady=12)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text=f'💰  Distribución óptima de €{capital:,.0f}',
                 bg=COLORS['surface'], fg=COLORS['primary'],
                 font=('Segoe UI', 12, 'bold')).pack(anchor='w')
        tk.Label(hdr,
                 text='Pesos basados en % Kelly (riesgo/retorno óptimo). '
                      'Ajusta según tu perfil de riesgo. No es asesoramiento financiero.',
                 bg=COLORS['surface'], fg=COLORS['text_dim'],
                 font=('Segoe UI', 8), justify='left').pack(anchor='w', pady=(2, 0))

        # Tabla
        cols = ('rank', 'sym', 'sect', 'score', 'kelly', 'pct', 'amount', 'shares', 'price', 'stop')
        hdrs = ('#', 'Símbolo', 'Sector', 'Score', 'Kelly%', '% Cartera', 'Monto €', 'Acciones ~', 'Precio', 'Stop Loss')
        tree, sf = self._make_treeview(
            win, cols, hdrs,
            col_widths=[30, 68, 140, 58, 58, 72, 80, 72, 68, 68])
        sf.pack(fill=tk.BOTH, expand=True, padx=16, pady=(12, 4))

        for i, ((_, row), amt, w) in enumerate(zip(df.iterrows(), amounts, weights), 1):
            price   = row.get('price')
            atr_pct = row.get('atr_pct')
            kelly   = row.get('kelly_pct')
            pct_w   = w / total_w * 100
            shares  = amt / price if price else 0
            stop    = price - 2 * price * atr_pct / 100 if price and atr_pct else None
            score   = row.get('_buy_score', 0)

            tag = 'buy' if score >= 70 else 'neutral'
            tree.insert('', 'end', tags=(tag,), values=(
                i,
                row.get('symbol', ''),
                row.get('sector', ''),
                f"{score:.0f}",
                f"{kelly:.1f}%" if kelly and kelly > 0 else "—",
                f"{pct_w:.1f}%",
                f"€{amt:,.0f}",
                f"{shares:.3f}",
                f"${price:.2f}" if price else '—',
                f"${stop:.2f}" if stop else '—',
            ))

        # Resumen accionable en texto
        summary_parts = []
        for i, ((_, row), amt) in enumerate(zip(df.iterrows(), amounts)):
            summary_parts.append(f"{row.get('symbol','')} €{amt:,.0f}")
        summary_txt = '  ·  '.join(summary_parts)

        tk.Label(win, text=f"  {summary_txt}",
                 bg=COLORS['surface2'], fg=COLORS['warning'],
                 font=('Segoe UI', 9, 'bold'),
                 anchor='w', padx=12, pady=7).pack(fill=tk.X, padx=16, pady=(0, 4))

        tk.Button(win, text='Cerrar', command=win.destroy,
                  bg=COLORS['surface3'], fg=COLORS['text'],
                  relief='flat', cursor='hand2',
                  font=('Segoe UI', 9), padx=12, pady=4).pack(pady=(0, 12))

    # ─── TAB PREDICCIÓN ──────────────────────

    def _build_prediction_tab(self):
        p = self.tab_pred

        # ── Título ──
        hdr = tk.Frame(p, bg=COLORS['bg'])
        hdr.pack(fill=tk.X, padx=24, pady=(20, 0))
        tk.Label(hdr, text='Predicción a 10 días — Análisis estadístico histórico',
                 bg=COLORS['bg'], fg=COLORS['text'],
                 font=('Segoe UI', 14, 'bold')).pack(side=tk.LEFT)

        # ── Disclaimer honesto ──
        disc = tk.Frame(p, bg='#2d1f00', padx=16, pady=10)
        disc.pack(fill=tk.X, padx=24, pady=(8, 0))
        tk.Label(disc,
                 text='⚠️  IMPORTANTE — Esto no es adivinación:',
                 bg='#2d1f00', fg=COLORS['warning'],
                 font=('Segoe UI', 9, 'bold')).pack(anchor='w')
        tk.Label(disc, bg='#2d1f00', fg='#c8a84b',
                 font=('Segoe UI', 9), justify='left',
                 text=(
                     'Este módulo usa dos métodos estadísticos sobre datos históricos reales:\n'
                     '  1. TENDENCIA (Regresión Lineal): si el precio lleva X días subiendo Y€/día, '
                     'proyecta que continuará ~10 días más. El R² indica qué tan limpia es esa tendencia (>0.6 = confiable).\n'
                     '  2. PATRONES SIMILARES: busca en todo el histórico momentos donde el precio se movió de forma parecida '
                     'a los últimos 10 días, y mide qué % de esas veces el precio subió 10 días después.\n\n'
                     'Úsalos como un argumento más, NO como certeza. Un 70% de win rate histórico significa que falló el 30% de las veces.'
                 )).pack(anchor='w', pady=(4, 0))

        # ── Controles ──
        ctrl = tk.Frame(p, bg=COLORS['bg'])
        ctrl.pack(fill=tk.X, padx=24, pady=(10, 6))
        tk.Label(ctrl, text='Ordenar por:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)
        self.pred_sort = ttk.Combobox(ctrl, width=22, state='readonly',
                                       font=('Segoe UI', 9),
                                       values=['Win Rate Patrones', 'Tendencia 10d',
                                               'Retorno Medio Esperado', 'R² Tendencia'])
        self.pred_sort.current(0)
        self.pred_sort.pack(side=tk.LEFT, padx=(6, 20))
        self.pred_sort.bind('<<ComboboxSelected>>', lambda e: self._populate_prediction_tab())

        tk.Label(ctrl, text='Solo señal SUBE:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)
        self.pred_only_up = tk.BooleanVar(value=False)
        tk.Checkbutton(ctrl, variable=self.pred_only_up,
                       bg=COLORS['bg'], fg=COLORS['text'], selectcolor=COLORS['surface2'],
                       activebackground=COLORS['bg'],
                       command=self._populate_prediction_tab).pack(side=tk.LEFT, padx=(4, 0))

        # ── Tabla ──
        cols = ('p_sym', 'p_sector', 'p_signal',
                'p_trend', 'p_r2',
                'p_pat_wr', 'p_pat_avg', 'p_pat_n',
                'p_zscore', 'p_price', 'p_proj',
                'p_interp')
        hdrs = ('Símbolo', 'Sector', 'Señal',
                'Tendencia 10d', 'R² Tend.',
                'WR Patrones', 'Ret. Medio Pat.', 'N Patrones',
                'Z-Score', 'Precio', 'Precio ~10d',
                'Interpretación')

        self.pred_tree, scroll_pred = self._make_treeview(
            p, cols, hdrs,
            col_widths=[70, 140, 100,
                        100, 72,
                        90, 110, 78,
                        72, 72, 88,
                        310])
        scroll_pred.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 8))
        self.pred_tree.bind('<Double-1>', lambda e: self._detail_from_tree(self.pred_tree, 0))

        # ── Resumen ──
        self.pred_summary = tk.Label(p, text='',
                                      bg=COLORS['surface'], fg=COLORS['text_dim'],
                                      font=('Segoe UI', 9), anchor='w',
                                      padx=16, pady=6)
        self.pred_summary.pack(fill=tk.X, padx=24, pady=(0, 12))

    def _populate_prediction_tab(self):
        if self.df_results is None or self.df_results.empty:
            return

        df = self.df_results.copy()

        # Filtro solo señal SUBE
        if self.pred_only_up.get():
            df = df[df['pred_signal'].isin(['SUBE', 'SUBE FUERTE'])]

        if df.empty:
            self.pred_tree.delete(*self.pred_tree.get_children())
            self.pred_summary.config(text='Sin candidatos con los filtros actuales.')
            return

        # Ordenar
        sort_col_map = {
            'Win Rate Patrones':       'pred_pat_wr',
            'Tendencia 10d':           'pred_trend_10d',
            'Retorno Medio Esperado':  'pred_pat_avg',
            'R² Tendencia':            'pred_r2',
        }
        sc = sort_col_map.get(self.pred_sort.get(), 'pred_pat_wr')
        if sc in df.columns:
            df = df.sort_values(sc, ascending=False, na_position='last')

        tree = self.pred_tree
        tree.delete(*tree.get_children())

        sig_color = {
            'SUBE FUERTE': 'buy',
            'SUBE':        'buy',
            'BAJA':        'sell',
            'BAJA FUERTE': 'sell',
            'NEUTRO':      'neutral',
        }

        for _, row in df.iterrows():
            sym       = row.get('symbol', '')
            price     = row.get('price')
            trend10   = row.get('pred_trend_10d')
            r2        = row.get('pred_r2')
            pat_wr    = row.get('pred_pat_wr')
            pat_avg   = row.get('pred_pat_avg')
            pat_n     = row.get('pred_pat_n', 0)
            zscore    = row.get('pred_zscore')
            signal    = row.get('pred_signal', 'NEUTRO')

            # Precio proyectado a 10d según tendencia
            if price and trend10 is not None:
                proj_price = price * (1 + trend10 / 100)
                proj_str   = f"${proj_price:.2f} ({trend10:+.1f}%)"
            else:
                proj_str   = '—'

            # Construcción de interpretación legible
            parts = []
            if trend10 is not None and r2 is not None:
                trend_dir  = '↑' if trend10 > 0 else '↓'
                r2_quality = 'tendencia clara' if r2 > 0.6 else ('tendencia moderada' if r2 > 0.35 else 'tendencia débil')
                parts.append(f"Tendencia {trend_dir}{abs(trend10):.1f}% / 10d ({r2_quality} R²={r2:.2f})")
            if pat_wr is not None and pat_n >= 3:
                direction = 'subió' if pat_wr >= 50 else 'bajó'
                parts.append(f"En {pat_n} patrones similares el precio {direction} {pat_wr:.0f}% de las veces (media {pat_avg:+.1f}%)")
            if zscore is not None:
                if zscore <= -1.5:
                    parts.append(f"Precio {abs(zscore):.1f}σ bajo su media 20d → posible rebote")
                elif zscore >= 1.5:
                    parts.append(f"Precio {zscore:.1f}σ sobre su media 20d → posible corrección")
            interp = '  ·  '.join(parts) if parts else 'Datos insuficientes para predicción'

            tag = sig_color.get(signal, 'neutral')

            tree.insert('', 'end', tags=(tag,), values=(
                sym,
                row.get('sector', ''),
                signal,
                f"{trend10:+.1f}%" if trend10 is not None else '—',
                f"{r2:.2f}"        if r2       is not None else '—',
                f"{pat_wr:.0f}%"   if pat_wr   is not None else '—',
                f"{pat_avg:+.2f}%" if pat_avg  is not None else '—',
                str(int(pat_n))    if pat_n    else '—',
                f"{zscore:+.2f}"   if zscore   is not None else '—',
                f"${price:.2f}"    if price    else '—',
                proj_str,
                interp,
            ))

        # Resumen
        n_sube  = len(df[df['pred_signal'].isin(['SUBE', 'SUBE FUERTE'])])
        n_baja  = len(df[df['pred_signal'].isin(['BAJA', 'BAJA FUERTE'])])
        n_neu   = len(df) - n_sube - n_baja
        top_up  = df[df['pred_signal'].isin(['SUBE', 'SUBE FUERTE'])].head(3)['symbol'].tolist()

        txt = (f"  De {len(df)} acciones — {n_sube} señal SUBE · {n_baja} señal BAJA · {n_neu} NEUTRO")
        if top_up:
            txt += f"   |   Mejores predicciones: {', '.join(top_up)}"
        self.pred_summary.config(text=txt)

    def _compute_bigdip_score(self, row):
        return compute_bigdip_score(row)

    def _compute_dip_score(self, row):
        return compute_dip_score(row)

    def _compute_price_targets(self, row, invested_eur=1000):
        """
        Calcula entrada, stop, TP1, TP2 y valor esperado por trade.
        Retorna dict con todos los valores ya formateados como strings.
        """
        price   = row.get('price')
        atr_pct = row.get('atr_pct')
        res_52w = row.get('resistance_52w')
        bt_wr   = row.get('bt_win_rate_vol')  # histórico win rate (fracción)

        if not price or not atr_pct:
            return None

        atr_val  = price * atr_pct / 100

        # Entry zone: ligeramente por debajo del precio actual (espera retroceso leve)
        entry_lo = price - 0.5 * atr_val
        entry_hi = price + 0.3 * atr_val

        # Stop loss: precio - 2×ATR (margen amplio para no saltarse en ruido)
        stop      = price - 2.0 * atr_val
        risk_pct  = (price - stop) / price * 100

        # TP1 = 1.5× el riesgo desde entrada (R/R 1.5:1)
        risk_pts  = price - stop
        tp1       = price + 1.5 * risk_pts
        tp1_pct   = (tp1 - price) / price * 100

        # TP2 = 2.5× el riesgo (R/R 2.5:1) o resistencia 52w si está más cerca
        tp2_base  = price + 2.5 * risk_pts
        if res_52w and res_52w > price:
            tp2 = min(tp2_base, res_52w * 0.98)   # 2% bajo resistencia
        else:
            tp2 = tp2_base
        tp2_pct   = (tp2 - price) / price * 100

        # Valor esperado si hay win rate histórico
        wr = bt_wr if (bt_wr and bt_wr > 0) else 0.55   # 55% por defecto
        ev_pct = wr * tp1_pct + (1 - wr) * (-risk_pct)

        # En euros
        shares_approx = invested_eur / price
        loss_eur      = shares_approx * (price - stop)
        gain_tp1_eur  = shares_approx * (tp1 - price)
        gain_tp2_eur  = shares_approx * (tp2 - price)
        ev_eur        = invested_eur * ev_pct / 100

        rr1 = tp1_pct / risk_pct if risk_pct > 0 else 0
        rr2 = tp2_pct / risk_pct if risk_pct > 0 else 0

        return {
            'entry_lo':     entry_lo,
            'entry_hi':     entry_hi,
            'stop':         stop,
            'tp1':          tp1,
            'tp2':          tp2,
            'risk_pct':     risk_pct,
            'tp1_pct':      tp1_pct,
            'tp2_pct':      tp2_pct,
            'rr1':          rr1,
            'rr2':          rr2,
            'win_rate_pct': wr * 100,
            'ev_pct':       ev_pct,
            'loss_eur':     loss_eur,
            'gain_tp1_eur': gain_tp1_eur,
            'gain_tp2_eur': gain_tp2_eur,
            'ev_eur':       ev_eur,
        }

    def _show_price_target_popup(self, row, invested_eur=1000):
        """Ventana emergente con análisis completo de precio objetivo para una acción."""
        sym    = row.get('symbol', '?')
        price  = row.get('price')
        t      = self._compute_price_targets(row, invested_eur)

        win = tk.Toplevel(self.root)
        win.title(f'Precio Objetivo — {sym}')
        win.configure(bg=COLORS['bg'])
        win.geometry('560x440')
        win.resizable(False, False)

        # Título
        hdr = tk.Frame(win, bg=COLORS['surface'], padx=20, pady=12)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text=f'{sym}  —  Análisis de Precio Objetivo',
                 bg=COLORS['surface'], fg=COLORS['primary'],
                 font=('Segoe UI', 12, 'bold')).pack(anchor='w')
        tk.Label(hdr, text=f'Precio actual: ${price:.2f}   |   Capital simulado: €{invested_eur:,}',
                 bg=COLORS['surface'], fg=COLORS['text_dim'],
                 font=('Segoe UI', 9)).pack(anchor='w', pady=(4, 0))

        if not t:
            tk.Label(win, text='No hay suficientes datos para calcular objetivos.',
                     bg=COLORS['bg'], fg=COLORS['text_dim'],
                     font=('Segoe UI', 10)).pack(pady=40)
            return

        # ── Cuadro principal ──
        body = tk.Frame(win, bg=COLORS['bg'], padx=24, pady=16)
        body.pack(fill=tk.BOTH, expand=True)

        def row_label(parent, label, value, fg=COLORS['text'], large=False):
            f = tk.Frame(parent, bg=COLORS['bg'])
            f.pack(fill=tk.X, pady=2)
            tk.Label(f, text=label, bg=COLORS['bg'], fg=COLORS['text_dim'],
                     font=('Segoe UI', 10), width=26, anchor='w').pack(side=tk.LEFT)
            tk.Label(f, text=value, bg=COLORS['bg'], fg=fg,
                     font=('Segoe UI', 11 if large else 10, 'bold' if large else 'normal'),
                     anchor='w').pack(side=tk.LEFT)

        tk.Label(body, text='── DÓNDE ENTRAR ──',
                 bg=COLORS['bg'], fg=COLORS['text_muted'],
                 font=('Segoe UI', 9, 'bold')).pack(anchor='w', pady=(0, 4))

        row_label(body, 'Zona de entrada:',
                  f"${t['entry_lo']:.2f} – ${t['entry_hi']:.2f}",
                  fg=COLORS['primary'], large=True)
        row_label(body, 'Precio actual:',
                  f"${price:.2f}  (entra si cae a la zona o compra ya)")

        tk.Label(body, text='── GESTIÓN DE RIESGO ──',
                 bg=COLORS['bg'], fg=COLORS['text_muted'],
                 font=('Segoe UI', 9, 'bold')).pack(anchor='w', pady=(10, 4))

        row_label(body, 'Stop Loss  (sal si baja a):',
                  f"${t['stop']:.2f}  (−{t['risk_pct']:.1f}%)",
                  fg=COLORS['danger'])
        row_label(body, 'Objetivo 1 (TP1):',
                  f"${t['tp1']:.2f}  (+{t['tp1_pct']:.1f}%)  R/R {t['rr1']:.1f}:1",
                  fg=COLORS['success'])
        row_label(body, 'Objetivo 2 (TP2):',
                  f"${t['tp2']:.2f}  (+{t['tp2_pct']:.1f}%)  R/R {t['rr2']:.1f}:1",
                  fg=COLORS['success'])

        tk.Label(body, text='── SI INVIERTES €1.000 ──',
                 bg=COLORS['bg'], fg=COLORS['text_muted'],
                 font=('Segoe UI', 9, 'bold')).pack(anchor='w', pady=(10, 4))

        row_label(body, 'Pérdida máxima (stop):',
                  f"−€{t['loss_eur']:.0f}",
                  fg=COLORS['danger'])
        row_label(body, 'Ganancia TP1:',
                  f"+€{t['gain_tp1_eur']:.0f}",
                  fg=COLORS['success'])
        row_label(body, 'Ganancia TP2:',
                  f"+€{t['gain_tp2_eur']:.0f}",
                  fg=COLORS['success'])

        ev_color = COLORS['success'] if t['ev_eur'] >= 0 else COLORS['danger']
        row_label(body,
                  f"Valor esperado ({t['win_rate_pct']:.0f}% win rate histórico):",
                  f"{t['ev_eur']:+.0f}€ por trade  ({t['ev_pct']:+.1f}%)",
                  fg=ev_color, large=True)

        verdict = ('✅ Relación riesgo/ganancia FAVORABLE — vale la pena considerar'
                   if t['rr1'] >= 1.5 and t['ev_eur'] >= 0
                   else '⚠️ Relación moderada — define bien tu stop antes de entrar')
        tk.Label(body, text=verdict,
                 bg=COLORS['surface'], fg=COLORS['text'],
                 font=('Segoe UI', 9), padx=12, pady=8,
                 wraplength=500, justify='left').pack(fill=tk.X, pady=(12, 0))

        tk.Button(win, text='Cerrar', bg=COLORS['surface3'], fg=COLORS['text'],
                  relief='flat', cursor='hand2',
                  command=win.destroy).pack(pady=12)

    def _build_dip_tab(self):
        p = self.tab_dip

        # ── Título ──
        hdr = tk.Frame(p, bg=COLORS['bg'])
        hdr.pack(fill=tk.X, padx=24, pady=(20, 0))
        tk.Label(hdr, text='Caída Atractiva — Comprar en el Dip',
                 bg=COLORS['bg'], fg=COLORS['text'],
                 font=('Segoe UI', 14, 'bold')).pack(side=tk.LEFT)

        # ── Banner explicativo ──
        info = tk.Frame(p, bg=COLORS['surface'], padx=16, pady=10)
        info.pack(fill=tk.X, padx=24, pady=(8, 0))
        tk.Label(info, text='¿Qué es una caída atractiva?',
                 bg=COLORS['surface'], fg=COLORS['primary'],
                 font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        tk.Label(info, bg=COLORS['surface'], fg=COLORS['text_dim'],
                 font=('Segoe UI', 9), justify='left',
                 text=(
                     "Acciones de calidad (score alto) que van al alza en el medio plazo (20d positivo),\n"
                     "pero que han tenido una caída reciente inusual en los últimos 3-5 días.\n"
                     "El objetivo es entrar a un precio más bajo que hace una semana en un activo que sigue siendo bueno.\n\n"
                     "  Caída%    = cuánto ha caído en los últimos 5 días  (negativo = bajó)\n"
                     "  Desde Máx = % que está por debajo del máximo de los últimos 15 días\n"
                     "  Inusual×  = veces que la caída supera la volatilidad diaria normal (≥2× = muy inusual)\n"
                     "  Tendencia = retorno a 20 días — debe ser positivo para confirmar uptrend\n\n"
                     "  Un buen candidato: Tendencia 20d > 0% · Caída 5d < -3% · RSI entre 40-55 · Inusual× ≥ 1.5"
                 )).pack(anchor='w', pady=(4, 0))

        # ── Controles ──
        ctrl = tk.Frame(p, bg=COLORS['bg'])
        ctrl.pack(fill=tk.X, padx=24, pady=(10, 6))

        tk.Label(ctrl, text='Mostrar Top:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)
        self.dip_n = ttk.Combobox(ctrl, width=6, state='readonly',
                                   font=('Segoe UI', 9),
                                   values=['5', '10', '15', '20', 'Todos'])
        self.dip_n.current(1)
        self.dip_n.pack(side=tk.LEFT, padx=(6, 20))
        self.dip_n.bind('<<ComboboxSelected>>', lambda e: self._populate_dip_tab())

        tk.Label(ctrl, text='Tendencia 20d positiva:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)
        self.dip_require_uptrend = tk.BooleanVar(value=True)
        tk.Checkbutton(ctrl, variable=self.dip_require_uptrend,
                       bg=COLORS['bg'], fg=COLORS['text'], selectcolor=COLORS['surface2'],
                       activebackground=COLORS['bg'],
                       command=self._populate_dip_tab).pack(side=tk.LEFT, padx=(4, 20))

        tk.Label(ctrl, text='Score calidad mín:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)
        self.dip_min_score = ttk.Combobox(ctrl, width=5, state='readonly',
                                           font=('Segoe UI', 9),
                                           values=['40', '50', '55', '60'])
        self.dip_min_score.current(1)   # 50 por defecto
        self.dip_min_score.pack(side=tk.LEFT, padx=(6, 0))
        self.dip_min_score.bind('<<ComboboxSelected>>', lambda e: self._populate_dip_tab())

        # ── Tabla ──
        cols = ('dip_rank', 'dip_decision', 'symbol', 'sector',
                'dip_score', 'confidence', 'precio',
                'caida_5d', 'tendencia_20d', 'desde_max', 'inusual',
                'rsi', 'vol', 'stop_atr', 'razones')
        hdrs = ('#', 'Decisión', 'Símbolo', 'Sector',
                'DipScore', 'Confianza', 'Precio',
                'Caída 5d', 'Tend. 20d', 'Desde Máx', 'Inusual×',
                'RSI', 'Vol MA', 'Stop ATR', 'Por qué es atractiva')

        self.dip_tree, scroll_dip = self._make_treeview(
            p, cols, hdrs,
            col_widths=[32, 92, 68, 140,
                        72, 86, 72,
                        78, 82, 80, 72,
                        52, 78, 72, 300])
        scroll_dip.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 4))
        self.dip_tree.bind('<Double-1>', lambda e: self._detail_from_tree(self.dip_tree, 2))

        # ── Botón precio objetivo ──
        btn_row_dip = tk.Frame(p, bg=COLORS['bg'])
        btn_row_dip.pack(fill=tk.X, padx=24, pady=(2, 4))
        tk.Button(btn_row_dip,
                  text='🎯 Ver Precio Objetivo del seleccionado',
                  bg=COLORS['success'], fg='white',
                  font=('Segoe UI', 9, 'bold'), relief='flat', cursor='hand2',
                  command=lambda: self._show_target_from_tree(self.dip_tree, 2)
                  ).pack(side=tk.LEFT)
        tk.Label(btn_row_dip,
                 text='  Entrada · Stop · TP1 · TP2 · R/R · "si meto €1000"',
                 bg=COLORS['bg'], fg=COLORS['text_dim'],
                 font=('Segoe UI', 9)).pack(side=tk.LEFT)

        # ── Resumen ──
        self.dip_summary = tk.Label(p, text='',
                                     bg=COLORS['surface'], fg=COLORS['text_dim'],
                                     font=('Segoe UI', 9), anchor='w',
                                     padx=16, pady=6)
        self.dip_summary.pack(fill=tk.X, padx=24, pady=(0, 12))

    def _show_target_from_tree(self, tree, sym_col):
        """Obtiene el símbolo seleccionado del tree y muestra popup de precio objetivo."""
        if self.df_results is None or self.df_results.empty:
            messagebox.showinfo('Sin datos', 'Ejecuta primero el análisis.')
            return
        sel = tree.selection()
        if not sel:
            messagebox.showinfo('Selección', 'Selecciona una fila primero.')
            return
        sym = str(tree.item(sel[0], 'values')[sym_col]).strip()
        rows = self.df_results[self.df_results['symbol'] == sym]
        if rows.empty:
            return
        self._show_price_target_popup(rows.iloc[0].to_dict())

    def _populate_dip_tab(self):
        if self.df_results is None or self.df_results.empty:
            return

        df = self.df_results.copy()

        # Filtro tendencia positiva
        if self.dip_require_uptrend.get():
            df = df[df['change_20d'].fillna(-999) > 0]

        # Filtro score de calidad mínimo
        min_q = int(self.dip_min_score.get())
        df['_buy_score'] = df.apply(self._compute_buy_score, axis=1)
        df = df[df['_buy_score'] >= min_q]

        if df.empty:
            self.dip_tree.delete(*self.dip_tree.get_children())
            self.dip_summary.config(text='Sin candidatos con los filtros actuales.')
            return

        # Calcular dip score y ordenar
        df['_dip_score'] = df.apply(self._compute_dip_score, axis=1)
        df = df.sort_values('_dip_score', ascending=False, na_position='last')

        # Limitar a Top N
        n_sel = self.dip_n.get()
        if n_sel != 'Todos':
            df = df.head(int(n_sel))

        tree = self.dip_tree
        tree.delete(*tree.get_children())

        for i, (_, row) in enumerate(df.iterrows(), 1):
            dip_sc  = row['_dip_score']
            c5      = row.get('change_5d')
            c20     = row.get('change_20d')
            dip_h   = row.get('dip_from_high_pct')
            inusual = row.get('dip_atr_multiple')
            rsi     = row.get('rsi')
            price   = row.get('price')
            atr_pct = row.get('atr_pct')
            vol     = row.get('vol_signal', '—')
            conf    = row.get('confidence', '—')

            # Stop ATR
            if price and atr_pct:
                atr_val  = price * atr_pct / 100
                stop_atr = f"${price - 2 * atr_val:.2f}"
            else:
                stop_atr = '—'

            # Decisión
            if dip_sc >= 65:
                decision = '🟢 COMPRAR DIP'
            elif dip_sc >= 50:
                decision = '🟡 VIGILAR'
            else:
                decision = '🔴 ESPERAR'

            # Razón legible
            reasons = []
            if c20 is not None and c20 > 0:
                reasons.append(f"↑ Uptrend {c20:+.1f}%")
            if c5 is not None and c5 < -2:
                reasons.append(f"↓ Caída {c5:.1f}%")
            if dip_h is not None and dip_h >= 3:
                reasons.append(f"Bajo máx {dip_h:.1f}%")
            if inusual is not None and inusual >= 1.5:
                reasons.append(f"Caída inusual {inusual:.1f}×ATR")
            if rsi is not None and 35 <= rsi <= 55:
                reasons.append(f"RSI neutro ({rsi:.0f})")
            elif rsi is not None and rsi < 35:
                reasons.append(f"RSI oversold ({rsi:.0f})")
            if vol == 'COMPRAR':
                reasons.append("Vol MA confirma compra")
            reason_str = ' · '.join(reasons) if reasons else '—'

            # Color de fila
            tag = 'buy' if dip_sc >= 65 else 'neutral' if dip_sc >= 50 else 'sell'

            tree.insert('', 'end', tags=(tag,), values=(
                i,
                decision,
                row.get('symbol', ''),
                row.get('sector', ''),
                f"{dip_sc:.1f}",
                conf,
                f"${price:.2f}" if price else '—',
                f"{c5:+.1f}%"  if c5  is not None else '—',
                f"{c20:+.1f}%" if c20 is not None else '—',
                f"{dip_h:.1f}%" if dip_h is not None else '—',
                f"{inusual:.1f}×" if inusual is not None else '—',
                f"{rsi:.0f}"   if rsi is not None else '—',
                vol,
                stop_atr,
                reason_str,
            ))

        # ── Resumen accionable ──
        comprar_dip = df[df['_dip_score'] >= 65]
        if not comprar_dip.empty:
            syms = ', '.join(comprar_dip['symbol'].head(5).tolist())
            txt = f"🟢 Oportunidades de compra en dip: {syms}"
        elif not df.empty:
            syms = ', '.join(df['symbol'].head(3).tolist())
            txt = f"🟡 Sin caídas claras ahora. Candidatos a vigilar: {syms}"
        else:
            txt = "Sin candidatos con los filtros actuales."

        best = df.iloc[0]['_dip_score'] if not df.empty else 0
        self.dip_summary.config(
            text=f"  {txt}   |   DipScore máximo: {best:.1f}/100"
        )

    # ─── TAB 3: RANKINGS ─────────────────────

    def _build_rankings(self):
        p = self.tab_rankings

        # Filtros
        flt = tk.Frame(p, bg=COLORS['bg'])
        flt.pack(fill=tk.X, padx=24, pady=(16, 8))

        tk.Label(flt, text='Filtrar sector:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)

        self.filter_sector = ttk.Combobox(flt, width=22, state='readonly',
                                           font=('Segoe UI', 9))
        self.filter_sector.pack(side=tk.LEFT, padx=(6, 20))
        self.filter_sector.bind('<<ComboboxSelected>>', self._apply_ranking_filter)

        tk.Label(flt, text='Ordenar por:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)

        self.sort_col = ttk.Combobox(flt, width=20, state='readonly',
                                      font=('Segoe UI', 9),
                                      values=['Confianza', 'Score', 'RSI', 'Cambio 5d',
                                              'Cambio 20d', 'Señal Vol', 'Ratio Vol'])
        self.sort_col.current(0)
        self.sort_col.pack(side=tk.LEFT, padx=(6, 20))
        self.sort_col.bind('<<ComboboxSelected>>', self._apply_ranking_filter)

        tk.Label(flt, text='Buscar:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        search_entry = tk.Entry(flt, textvariable=self.search_var, width=10,
                                font=('Segoe UI', 9),
                                bg=COLORS['surface2'], fg=COLORS['text'],
                                insertbackground=COLORS['text'],
                                relief='flat', bd=4)
        search_entry.pack(side=tk.LEFT, padx=(6, 4))
        self.search_var.trace_add('write', lambda *_: self._apply_ranking_filter())

        tk.Button(flt, text='✕', font=('Segoe UI', 8), relief='flat',
                  bg=COLORS['surface2'], fg=COLORS['text_dim'],
                  cursor='hand2', bd=0,
                  command=lambda: self.search_var.set('')).pack(side=tk.LEFT)

        # Tabla
        cols = ('rank', 'symbol', 'sector', 'score', 'precio', 'cambio5d',
                'cambio20d', 'rsi', 'adx', 'vol_signal', 'tf_align', 'candle', 'fund', 'confidence')
        hdrs = ('#', 'Símbolo', 'Sector', 'Score', 'Precio', 'Δ 5d%',
                'Δ 20d%', 'RSI', 'ADX', 'Vol MA', 'Multi-TF', 'Patrón Vela', 'Fund.', 'Confianza')

        self.rank_tree, scroll_f = self._make_treeview(p, cols, hdrs,
                                                        col_widths=[32, 62, 140, 70, 78, 62,
                                                                    62, 48, 80, 78, 170, 150, 86, 88])
        scroll_f.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 16))
        self.rank_tree.bind('<Double-1>', lambda e: self._detail_from_tree(self.rank_tree, 1))

    # ─── TAB 3: VOLUMEN MA ───────────────────

    def _build_volume(self):
        p = self.tab_volume

        # Explicación
        info_f = tk.Frame(p, bg=COLORS['surface'], pady=16, padx=20)
        info_f.pack(fill=tk.X, padx=24, pady=(16, 0))

        tk.Label(info_f, text='Estrategia de Volumen MA (idea de tu padre)',
                 bg=COLORS['surface'], fg=COLORS['primary'],
                 font=('Segoe UI', 11, 'bold')).pack(anchor='w')

        explanation = (
            "Monto Transado = Precio de Cierre × Volumen  (cuánto dinero se movió ese día)\n"
            "MA 7 días  →  promedio del monto transado en los últimos 7 días\n"
            "MA 60 días →  promedio del monto transado en los últimos 60 días\n\n"
            "   COMPRAR  si  MA7 > MA60   (actividad creciente = dinero entrando)\n"
            "   VENDER   si  MA7 < MA60   (actividad decreciente = dinero saliendo)\n\n"
            "Ratio = MA7 / MA60 · Ratio > 1 → señal de compra · Ratio < 1 → señal de venta"
        )
        tk.Label(info_f, text=explanation, bg=COLORS['surface'], fg=COLORS['text_dim'],
                 font=('Consolas', 9), justify='left').pack(anchor='w', pady=(8, 0))

        # Filtro señal
        flt = tk.Frame(p, bg=COLORS['bg'])
        flt.pack(fill=tk.X, padx=24, pady=(12, 8))

        tk.Label(flt, text='Mostrar:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)

        self.vol_filter = ttk.Combobox(flt, width=14, state='readonly',
                                        font=('Segoe UI', 9),
                                        values=['Todas', 'Solo COMPRAR', 'Solo VENDER'])
        self.vol_filter.current(0)
        self.vol_filter.pack(side=tk.LEFT, padx=(6, 0))
        self.vol_filter.bind('<<ComboboxSelected>>', self._apply_volume_filter)

        # Tabla
        cols = ('symbol', 'sector', 'precio', 'signal', 'ratio', 'strength',
                'ma7', 'ma60', 'ma7_trend', 'score_tec')
        hdrs = ('Símbolo', 'Sector', 'Precio', 'Señal', 'Ratio MA7/MA60',
                'Fuerza %', 'MA7 ($M)', 'MA60 ($M)', 'Tendencia MA7', 'Score Téc.')

        self.vol_tree, scroll_f = self._make_treeview(p, cols, hdrs,
                                                       col_widths=[70, 160, 80, 90, 120,
                                                                   80, 90, 90, 110, 80])
        scroll_f.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 16))
        self.vol_tree.bind('<Double-1>', lambda e: self._detail_from_tree(self.vol_tree, 0))

    # ─── TAB 4: PATRONES DE VELA ─────────────

    def _build_candles(self):
        p = self.tab_candles

        # Info card
        info_f = tk.Frame(p, bg=COLORS['surface'], pady=16, padx=20)
        info_f.pack(fill=tk.X, padx=24, pady=(16, 0))

        tk.Label(info_f, text='Patrones de Velas Japonesas (última sesión)',
                 bg=COLORS['surface'], fg=COLORS['primary'],
                 font=('Segoe UI', 11, 'bold')).pack(anchor='w')

        explanation = (
            "Se detectan los patrones de las últimas 3 velas diarias de cada acción.\n\n"
            "  COMPRA  →  Hammer · Bullish Engulfing · Morning Star\n"
            "  VENTA   →  Hanging Man · Bearish Engulfing · Evening Star\n"
            "  NEUTRO  →  Doji (indecisión del mercado)\n\n"
            "Fuerza: muy fuerte = señal de alta confianza  |  fuerte = señal válida  |  débil = confirmar con otros indicadores\n"
            "Tip: un patrón es más fiable cuando coincide con la señal de Volumen MA."
        )
        tk.Label(info_f, text=explanation, bg=COLORS['surface'], fg=COLORS['text_dim'],
                 font=('Consolas', 9), justify='left').pack(anchor='w', pady=(8, 0))

        # Filtros
        flt = tk.Frame(p, bg=COLORS['bg'])
        flt.pack(fill=tk.X, padx=24, pady=(12, 8))

        tk.Label(flt, text='Señal:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)

        self.candle_filter_signal = ttk.Combobox(flt, width=14, state='readonly',
                                                  font=('Segoe UI', 9),
                                                  values=['Todas', 'Solo COMPRA', 'Solo VENTA', 'Solo NEUTRO'])
        self.candle_filter_signal.current(0)
        self.candle_filter_signal.pack(side=tk.LEFT, padx=(6, 20))
        self.candle_filter_signal.bind('<<ComboboxSelected>>', self._apply_candle_filter)

        tk.Label(flt, text='Fuerza:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)

        self.candle_filter_strength = ttk.Combobox(flt, width=14, state='readonly',
                                                    font=('Segoe UI', 9),
                                                    values=['Todas', 'muy fuerte', 'fuerte', 'débil'])
        self.candle_filter_strength.current(0)
        self.candle_filter_strength.pack(side=tk.LEFT, padx=(6, 0))
        self.candle_filter_strength.bind('<<ComboboxSelected>>', self._apply_candle_filter)

        # Tabla
        cols = ('symbol', 'sector', 'precio', 'pattern', 'signal', 'strength',
                'vol_signal', 'rsi', 'score')
        hdrs = ('Símbolo', 'Sector', 'Precio', 'Patrón Detectado', 'Señal Vela',
                'Fuerza', 'Vol MA Señal', 'RSI', 'Score Téc.')

        self.candle_tree, scroll_f = self._make_treeview(p, cols, hdrs,
                                                          col_widths=[70, 160, 80, 200,
                                                                      90, 90, 100, 55, 80])
        scroll_f.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 16))
        self.candle_tree.bind('<Double-1>', lambda e: self._detail_from_tree(self.candle_tree, 0))

    def _apply_candle_filter(self, event=None):
        if self.df_results is None or self.df_results.empty:
            return
        df = self.df_results.copy()

        sig = self.candle_filter_signal.get()
        if sig == 'Solo COMPRA':
            df = df[df['candle_signal'] == 'COMPRA']
        elif sig == 'Solo VENTA':
            df = df[df['candle_signal'] == 'VENTA']
        elif sig == 'Solo NEUTRO':
            df = df[df['candle_signal'] == 'NEUTRO']

        strength = self.candle_filter_strength.get()
        if strength in ('muy fuerte', 'fuerte', 'débil'):
            df = df[df['candle_strength'] == strength]

        # Ordenar: muy fuerte primero, luego COMPRA antes que VENTA
        strength_order = {'muy fuerte': 3, 'fuerte': 2, 'débil': 1, '—': 0}
        signal_order   = {'COMPRA': 2, 'VENTA': 1, 'NEUTRO': 0}
        df['_so'] = df['candle_strength'].map(strength_order).fillna(0)
        df['_ss'] = df['candle_signal'].map(signal_order).fillna(0)
        df = df.sort_values(['_so', '_ss', 'total_score'], ascending=[False, False, False])

        tree = self.candle_tree
        tree.delete(*tree.get_children())

        for i, (_, row) in enumerate(df.iterrows()):
            tag = 'odd' if i % 2 == 0 else 'even'
            cs = row.get('candle_signal', 'NEUTRO')
            cp = row.get('candle_pattern', 'Sin patrón')
            ck = row.get('candle_strength', '—')
            vs = row.get('vol_signal', 'N/A')

            if cs == 'COMPRA':
                row_tag = (tag, 'buy')
            elif cs == 'VENTA':
                row_tag = (tag, 'sell')
            else:
                row_tag = (tag,)

            tree.insert('', 'end', tags=row_tag, values=(
                row['symbol'],
                row['sector'],
                f"${row['price']:.2f}",
                cp,
                cs,
                ck,
                vs,
                f"{row['rsi']:.0f}",
                f"{row['total_score']:.1f}",
            ))

    # ─── TAB 5: MULTI-TIMEFRAME ──────────────

    def _build_multitf(self):
        p = self.tab_multitf

        info_f = tk.Frame(p, bg=COLORS['surface'], pady=16, padx=20)
        info_f.pack(fill=tk.X, padx=24, pady=(16, 0))

        tk.Label(info_f, text='Alineacion Multi-Temporalidad (Diario vs Semanal)',
                 bg=COLORS['surface'], fg=COLORS['primary'],
                 font=('Segoe UI', 11, 'bold')).pack(anchor='w')

        explanation = (
            "Compara EMAs 9/21/50 en marco SEMANAL con EMAs 9/21 en marco DIARIO.\n\n"
            "  ALINEADO ALCISTA      (+2 pts)  Semanal y diario ambos alcistas — maximo momentum\n"
            "  CONSOLIDANDO ALCISTA  (+1 pt)   Semanal alcista, diario lateral — posible continuacion\n"
            "  CORRECCION EN ALCISTA ( 0 pts)  Semanal alcista, diario cae — buy the dip\n"
            "  REBOTE EN LATERAL     ( 0 pts)  Lateral semanal, diario sube — confirmar rotura\n"
            "  CAIDA EN LATERAL      (-1 pt)   Lateral semanal, diario cae — cuidado\n"
            "  REBOTE EN BAJISTA     (-1 pt)   Semanal bajista, diario sube — trampa alcista\n"
            "  CONSOLIDANDO BAJISTA  (-1 pt)   Semanal bajista, diario lateral — presion vendedora\n"
            "  ALINEADO BAJISTA      (-2 pts)  Semanal y diario ambos bajistas — evitar\n\n"
            "Tip: las mejores entradas son ALINEADO ALCISTA con ADX > 25 y Vol MA = COMPRAR."
        )
        tk.Label(info_f, text=explanation, bg=COLORS['surface'], fg=COLORS['text_dim'],
                 font=('Consolas', 9), justify='left').pack(anchor='w', pady=(8, 0))

        # Filtro
        flt = tk.Frame(p, bg=COLORS['bg'])
        flt.pack(fill=tk.X, padx=24, pady=(12, 8))

        tk.Label(flt, text='Señal TF:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)

        self.tf_filter = ttk.Combobox(flt, width=20, state='readonly',
                                      font=('Segoe UI', 9),
                                      values=['Todas', 'Solo ALCISTA', 'Solo BAJISTA', 'Solo NEUTRO'])
        self.tf_filter.current(0)
        self.tf_filter.pack(side=tk.LEFT, padx=(6, 0))
        self.tf_filter.bind('<<ComboboxSelected>>', self._apply_tf_filter)

        # Tabla
        cols = ('symbol', 'sector', 'precio', 'tf_alignment', 'tf_signal',
                'tf_weekly_trend', 'tf_daily_trend', 'tf_weekly_rsi',
                'tf_pts', 'adx', 'vol_signal', 'score')
        hdrs = ('Simbolo', 'Sector', 'Precio', 'Alineacion', 'Señal TF',
                'Tend. Semanal', 'Tend. Diaria', 'RSI Semanal',
                'Pts TF', 'ADX', 'Vol MA', 'Score Tec.')

        self.tf_tree, scroll_f = self._make_treeview(p, cols, hdrs,
                                                      col_widths=[70, 150, 80, 200, 90,
                                                                  110, 100, 90,
                                                                  60, 70, 80, 80])
        scroll_f.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 16))
        self.tf_tree.bind('<Double-1>', lambda e: self._detail_from_tree(self.tf_tree, 0))

    def _apply_tf_filter(self, event=None):
        if self.df_results is None or self.df_results.empty:
            return
        df = self.df_results.copy()

        sig = self.tf_filter.get()
        if sig == 'Solo ALCISTA':
            df = df[df['tf_signal'] == 'ALCISTA']
        elif sig == 'Solo BAJISTA':
            df = df[df['tf_signal'] == 'BAJISTA']
        elif sig == 'Solo NEUTRO':
            df = df[df['tf_signal'] == 'NEUTRO']

        # Ordenar: mejor alineacion (mayor conf_pts) primero
        if 'tf_conf_pts' in df.columns:
            df = df.sort_values(['tf_conf_pts', 'total_score'], ascending=[False, False])

        tree = self.tf_tree
        tree.delete(*tree.get_children())

        for i, (_, row) in enumerate(df.iterrows()):
            tag = 'odd' if i % 2 == 0 else 'even'
            tf_sig = row.get('tf_signal', 'NEUTRO')

            if tf_sig == 'ALCISTA':
                row_tag = (tag, 'buy')
            elif tf_sig == 'BAJISTA':
                row_tag = (tag, 'sell')
            else:
                row_tag = (tag,)

            pts = row.get('tf_conf_pts', 0)
            adx_v = row.get('adx_value')
            adx_s = row.get('adx_strength', '—')
            adx_display = f"{adx_v:.0f} ({adx_s})" if adx_v else '—'
            wrsi = row.get('tf_weekly_rsi')
            pts_str = f"{int(pts):+d}" if pts is not None else '—'

            tree.insert('', 'end', tags=row_tag, values=(
                row['symbol'],
                row['sector'],
                f"${row['price']:.2f}",
                row.get('tf_alignment', 'N/A'),
                tf_sig,
                row.get('tf_weekly_trend', '—'),
                row.get('tf_daily_trend', '—'),
                f"{wrsi:.1f}" if wrsi else '—',
                pts_str,
                adx_display,
                row.get('vol_signal', 'N/A'),
                f"{row['total_score']:.1f}",
            ))

    # ─── TAB 6: FUERZA RELATIVA vs SPY ──────

    def _build_rs(self):
        p = self.tab_rs

        info_f = tk.Frame(p, bg=COLORS['surface'], pady=16, padx=20)
        info_f.pack(fill=tk.X, padx=24, pady=(16, 0))

        tk.Label(info_f, text='Fuerza Relativa vs S&P 500 (SPY)',
                 bg=COLORS['surface'], fg=COLORS['primary'],
                 font=('Segoe UI', 11, 'bold')).pack(anchor='w')

        explanation = (
            "RS = Retorno del activo - Retorno de SPY en el mismo periodo\n\n"
            "  RS positivo → el activo SUPERA al mercado (lider)\n"
            "  RS negativo → el activo PIERDE contra el mercado (rezagado)\n\n"
            "  Periodos:  RS 20d (corto)  |  RS 60d (medio)  |  RS 252d (largo)\n"
            "  Score RS:  promedio ponderado 20d×30% + 60d×40% + 252d×30%\n\n"
            "  LIDER     (score >= 65)  Supera consistentemente al S&P 500\n"
            "  MERCADO   (score >= 40)  Rendimiento similar al indice\n"
            "  REZAGADO  (score <  40)  Queda por detras del mercado\n\n"
            "Tip: los lideres historicos tienden a seguir siendo lideres (momentum de RS)."
        )
        tk.Label(info_f, text=explanation, bg=COLORS['surface'], fg=COLORS['text_dim'],
                 font=('Consolas', 9), justify='left').pack(anchor='w', pady=(8, 0))

        # Filtro
        flt = tk.Frame(p, bg=COLORS['bg'])
        flt.pack(fill=tk.X, padx=24, pady=(12, 8))

        tk.Label(flt, text='Señal RS:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)

        self.rs_filter = ttk.Combobox(flt, width=18, state='readonly',
                                      font=('Segoe UI', 9),
                                      values=['Todas', 'Solo LIDER', 'Solo MERCADO', 'Solo REZAGADO'])
        self.rs_filter.current(0)
        self.rs_filter.pack(side=tk.LEFT, padx=(6, 0))
        self.rs_filter.bind('<<ComboboxSelected>>', self._apply_rs_filter)

        # Tabla
        cols = ('symbol', 'sector', 'precio', 'rs_signal', 'rs_score',
                'rs_20d', 'rs_60d', 'rs_252d', 'vol_signal', 'tf_align', 'score')
        hdrs = ('Simbolo', 'Sector', 'Precio', 'Señal RS', 'Score RS',
                'RS 20d', 'RS 60d', 'RS 252d', 'Vol MA', 'Multi-TF', 'Score Tec.')

        self.rs_tree, scroll_f = self._make_treeview(p, cols, hdrs,
                                                      col_widths=[70, 150, 80, 90, 80,
                                                                  80, 80, 80, 80, 170, 80])
        scroll_f.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 16))
        self.rs_tree.bind('<Double-1>', lambda e: self._detail_from_tree(self.rs_tree, 0))

    def _apply_rs_filter(self, event=None):
        if self.df_results is None or self.df_results.empty:
            return
        df = self.df_results.copy()

        sig = self.rs_filter.get()
        if sig == 'Solo LIDER':
            df = df[df['rs_signal'] == 'LIDER']
        elif sig == 'Solo MERCADO':
            df = df[df['rs_signal'] == 'MERCADO']
        elif sig == 'Solo REZAGADO':
            df = df[df['rs_signal'] == 'REZAGADO']

        # Ordenar por score RS
        if 'rs_score' in df.columns:
            df = df.sort_values('rs_score', ascending=False)

        tree = self.rs_tree
        tree.delete(*tree.get_children())

        for i, (_, row) in enumerate(df.iterrows()):
            tag = 'odd' if i % 2 == 0 else 'even'
            rs_sig = row.get('rs_signal', 'N/A')

            if rs_sig == 'LIDER':
                row_tag = (tag, 'buy')
            elif rs_sig == 'REZAGADO':
                row_tag = (tag, 'sell')
            else:
                row_tag = (tag,)

            rs_sc  = row.get('rs_score')
            rs_20  = row.get('rs_20d')
            rs_60  = row.get('rs_60d')
            rs_252 = row.get('rs_252d')
            tf_align = row.get('tf_alignment', '—')
            tf_pts   = row.get('tf_conf_pts', 0)
            tf_display = f"{tf_align} ({int(tf_pts):+d})" if tf_align not in ('N/A', '—') else '—'

            tree.insert('', 'end', tags=row_tag, values=(
                row['symbol'],
                row['sector'],
                f"${row['price']:.2f}",
                rs_sig,
                f"{rs_sc:.1f}" if rs_sc is not None else '—',
                f"{rs_20:+.1f}%" if rs_20 is not None else '—',
                f"{rs_60:+.1f}%" if rs_60 is not None else '—',
                f"{rs_252:+.1f}%" if rs_252 is not None else '—',
                row.get('vol_signal', 'N/A'),
                tf_display,
                f"{row['total_score']:.1f}",
            ))

    # ─── TAB 7: RIESGO ───────────────────────

    def _build_risk(self):
        p = self.tab_risk

        info_f = tk.Frame(p, bg=COLORS['surface'], pady=16, padx=20)
        info_f.pack(fill=tk.X, padx=24, pady=(16, 0))

        tk.Label(info_f, text='Gestión de Riesgo — Sharpe · Max Drawdown · Kelly',
                 bg=COLORS['surface'], fg=COLORS['primary'],
                 font=('Segoe UI', 11, 'bold')).pack(anchor='w')

        explanation = (
            "SHARPE RATIO   Retorno anual ajustado por volatilidad (tasa libre de riesgo: 4.5%)\n"
            "               >= 1.5 excelente  |  1.0-1.5 bueno  |  0.5-1.0 aceptable  |  < 0 pierde vs cash\n\n"
            "MAX DRAWDOWN   Mayor caida historica desde máximo hasta mínimo (últimos 252d)\n"
            "               > -10% bajo  |  -10% a -25% moderado  |  < -25% alto\n\n"
            "KELLY %        Porcentaje sugerido del capital a invertir (Half-Kelly, max 20%)\n"
            "               Basado en tasa de acierto y ratio ganancia/perdida de trades de 20d\n"
            "               Ej: Kelly 8% → de cada €1000, invertir €80 en esta posicion\n\n"
            "SCORE RIESGO   Promedio Sharpe (50%) + MaxDD (50%) → BAJO / MEDIO / ALTO RIESGO"
        )
        tk.Label(info_f, text=explanation, bg=COLORS['surface'], fg=COLORS['text_dim'],
                 font=('Consolas', 9), justify='left').pack(anchor='w', pady=(8, 0))

        # Filtro
        flt = tk.Frame(p, bg=COLORS['bg'])
        flt.pack(fill=tk.X, padx=24, pady=(12, 8))

        tk.Label(flt, text='Perfil de riesgo:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)

        self.risk_filter = ttk.Combobox(flt, width=20, state='readonly',
                                        font=('Segoe UI', 9),
                                        values=['Todos', 'BAJO RIESGO', 'RIESGO MEDIO', 'ALTO RIESGO'])
        self.risk_filter.current(0)
        self.risk_filter.pack(side=tk.LEFT, padx=(6, 0))
        self.risk_filter.bind('<<ComboboxSelected>>', self._apply_risk_filter)

        # Tabla
        cols = ('symbol', 'sector', 'precio', 'risk_signal', 'risk_score',
                'sharpe', 'max_dd', 'kelly_pct', 'win_rate', 'ann_ret', 'ann_vol', 'score')
        hdrs = ('Simbolo', 'Sector', 'Precio', 'Perfil Riesgo', 'Score Riesgo',
                'Sharpe', 'Max DD', 'Kelly %', 'Tasa Acierto', 'Ret. Anual', 'Vol. Anual', 'Score Tec.')

        self.risk_tree, scroll_f = self._make_treeview(p, cols, hdrs,
                                                        col_widths=[70, 150, 78, 110, 90,
                                                                    70, 80, 70, 90, 90, 85, 80])
        scroll_f.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 16))
        self.risk_tree.bind('<Double-1>', lambda e: self._detail_from_tree(self.risk_tree, 0))

    def _apply_risk_filter(self, event=None):
        if self.df_results is None or self.df_results.empty:
            return
        df = self.df_results.copy()

        sig = self.risk_filter.get()
        if sig in ('BAJO RIESGO', 'RIESGO MEDIO', 'ALTO RIESGO'):
            df = df[df['risk_signal'] == sig]

        # Ordenar: mejor Sharpe primero
        if 'sharpe' in df.columns:
            df = df.sort_values('sharpe', ascending=False)

        tree = self.risk_tree
        tree.delete(*tree.get_children())

        for i, (_, row) in enumerate(df.iterrows()):
            tag = 'odd' if i % 2 == 0 else 'even'
            risk_sig = row.get('risk_signal', 'N/A')

            if risk_sig == 'BAJO RIESGO':
                row_tag = (tag, 'buy')
            elif risk_sig == 'ALTO RIESGO':
                row_tag = (tag, 'sell')
            else:
                row_tag = (tag,)

            sharpe   = row.get('sharpe')
            max_dd   = row.get('max_dd')
            kelly    = row.get('kelly_pct')
            wr       = row.get('win_rate')
            ann_r    = row.get('ann_return')
            ann_v    = row.get('ann_vol')
            r_score  = row.get('risk_score')

            tree.insert('', 'end', tags=row_tag, values=(
                row['symbol'],
                row['sector'],
                f"${row['price']:.2f}",
                risk_sig,
                f"{r_score:.0f}"   if r_score  is not None else '—',
                f"{sharpe:.2f}"    if sharpe   is not None else '—',
                f"{max_dd:.1f}%"   if max_dd   is not None else '—',
                f"{kelly:.1f}%"    if kelly    is not None else '—',
                f"{wr:.1f}%"       if wr       is not None else '—',
                f"{ann_r:+.1f}%"   if ann_r    is not None else '—',
                f"{ann_v:.1f}%"    if ann_v    is not None else '—',
                f"{row['total_score']:.1f}",
            ))

    # ─── TAB 8: BACKTESTING ──────────────────

    def _build_backtest(self):
        p = self.tab_backtest

        info_f = tk.Frame(p, bg=COLORS['surface'], pady=16, padx=20)
        info_f.pack(fill=tk.X, padx=24, pady=(16, 0))

        tk.Label(info_f, text='Backtesting Histórico — Fiabilidad de la Estrategia',
                 bg=COLORS['surface'], fg=COLORS['primary'],
                 font=('Segoe UI', 11, 'bold')).pack(anchor='w')

        explanation = (
            "Se simulan entradas en cada punto histórico donde se activó la señal\n"
            "y se mide qué ocurrió 20 días después. Sin lookahead — datos reales.\n\n"
            "  TÉCNICA SOLA   Entrada cuando EMA9 > EMA21 > EMA50\n"
            "  TÉC. + VOLUMEN Entrada cuando además Vol MA7 > Vol MA60 (señal más selectiva)\n\n"
            "  Win Rate       % de señales que resultaron en retorno positivo a 20 días\n"
            "  Ret. Medio     Retorno promedio por señal (incluyendo pérdidas)\n"
            "  Expectativa    Ganancia esperada por trade = WR×avg_win + LR×avg_loss\n"
            "  Buy & Hold     Retorno de comprar y mantener todo el período (referencia)\n\n"
            "  >= 65% win rate → señal históricamente fiable (+1 punto de confianza)\n"
            "  <  40% win rate → señal poco fiable (-1 punto de confianza)"
        )
        tk.Label(info_f, text=explanation, bg=COLORS['surface'], fg=COLORS['text_dim'],
                 font=('Consolas', 9), justify='left').pack(anchor='w', pady=(8, 0))

        # Filtro orden
        flt = tk.Frame(p, bg=COLORS['bg'])
        flt.pack(fill=tk.X, padx=24, pady=(12, 8))

        tk.Label(flt, text='Ordenar por:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)

        self.bt_sort = ttk.Combobox(flt, width=22, state='readonly',
                                    font=('Segoe UI', 9),
                                    values=['Win Rate (Téc+Vol)', 'Win Rate (Técnica)',
                                            'Expectativa', 'Ret. Medio', 'Buy & Hold'])
        self.bt_sort.current(0)
        self.bt_sort.pack(side=tk.LEFT, padx=(6, 0))
        self.bt_sort.bind('<<ComboboxSelected>>', self._apply_bt_sort)

        # Tabla
        cols = ('symbol', 'sector',
                'bt_wr_vol', 'bt_ret_vol', 'bt_exp_vol', 'bt_n_vol',
                'bt_wr_tech', 'bt_ret_tech', 'bt_n_tech',
                'bt_best', 'bt_worst', 'bt_bh')
        hdrs = ('Simbolo', 'Sector',
                'WR Téc+Vol', 'Ret Téc+Vol', 'Expect. T+V', 'N T+V',
                'WR Téc', 'Ret Téc', 'N Téc',
                'Mejor', 'Peor', 'Buy&Hold')

        self.bt_tree, scroll_f = self._make_treeview(p, cols, hdrs,
                                                      col_widths=[70, 148,
                                                                  90, 90, 90, 55,
                                                                  80, 80, 55,
                                                                  75, 75, 80])
        scroll_f.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 4))
        self.bt_tree.bind('<Double-1>', lambda e: self._detail_from_tree(self.bt_tree, 0))

        # Botón validación estadística
        btn_row = tk.Frame(p, bg=COLORS['bg'])
        btn_row.pack(fill=tk.X, padx=24, pady=(4, 12))
        tk.Button(btn_row, text='Validar Score Compuesto Estadísticamente',
                  bg=COLORS['primary'], fg='white',
                  font=('Segoe UI', 9, 'bold'), relief='flat', cursor='hand2',
                  command=self._run_score_validation).pack(side=tk.LEFT)
        tk.Label(btn_row,
                 text='  Correlaciona el score de compra (0-100) con retornos históricos (5d / 20d)',
                 bg=COLORS['bg'], fg=COLORS['text_dim'],
                 font=('Segoe UI', 9)).pack(side=tk.LEFT)

    def _apply_bt_sort(self, event=None):
        if self.df_results is None or self.df_results.empty:
            return
        df = self.df_results.copy()

        sort_map = {
            'Win Rate (Téc+Vol)': 'bt_win_rate_vol',
            'Win Rate (Técnica)': 'bt_win_rate_tech',
            'Expectativa':        'bt_expectancy_vol',
            'Ret. Medio':         'bt_avg_ret_vol',
            'Buy & Hold':         'bt_buy_hold',
        }
        col = sort_map.get(self.bt_sort.get(), 'bt_win_rate_vol')
        if col in df.columns:
            df = df.sort_values(col, ascending=False)

        tree = self.bt_tree
        tree.delete(*tree.get_children())

        for i, (_, row) in enumerate(df.iterrows()):
            tag = 'odd' if i % 2 == 0 else 'even'

            wr_vol  = row.get('bt_win_rate_vol')
            if wr_vol is not None and wr_vol >= 65:
                row_tag = (tag, 'buy')
            elif wr_vol is not None and wr_vol < 40:
                row_tag = (tag, 'sell')
            else:
                row_tag = (tag,)

            def fmt_pct(v, sign=False):
                if v is None: return '—'
                return f"{v:+.1f}%" if sign else f"{v:.1f}%"

            def fmt_n(v):
                return str(int(v)) if v else '—'

            tree.insert('', 'end', tags=row_tag, values=(
                row['symbol'],
                row['sector'],
                fmt_pct(row.get('bt_win_rate_vol')),
                fmt_pct(row.get('bt_avg_ret_vol'),    sign=True),
                fmt_pct(row.get('bt_expectancy_vol'), sign=True),
                fmt_n(row.get('bt_n_vol')),
                fmt_pct(row.get('bt_win_rate_tech')),
                fmt_pct(row.get('bt_avg_ret_tech'),   sign=True),
                fmt_n(row.get('bt_n_tech')),
                fmt_pct(row.get('bt_best_tech'),      sign=True),
                fmt_pct(row.get('bt_worst_tech'),     sign=True),
                fmt_pct(row.get('bt_buy_hold'),       sign=True),
            ))

    # ─── VALIDACIÓN ESTADÍSTICA ───────────────

    def _run_score_validation(self):
        """Valida si el Score Compuesto (0-100) predice retornos positivos."""
        if self.df_results is None or self.df_results.empty:
            messagebox.showinfo('Sin datos',
                                'Ejecuta primero el análisis completo.')
            return

        import math
        try:
            import matplotlib
            matplotlib.use('TkAgg')
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            import matplotlib.gridspec as gridspec
        except ImportError:
            messagebox.showerror('Error', 'matplotlib no disponible.')
            return

        df = self.df_results.copy()
        df['_buy_score'] = df.apply(self._compute_buy_score, axis=1)

        # Necesitamos change_5d y change_20d
        df = df.dropna(subset=['_buy_score'])
        df_5d  = df.dropna(subset=['change_5d'])
        df_20d = df.dropna(subset=['change_20d'])

        # ── Correlación Pearson ────────────────────────────────────────────
        def pearson_r(x, y):
            n = len(x)
            if n < 3:
                return 0.0, 1.0
            mx, my = x.mean(), y.mean()
            num = ((x - mx) * (y - my)).sum()
            den = (((x - mx)**2).sum() * ((y - my)**2).sum()) ** 0.5
            r = num / den if den != 0 else 0.0
            # t-statistic → p-value approximation
            if abs(r) >= 1.0:
                return r, 0.0
            t = r * math.sqrt(n - 2) / math.sqrt(1 - r**2)
            # two-tailed p from t-distribution (approx via beta function)
            try:
                import scipy.stats as stats
                p = 2 * stats.t.sf(abs(t), df=n - 2)
            except ImportError:
                # Rough approximation without scipy
                p = math.exp(-0.717 * abs(t) - 0.416 * t**2)
            return r, p

        r5,  p5  = pearson_r(df_5d['_buy_score'],  df_5d['change_5d'])
        r20, p20 = pearson_r(df_20d['_buy_score'], df_20d['change_20d'])

        # ── Buckets por score ──────────────────────────────────────────────
        bins   = [0, 30, 45, 55, 65, 80, 101]
        labels = ['<30', '30-45', '45-55', '55-65', '65-80', '>80']

        def bucket_stats(df_in, ret_col):
            rows = []
            for lo, hi, lbl in zip(bins, bins[1:], labels):
                sub = df_in[(df_in['_buy_score'] >= lo) & (df_in['_buy_score'] < hi)]
                if len(sub) == 0:
                    rows.append({'bucket': lbl, 'n': 0,
                                 'avg_ret': None, 'win_rate': None})
                    continue
                avg = sub[ret_col].mean()
                wr  = (sub[ret_col] > 0).mean() * 100
                rows.append({'bucket': lbl, 'n': len(sub),
                             'avg_ret': avg, 'win_rate': wr})
            return pd.DataFrame(rows)

        bkt5  = bucket_stats(df_5d,  'change_5d')
        bkt20 = bucket_stats(df_20d, 'change_20d')

        # ── Popup ──────────────────────────────────────────────────────────
        win = tk.Toplevel(self.root)
        win.title('Validación Estadística del Score Compuesto')
        win.configure(bg=COLORS['bg'])
        win.geometry('1080x720')

        # Encabezado
        hdr = tk.Frame(win, bg=COLORS['surface'], pady=12, padx=20)
        hdr.pack(fill=tk.X)
        tk.Label(hdr, text='Validación Estadística — Score Compuesto vs Retornos Históricos',
                 bg=COLORS['surface'], fg=COLORS['primary'],
                 font=('Segoe UI', 12, 'bold')).pack(anchor='w')
        n_sym = len(df)

        # Significancia
        def sig_label(p):
            if p < 0.001: return '*** (p<0.001)'
            if p < 0.01:  return '** (p<0.01)'
            if p < 0.05:  return '* (p<0.05)'
            return '(no significativo)'

        r5s  = sig_label(p5)
        r20s = sig_label(p20)

        # ── Veredicto simple ───────────────────────────────────────────────
        # Win rate del bucket más alto disponible en 20d
        top_wr20 = None
        for lbl in reversed(['<30', '30-45', '45-55', '55-65', '65-80', '>80']):
            sub = bkt20[bkt20['bucket'] == lbl]
            if not sub.empty and sub.iloc[0]['win_rate'] is not None and sub.iloc[0]['n'] >= 3:
                top_wr20 = sub.iloc[0]['win_rate']
                break

        sig_ok   = p20 < 0.05 and r20 > 0.15
        wr_ok    = top_wr20 is not None and top_wr20 >= 58
        r_strong = r20 > 0.25

        if sig_ok and wr_ok and r_strong:
            verdict_icon  = '✅'
            verdict_title = 'EL MODELO FUNCIONA — SÍ, sigue el ranking para invertir'
            verdict_color = COLORS['success']
            verdict_exp   = (
                f"El score de compra predice bien los retornos recientes (r={r20:+.2f}, correlación estadísticamente significativa).\n"
                f"Las acciones con score alto tuvieron retorno positivo en ~{top_wr20:.0f}% de los casos a 20 días.\n"
                "➤ Puedes confiar en el orden del Top Compras. #1 = mejor oportunidad ahora mismo."
            )
        elif r20 > 0.05 or (top_wr20 is not None and top_wr20 >= 52):
            verdict_icon  = '⚠️'
            verdict_title = 'RESULTADOS MIXTOS — Usa el ranking como guía, no como única señal'
            verdict_color = COLORS['warning']
            verdict_exp   = (
                f"El modelo muestra correlación positiva débil (r={r20:+.2f}) pero no es concluyente estadísticamente.\n"
                f"Las acciones con score alto tuvieron retorno positivo en ~{top_wr20:.0f}% de los casos a 20 días.\n"
                "➤ El ranking es útil como filtro inicial. Verifica con la pestaña Fundamentales antes de invertir."
            )
        else:
            verdict_icon  = '❌'
            verdict_title = 'MODELO NO CONFIABLE EN ESTE PERIODO — Verifica manualmente'
            verdict_color = COLORS['danger']
            verdict_exp   = (
                f"No se encontró correlación significativa entre el score y los retornos recientes (r={r20:+.2f}).\n"
                "Esto puede ocurrir en mercados muy volátiles o sin tendencia clara.\n"
                "➤ Usa el ranking como referencia pero no como señal de entrada directa."
            )

        verdict_frame = tk.Frame(hdr, bg=verdict_color, padx=14, pady=8)
        verdict_frame.pack(fill=tk.X, pady=(10, 4))
        tk.Label(verdict_frame,
                 text=f"{verdict_icon}  {verdict_title}",
                 bg=verdict_color, fg='white',
                 font=('Segoe UI', 11, 'bold')).pack(anchor='w')
        tk.Label(verdict_frame, text=verdict_exp,
                 bg=verdict_color, fg='white',
                 font=('Segoe UI', 9), justify='left').pack(anchor='w', pady=(4, 0))

        # ── Estadísticas técnicas (para quien quiera profundizar) ──────────
        summary = (
            f"Para curiosos — Estadística:  r(score, 5d) = {r5:+.3f} {r5s}   |   "
            f"r(score, 20d) = {r20:+.3f} {r20s}   |   n={n_sym} símbolos"
        )
        tk.Label(hdr, text=summary, bg=COLORS['surface'], fg=COLORS['text_muted'],
                 font=('Consolas', 8), justify='left').pack(anchor='w', pady=(4, 0))

        # ── Matplotlib ─────────────────────────────────────────────────────
        BG  = COLORS['bg']
        SRF = COLORS['surface']
        TXT = COLORS['text']
        DIM = COLORS['text_dim']
        GRN = COLORS['success']
        RED = COLORS['danger']
        BLU = COLORS['primary']

        fig = plt.Figure(figsize=(14, 7), facecolor=BG)
        gs  = gridspec.GridSpec(2, 3, figure=fig,
                                hspace=0.45, wspace=0.38,
                                left=0.06, right=0.97,
                                top=0.93, bottom=0.10)

        def style_ax(ax, title):
            ax.set_facecolor(SRF)
            ax.set_title(title, color=TXT, fontsize=9, fontweight='bold', pad=6)
            ax.tick_params(colors=DIM, labelsize=8)
            for spine in ax.spines.values():
                spine.set_edgecolor(COLORS['border'])
            ax.xaxis.label.set_color(DIM)
            ax.yaxis.label.set_color(DIM)

        def bar_colors(vals):
            return [GRN if (v is not None and v >= 50) else RED
                    for v in vals]

        # ── Plot 1: Scatter score vs 5d ────────────────────────────────────
        ax1 = fig.add_subplot(gs[0, 0])
        style_ax(ax1, 'Score vs Retorno 5 días')
        if not df_5d.empty:
            sc5  = df_5d['_buy_score'].values
            ret5 = df_5d['change_5d'].values
            colors5 = [GRN if r > 0 else RED for r in ret5]
            ax1.scatter(sc5, ret5, c=colors5, alpha=0.65, s=25, linewidths=0)
            # tendencia
            if len(sc5) > 2:
                m = ((sc5 - sc5.mean()) * (ret5 - ret5.mean())).sum() / ((sc5 - sc5.mean())**2).sum()
                b = ret5.mean() - m * sc5.mean()
                xs = [sc5.min(), sc5.max()]
                ax1.plot(xs, [m*x+b for x in xs], color=BLU, lw=1.5, linestyle='--', alpha=0.8)
            ax1.axhline(0, color=COLORS['border'], lw=0.8, linestyle=':')
            ax1.set_xlabel('Score Compuesto')
            ax1.set_ylabel('Retorno 5d (%)')
            ax1.text(0.05, 0.93, f'r={r5:+.3f} {r5s}',
                     transform=ax1.transAxes, color=BLU, fontsize=7.5, va='top')

        # ── Plot 2: Scatter score vs 20d ───────────────────────────────────
        ax2 = fig.add_subplot(gs[0, 1])
        style_ax(ax2, 'Score vs Retorno 20 días')
        if not df_20d.empty:
            sc20  = df_20d['_buy_score'].values
            ret20 = df_20d['change_20d'].values
            colors20 = [GRN if r > 0 else RED for r in ret20]
            ax2.scatter(sc20, ret20, c=colors20, alpha=0.65, s=25, linewidths=0)
            if len(sc20) > 2:
                m = ((sc20 - sc20.mean()) * (ret20 - ret20.mean())).sum() / ((sc20 - sc20.mean())**2).sum()
                b = ret20.mean() - m * sc20.mean()
                xs = [sc20.min(), sc20.max()]
                ax2.plot(xs, [m*x+b for x in xs], color=BLU, lw=1.5, linestyle='--', alpha=0.8)
            ax2.axhline(0, color=COLORS['border'], lw=0.8, linestyle=':')
            ax2.set_xlabel('Score Compuesto')
            ax2.set_ylabel('Retorno 20d (%)')
            ax2.text(0.05, 0.93, f'r={r20:+.3f} {r20s}',
                     transform=ax2.transAxes, color=BLU, fontsize=7.5, va='top')

        # ── Plot 3: Win rate por bucket 5d ─────────────────────────────────
        ax3 = fig.add_subplot(gs[0, 2])
        style_ax(ax3, 'Win Rate por Bucket Score (5d)')
        valid3 = bkt5.dropna(subset=['win_rate'])
        if not valid3.empty:
            xpos = range(len(valid3))
            bars3 = ax3.bar(xpos, valid3['win_rate'],
                            color=bar_colors(valid3['win_rate']),
                            width=0.65, alpha=0.85)
            ax3.axhline(50, color=COLORS['warning'], lw=1, linestyle='--', alpha=0.7)
            ax3.set_xticks(xpos)
            ax3.set_xticklabels(valid3['bucket'], fontsize=8)
            ax3.set_ylabel('Win Rate (%)')
            ax3.set_ylim(0, 105)
            for bar, (_, row) in zip(bars3, valid3.iterrows()):
                ax3.text(bar.get_x() + bar.get_width()/2,
                         bar.get_height() + 1,
                         f"{row['win_rate']:.0f}%\n(n={int(row['n'])})",
                         ha='center', va='bottom', color=TXT, fontsize=7)

        # ── Plot 4: Win rate por bucket 20d ────────────────────────────────
        ax4 = fig.add_subplot(gs[1, 0])
        style_ax(ax4, 'Win Rate por Bucket Score (20d)')
        valid4 = bkt20.dropna(subset=['win_rate'])
        if not valid4.empty:
            xpos = range(len(valid4))
            bars4 = ax4.bar(xpos, valid4['win_rate'],
                            color=bar_colors(valid4['win_rate']),
                            width=0.65, alpha=0.85)
            ax4.axhline(50, color=COLORS['warning'], lw=1, linestyle='--', alpha=0.7)
            ax4.set_xticks(xpos)
            ax4.set_xticklabels(valid4['bucket'], fontsize=8)
            ax4.set_ylabel('Win Rate (%)')
            ax4.set_ylim(0, 105)
            for bar, (_, row) in zip(bars4, valid4.iterrows()):
                ax4.text(bar.get_x() + bar.get_width()/2,
                         bar.get_height() + 1,
                         f"{row['win_rate']:.0f}%\n(n={int(row['n'])})",
                         ha='center', va='bottom', color=TXT, fontsize=7)

        # ── Plot 5: Retorno medio por bucket 5d ────────────────────────────
        ax5 = fig.add_subplot(gs[1, 1])
        style_ax(ax5, 'Retorno Medio por Bucket Score (5d)')
        valid5 = bkt5.dropna(subset=['avg_ret'])
        if not valid5.empty:
            xpos = range(len(valid5))
            bars5 = ax5.bar(xpos, valid5['avg_ret'],
                            color=[GRN if v >= 0 else RED for v in valid5['avg_ret']],
                            width=0.65, alpha=0.85)
            ax5.axhline(0, color=COLORS['border'], lw=0.8, linestyle=':')
            ax5.set_xticks(xpos)
            ax5.set_xticklabels(valid5['bucket'], fontsize=8)
            ax5.set_ylabel('Retorno Medio (%)')
            for bar, (_, row) in zip(bars5, valid5.iterrows()):
                val = row['avg_ret']
                ypos = max(val, 0) + 0.05 if val >= 0 else min(val, 0) - 0.3
                ax5.text(bar.get_x() + bar.get_width()/2, ypos,
                         f"{val:+.2f}%", ha='center', va='bottom',
                         color=TXT, fontsize=7)

        # ── Plot 6: Retorno medio por bucket 20d ───────────────────────────
        ax6 = fig.add_subplot(gs[1, 2])
        style_ax(ax6, 'Retorno Medio por Bucket Score (20d)')
        valid6 = bkt20.dropna(subset=['avg_ret'])
        if not valid6.empty:
            xpos = range(len(valid6))
            bars6 = ax6.bar(xpos, valid6['avg_ret'],
                            color=[GRN if v >= 0 else RED for v in valid6['avg_ret']],
                            width=0.65, alpha=0.85)
            ax6.axhline(0, color=COLORS['border'], lw=0.8, linestyle=':')
            ax6.set_xticks(xpos)
            ax6.set_xticklabels(valid6['bucket'], fontsize=8)
            ax6.set_ylabel('Retorno Medio (%)')
            for bar, (_, row) in zip(bars6, valid6.iterrows()):
                val = row['avg_ret']
                ypos = max(val, 0) + 0.05 if val >= 0 else min(val, 0) - 0.3
                ax6.text(bar.get_x() + bar.get_width()/2, ypos,
                         f"{val:+.2f}%", ha='center', va='bottom',
                         color=TXT, fontsize=7)

        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True, padx=12, pady=(6, 6))

        # ── Tabla resumen de buckets ───────────────────────────────────────
        tbl_frame = tk.Frame(win, bg=COLORS['surface'], padx=16, pady=8)
        tbl_frame.pack(fill=tk.X, padx=12, pady=(0, 12))

        tk.Label(tbl_frame, text='Resumen por Bucket de Score:',
                 bg=COLORS['surface'], fg=COLORS['text'],
                 font=('Segoe UI', 9, 'bold')).grid(row=0, column=0, columnspan=9, sticky='w')

        hdrs_tbl = ['Bucket', 'N(5d)', 'WR 5d', 'Ret.Medio 5d',
                    '│', 'N(20d)', 'WR 20d', 'Ret.Medio 20d', '']
        for c, h in enumerate(hdrs_tbl):
            tk.Label(tbl_frame, text=h, bg=COLORS['surface'],
                     fg=COLORS['text_dim'], font=('Consolas', 8, 'bold'),
                     width=12).grid(row=1, column=c, sticky='w')

        def fret(v):  return f"{v:+.2f}%" if v is not None else '—'
        def fwr(v):   return f"{v:.1f}%"  if v is not None else '—'

        for i, lbl in enumerate(labels):
            r5_row  = bkt5[bkt5['bucket'] == lbl].iloc[0]  if lbl in bkt5['bucket'].values  else None
            r20_row = bkt20[bkt20['bucket'] == lbl].iloc[0] if lbl in bkt20['bucket'].values else None
            row_num = i + 2

            # Color de fila según win rate 20d
            wr_val = r20_row['win_rate'] if r20_row is not None and r20_row['win_rate'] is not None else None
            fc = (COLORS['success'] if wr_val is not None and wr_val >= 60
                  else COLORS['danger'] if wr_val is not None and wr_val < 40
                  else COLORS['text'])

            vals = [
                lbl,
                str(int(r5_row['n']))  if r5_row  is not None else '—',
                fwr(r5_row['win_rate']  if r5_row  is not None else None),
                fret(r5_row['avg_ret']  if r5_row  is not None else None),
                '│',
                str(int(r20_row['n'])) if r20_row is not None else '—',
                fwr(r20_row['win_rate'] if r20_row is not None else None),
                fret(r20_row['avg_ret'] if r20_row is not None else None),
                '',
            ]
            for c, v in enumerate(vals):
                fg = fc if c in (0, 2, 6) else COLORS['text_dim']
                tk.Label(tbl_frame, text=v, bg=COLORS['surface'],
                         fg=fg, font=('Consolas', 8),
                         width=12).grid(row=row_num, column=c, sticky='w')

        # Nota metodológica
        nota = (
            "NOTA METODOLÓGICA: Esta validación compara el score actual vs retornos PASADOS (5d y 20d).\n"
            "Un r positivo y significativo confirma que el modelo identifica correctamente momentum y tendencias recientes.\n"
            "Para validación forward real, se necesitaría esperar 20 días tras cada análisis."
        )
        tk.Label(tbl_frame, text=nota, bg=COLORS['surface'], fg=COLORS['text_muted'],
                 font=('Consolas', 8), justify='left').grid(
                     row=len(labels)+2, column=0, columnspan=9, sticky='w', pady=(8, 0))

        tk.Button(win, text='Cerrar', bg=COLORS['surface3'], fg=COLORS['text'],
                  relief='flat', cursor='hand2',
                  command=lambda: [plt.close(fig), win.destroy()]).pack(pady=(0, 8))

    # ─── TAB 9: FUNDAMENTALES ────────────────

    def _build_fundamentals(self):
        p = self.tab_fund

        info_f = tk.Frame(p, bg=COLORS['surface'], pady=14, padx=20)
        info_f.pack(fill=tk.X, padx=24, pady=(16, 0))

        tk.Label(info_f, text='Análisis Fundamental — P/E · PEG · Deuda · Crecimiento',
                 bg=COLORS['surface'], fg=COLORS['primary'],
                 font=('Segoe UI', 11, 'bold')).pack(anchor='w')

        explanation = (
            "P/E  — Precio / Ganancias. < 15 = barato · > 40 = caro\n"
            "PEG  — P/E ajustado por crecimiento. < 1 = subvalorado vs su crecimiento\n"
            "D/E  — Deuda / Capital. < 30% = sólido · > 200% = riesgo alto\n"
            "Rev% — Crecimiento de ingresos anual. > 15% = empresa en expansión\n"
            "Mg%  — Margen de beneficio neto. > 20% = muy rentable\n"
            "Score — 0-100 combinando todo. >= 65 = FAVORABLE · < 40 = DESFAVORABLE"
        )
        tk.Label(info_f, text=explanation, bg=COLORS['surface'], fg=COLORS['text_dim'],
                 font=('Consolas', 9), justify='left').pack(anchor='w', pady=(6, 0))

        # Filtro
        flt = tk.Frame(p, bg=COLORS['bg'])
        flt.pack(fill=tk.X, padx=24, pady=(12, 8))

        tk.Label(flt, text='Señal:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)

        self.fund_filter = ttk.Combobox(flt, width=16, state='readonly', font=('Segoe UI', 9),
                                         values=['Todas', 'FAVORABLE', 'NEUTRO', 'DESFAVORABLE'])
        self.fund_filter.current(0)
        self.fund_filter.pack(side=tk.LEFT, padx=(6, 0))
        self.fund_filter.bind('<<ComboboxSelected>>', self._apply_fund_filter)

        # Tabla
        cols = ('symbol', 'sector', 'precio', 'fund_score', 'fund_signal',
                'pe', 'peg', 'debt', 'growth', 'margins', 'score_tec', 'confidence')
        hdrs = ('Símbolo', 'Sector', 'Precio', 'Score Fund.', 'Señal Fund.',
                'P/E (ETF=—)', 'PEG (ETF=—)', 'D/E% / Yield%', 'Rev% / Ret3y%', 'Marg% / Ret5y%',
                'Score Téc.', 'Confianza')

        self.fund_tree, scroll_f = self._make_treeview(p, cols, hdrs,
                                                        col_widths=[68, 148, 78, 88, 100,
                                                                    60, 60, 68, 68, 72, 80, 88])
        scroll_f.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 16))

    def _apply_fund_filter(self, event=None):
        if self.df_results is None or self.df_results.empty:
            return
        df = self.df_results.copy()

        flt = self.fund_filter.get()
        if flt in ('FAVORABLE', 'NEUTRO', 'DESFAVORABLE'):
            df = df[df['fund_signal'] == flt]

        df = df.sort_values('fund_score', ascending=False)

        tree = self.fund_tree
        tree.delete(*tree.get_children())

        for i, (_, row) in enumerate(df.iterrows()):
            tag = 'odd' if i % 2 == 0 else 'even'
            fs  = row.get('fund_signal', 'N/A')

            if fs == 'FAVORABLE':
                row_tag = (tag, 'buy')
            elif fs == 'DESFAVORABLE':
                row_tag = (tag, 'sell')
            else:
                row_tag = (tag,)

            def fmt(v, suffix='', decimals=1):
                if v is None or (isinstance(v, float) and v != v):
                    return '—'
                return f"{v:.{decimals}f}{suffix}"

            tree.insert('', 'end', tags=row_tag, values=(
                row['symbol'],
                row['sector'],
                f"${row['price']:.2f}",
                fmt(row.get('fund_score')),
                fs,
                fmt(row.get('fund_pe')),
                fmt(row.get('fund_peg'), decimals=2),
                fmt(row.get('fund_debt')),
                fmt(row.get('fund_growth'), suffix='%'),
                fmt(row.get('fund_margins'), suffix='%'),
                fmt(row.get('total_score')),
                row.get('confidence', '—'),
            ))

    # ─── TAB 6: PORTFOLIO ────────────────────

    def _build_portfolio(self):
        p = self.tab_portfolio

        top = tk.Frame(p, bg=COLORS['bg'])
        top.pack(fill=tk.X, padx=24, pady=(16, 8))

        tk.Label(top, text='Portfolio Diversificado Recomendado',
                 bg=COLORS['bg'], fg=COLORS['text'],
                 font=('Segoe UI', 13, 'bold')).pack(side=tk.LEFT)

        self.btn_rebuild = StyledButton(top, '↺  Recalcular',
                                        self._rebuild_portfolio, color=COLORS['primary'])
        self.btn_rebuild.pack(side=tk.RIGHT)
        self.btn_rebuild.set_enabled(False)

        # Descripción
        desc_f = tk.Frame(p, bg=COLORS['surface'], padx=16, pady=10)
        desc_f.pack(fill=tk.X, padx=24, pady=(0, 10))

        tk.Label(desc_f,
                 text='Selección basada en: Mayor score técnico · Baja correlación (≤70%) · '
                      'Diversificación sectorial · Señal de Volumen MA positiva preferida',
                 bg=COLORS['surface'], fg=COLORS['text_dim'],
                 font=('Segoe UI', 9), wraplength=900, justify='left').pack(anchor='w')

        # Tabla
        cols = ('rank', 'symbol', 'sector', 'score', 'precio', 'vol_signal',
                'vol_ratio', 'rsi', 'cambio5d', 'position_size')
        hdrs = ('#', 'Símbolo', 'Sector', 'Score', 'Precio',
                'Vol Señal', 'Vol Ratio', 'RSI', 'Δ 5d%', 'Tamaño Posición')

        self.port_tree, scroll_f = self._make_treeview(p, cols, hdrs,
                                                        col_widths=[35, 70, 160, 75, 80,
                                                                    90, 90, 55, 70, 140])
        scroll_f.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 16))

    # ─── TAB: ROTACIÓN SECTORIAL ─────────────

    def _build_sector_rotation(self):
        p = self.tab_sector

        tk.Label(p, text='Rotación Sectorial — Sectores Líderes vs Rezagados',
                 bg=COLORS['bg'], fg=COLORS['text'],
                 font=('Segoe UI', 14, 'bold')).pack(anchor='w', padx=24, pady=(20, 4))
        tk.Label(p, text='Score de rotación = % COMPRAR × confianza media × RS medio. '
                         'Verde = sector líder, rojo = sector rezagado.',
                 bg=COLORS['bg'], fg=COLORS['text_dim'],
                 font=('Segoe UI', 9)).pack(anchor='w', padx=24, pady=(0, 10))

        cols = ('rank', 'sector', 'n_sym', 'pct_buy', 'conf_avg',
                'rs_avg', 'sharpe_avg', 'score_rot', 'top_symbols')
        hdrs = ('#', 'Sector', 'Símbolos', '% COMPRAR', 'Confianza Media',
                'RS Medio', 'Sharpe Medio', 'Score Rotación', 'Top Símbolos')

        self.sector_tree, scroll_s = self._make_treeview(
            p, cols, hdrs,
            col_widths=[32, 185, 72, 82, 110, 80, 90, 110, 300])
        scroll_s.pack(fill=tk.BOTH, expand=True, padx=24, pady=(0, 16))

    def _populate_sector_rotation(self):
        if self.df_results is None or self.df_results.empty:
            return
        df = self.df_results.copy()
        tree = self.sector_tree
        tree.delete(*tree.get_children())

        conf_map = {'MUY ALTA': 5, 'ALTA': 4, 'MEDIA': 3, 'BAJA': 2, 'MUY BAJA': 1}

        sectors = []
        for sector, grp in df.groupby('sector'):
            n = len(grp)
            if n == 0:
                continue
            pct_buy  = (grp['vol_signal'] == 'COMPRAR').sum() / n * 100
            conf_num = grp['confidence'].map(conf_map).mean()
            rs_avg   = grp['rs_score'].mean() if 'rs_score' in grp else None
            sh_avg   = grp['sharpe'].mean()   if 'sharpe'   in grp else None

            # Score de rotación: % comprar (0-100) × confianza (1-5) × RS (0-100) / norm
            rs_v  = rs_avg  if rs_avg  is not None and not pd.isna(rs_avg)  else 50
            sh_v  = sh_avg  if sh_avg  is not None and not pd.isna(sh_avg)  else 0
            conf_v = conf_num if not pd.isna(conf_num) else 3
            score = (pct_buy * 0.40) + (conf_v / 5 * 100 * 0.35) + (rs_v * 0.25)

            # Top 3 símbolos del sector por confidence_score
            top3 = grp.nlargest(3, 'confidence_score')['symbol'].tolist() \
                   if 'confidence_score' in grp else []

            sectors.append({
                'sector': sector, 'n': n, 'pct_buy': pct_buy,
                'conf_avg': conf_v, 'rs_avg': rs_v, 'sharpe_avg': sh_v,
                'score': score, 'top3': top3,
            })

        sectors.sort(key=lambda x: x['score'], reverse=True)

        for i, s in enumerate(sectors, 1):
            if s['score'] >= 65:   tag = 'buy'
            elif s['score'] >= 40: tag = 'neutral'
            else:                   tag = 'sell'

            sh = s['sharpe_avg']
            tree.insert('', 'end', tags=(tag,), values=(
                i,
                s['sector'],
                s['n'],
                f"{s['pct_buy']:.0f}%",
                f"{s['conf_avg']:.1f}/5",
                f"{s['rs_avg']:.1f}",
                f"{sh:.2f}" if sh and not pd.isna(sh) else '—',
                f"{s['score']:.1f}",
                ', '.join(s['top3']),
            ))

    # ─── TAB: ANÁLISIS IA (Claude) ───────────

    def _build_ia_tab(self):
        p = self.tab_ia

        # ── Cabecera con estado de disponibilidad ────────────────────────
        hdr = tk.Frame(p, bg=COLORS['surface'], padx=24, pady=14)
        hdr.pack(fill=tk.X)

        ia_col = COLORS['success'] if self._claude_available else COLORS['text_dim']
        ia_icon = '🤖' if self._claude_available else '⚫'
        tk.Label(hdr, text=f'{ia_icon}  Análisis Cualitativo con Claude IA',
                 bg=COLORS['surface'], fg=COLORS['text'],
                 font=('Segoe UI', 13, 'bold')).pack(side=tk.LEFT)

        mode_f = tk.Frame(hdr, bg=COLORS['surface2'], padx=10, pady=4)
        mode_f.pack(side=tk.LEFT, padx=20)
        tk.Label(mode_f, text=f"Modo: {self._claude_mode_label}",
                 bg=COLORS['surface2'], fg=ia_col,
                 font=('Segoe UI', 9, 'bold')).pack()

        # ── Explicación ──────────────────────────────────────────────────
        desc_f = tk.Frame(p, bg=COLORS['bg'], padx=24, pady=10)
        desc_f.pack(fill=tk.X)

        desc_text = (
            "Claude IA complementa el análisis técnico con perspectiva cualitativa:\n"
            "  • Tesis de inversión — por qué la empresa puede crecer a mediano plazo\n"
            "  • Riesgos cualitativos — competencia, regulación, deuda, management\n"
            "  • Catalizadores — nuevos productos, expansión, macro favorable\n"
            "  • Coherencia técnica — si el score técnico refleja la realidad del negocio\n\n"
            "Nota: El análisis se basa en el conocimiento de entrenamiento de Claude (hasta ago 2025). "
            "Para noticias muy recientes, contrasta con fuentes actualizadas."
        )
        if not self._claude_available:
            desc_text += (
                "\n\n⚠️  Claude IA no detectado. Para activarlo:\n"
                "  Opción A) pip install anthropic  +  set ANTHROPIC_API_KEY=sk-ant-...\n"
                "  Opción B) Instalar Claude CLI: https://claude.ai/code"
            )

        tk.Label(desc_f, text=desc_text, bg=COLORS['bg'], fg=COLORS['text_dim'],
                 font=('Segoe UI', 9), justify='left', anchor='w').pack(anchor='w')

        # ── Controles ────────────────────────────────────────────────────
        ctrl = tk.Frame(p, bg=COLORS['bg'])
        ctrl.pack(fill=tk.X, padx=24, pady=(0, 8))

        self.ia_top_n = tk.IntVar(value=10)
        tk.Label(ctrl, text='Analizar Top:', bg=COLORS['bg'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT)
        for n in (5, 10, 15):
            tk.Radiobutton(ctrl, text=str(n), variable=self.ia_top_n, value=n,
                           bg=COLORS['bg'], fg=COLORS['text'],
                           selectcolor=COLORS['surface2'],
                           activebackground=COLORS['bg']).pack(side=tk.LEFT, padx=4)

        self.ia_btn = tk.Button(
            ctrl,
            text='🤖 Analizar con Claude IA',
            bg=COLORS['success'] if self._claude_available else COLORS['surface3'],
            fg='white' if self._claude_available else COLORS['text_dim'],
            font=('Segoe UI', 10, 'bold'),
            relief='flat', padx=14, pady=5,
            cursor='hand2' if self._claude_available else 'arrow',
            state='normal' if self._claude_available else 'disabled',
            command=self._run_claude_analysis,
        )
        self.ia_btn.pack(side=tk.LEFT, padx=(20, 0))

        self.ia_status_lbl = tk.Label(ctrl, text='',
                                       bg=COLORS['bg'], fg=COLORS['text_dim'],
                                       font=('Segoe UI', 9))
        self.ia_status_lbl.pack(side=tk.LEFT, padx=12)

        # ── Tabla de resultados ──────────────────────────────────────────
        cols = ('sym', 'sector', 'ia_rating', 'ia_conf', 'tech_score',
                'ann_ret', 'coherence', 'thesis')
        hdrs = ('Símbolo', 'Sector', 'Rating IA', 'Conf.IA', 'Score Bot',
                'Ret.Anual', 'Coherencia', 'Tesis (doble-click para detalle)')
        self.ia_tree, ia_sf = self._make_treeview(
            p, cols, hdrs,
            col_widths=[72, 130, 90, 65, 72, 72, 88, 420],
            height=12
        )
        ia_sf.pack(fill=tk.BOTH, expand=True, padx=24, pady=(4, 0))
        self.ia_tree.bind('<Double-1>', self._ia_detail_popup)

        # ── Área de texto para detalles ──────────────────────────────────
        detail_lbl = tk.Label(p, text='DETALLE DEL ANÁLISIS IA',
                               bg=COLORS['bg'], fg=COLORS['text_dim'],
                               font=('Segoe UI', 9, 'bold'))
        detail_lbl.pack(anchor='w', padx=24, pady=(12, 4))

        self.ia_detail_box = scrolledtext.ScrolledText(
            p, height=8, wrap=tk.WORD,
            bg=COLORS['surface'], fg=COLORS['text'],
            font=('Consolas', 9), state='disabled',
            relief='flat', padx=10, pady=6
        )
        self.ia_detail_box.pack(fill=tk.X, padx=24, pady=(0, 12))

    def _run_claude_analysis(self):
        """Lanza el análisis Claude en background thread."""
        if self._ia_running:
            return
        if self.df_results is None or self.df_results.empty:
            messagebox.showinfo('Sin datos',
                                'Ejecuta el análisis técnico primero.')
            return
        if not self._claude_available or self._claude_analyzer is None:
            messagebox.showerror('Claude no disponible',
                                 'Configura ANTHROPIC_API_KEY o instala Claude CLI.')
            return

        self._ia_running = True
        self.ia_btn.config(state='disabled', text='⏳ Consultando a Claude...')
        self.ia_status_lbl.config(text='Enviando datos a Claude IA...',
                                   fg=COLORS['warning'])

        top_n = self.ia_top_n.get()

        def _worker():
            try:
                # Añadir _buy_score si no existe
                df = self.df_results.copy()
                if '_buy_score' not in df.columns:
                    df['_buy_score'] = df.apply(self._compute_buy_score, axis=1)
                results = self._claude_analyzer.analyze_top_stocks(
                    df, top_n=top_n, score_col='_buy_score', only_buy=True
                )
                self.q.put(('ia_result', results))
                self._q_notify()
            except Exception as e:
                self.q.put(('ia_error', str(e)))
                self._q_notify()

        import threading
        threading.Thread(target=_worker, daemon=True).start()

    def _populate_ia_table(self, results: list):
        """Llena la tabla IA con los resultados de Claude."""
        t = self.ia_tree
        t.delete(*t.get_children())

        for item in results:
            sym      = item.get('symbol', '')
            rating   = item.get('ia_rating', '—')
            conf_ia  = item.get('ia_confidence', '—')
            thesis   = item.get('thesis', '—')
            coherence = item.get('tech_coherence', '—')

            # Datos técnicos del df_results
            tech_score = '—'
            ann_ret    = '—'
            sector     = '—'
            if self.df_results is not None and not self.df_results.empty:
                rows = self.df_results[self.df_results['symbol'] == sym]
                if not rows.empty:
                    r = rows.iloc[0]
                    tech_score = f"{r.get('_buy_score', r.get('total_score', 0)):.0f}"
                    ar = r.get('ann_return')
                    ann_ret = f"{ar:+.1f}%" if ar is not None else '—'
                    sector  = r.get('sector', '—')

            tag = ('buy'  if rating == 'COMPRAR' else
                   'sell' if rating == 'EVITAR'  else 'neutral')
            t.insert('', 'end', tags=(tag,), values=(
                sym, sector, rating,
                str(conf_ia) if conf_ia != '—' else '—',
                tech_score, ann_ret,
                coherence,
                thesis[:120] + '…' if len(str(thesis)) > 120 else thesis,
            ))

        # Resumen
        n_buy    = sum(1 for r in results if r.get('ia_rating') == 'COMPRAR')
        n_avoid  = sum(1 for r in results if r.get('ia_rating') == 'EVITAR')
        self.ia_status_lbl.config(
            text=f"✓ {len(results)} analizados — IA recomienda: {n_buy} COMPRAR · {n_avoid} EVITAR",
            fg=COLORS['success']
        )

    def _ia_detail_popup(self, event=None):
        """Muestra el análisis IA completo de la acción seleccionada."""
        sel = self.ia_tree.selection()
        if not sel:
            return
        sym = str(self.ia_tree.item(sel[0], 'values')[0]).strip()

        # Buscar en los resultados
        result = next((r for r in self._ia_results if r.get('symbol') == sym), None)
        if not result:
            return

        # Formatear detalle
        rating    = result.get('ia_rating', '—')
        conf      = result.get('ia_confidence', '—')
        thesis    = result.get('thesis', '—')
        risks     = result.get('risks', [])
        catalysts = result.get('catalysts', [])
        macro     = result.get('macro_context', '—')
        coherence = result.get('tech_coherence', '—')
        coh_reason = result.get('tech_coherence_reason', '—')

        lines = [
            f"{'='*50}",
            f"  {sym} — Rating IA: {rating}  (Confianza: {conf}/5)",
            f"{'='*50}",
            f"\nTESIS DE INVERSIÓN:",
            f"  {thesis}",
            f"\nCONTEXTO MACRO/SECTORIAL:",
            f"  {macro}",
            f"\nRIESGOS:",
        ] + [f"  ⚠  {r}" for r in risks] + [
            f"\nCATALIZADORES:",
        ] + [f"  ✓  {c}" for c in catalysts] + [
            f"\nCOHERENCIA CON ANÁLISIS TÉCNICO: {coherence}",
            f"  {coh_reason}",
        ]

        # Mostrar en caja de detalle
        self.ia_detail_box.config(state='normal')
        self.ia_detail_box.delete('1.0', 'end')
        self.ia_detail_box.insert('end', '\n'.join(lines))
        self.ia_detail_box.config(state='disabled')

        # También mostrar en el tab IA para que el usuario lo vea
        self.nb.select(self.tab_ia)

    # ─── TAB 5: CONFIGURACIÓN ────────────────

    def _build_settings(self):
        p = self.tab_settings

        card = tk.Frame(p, bg=COLORS['surface'], padx=30, pady=24)
        card.pack(padx=40, pady=30, fill=tk.X)

        tk.Label(card, text='Parámetros de Análisis', bg=COLORS['surface'],
                 fg=COLORS['text'], font=('Segoe UI', 13, 'bold')).grid(
                 row=0, column=0, columnspan=2, sticky='w', pady=(0, 20))

        def row(r, label, widget):
            tk.Label(card, text=label, bg=COLORS['surface'],
                     fg=COLORS['text_dim'], font=('Segoe UI', 10)).grid(
                     row=r, column=0, sticky='w', pady=8, padx=(0, 20))
            widget.grid(row=r, column=1, sticky='w')

        # Portfolio N
        self._cfg_n = tk.StringVar(value=str(self.cfg['portfolio_n']))
        row(1, 'Acciones en portfolio:', tk.Entry(card, textvariable=self._cfg_n,
            bg=COLORS['surface2'], fg=COLORS['text'], insertbackground=COLORS['text'],
            relief='flat', bd=4, width=8, font=('Segoe UI', 10)))

        # Max corr
        self._cfg_corr = tk.StringVar(value=str(self.cfg['max_corr']))
        row(2, 'Correlación máxima:', tk.Entry(card, textvariable=self._cfg_corr,
            bg=COLORS['surface2'], fg=COLORS['text'], insertbackground=COLORS['text'],
            relief='flat', bd=4, width=8, font=('Segoe UI', 10)))

        # Force download
        self._cfg_force = tk.BooleanVar(value=self.cfg['force_download'])
        cb = tk.Checkbutton(card, variable=self._cfg_force,
                            bg=COLORS['surface'], fg=COLORS['text'],
                            selectcolor=COLORS['surface2'],
                            activebackground=COLORS['surface'],
                            font=('Segoe UI', 10), text='Forzar re-descarga de datos')
        row(3, 'Datos históricos:', cb)

        # Notificaciones de escritorio
        self._notif_enabled = tk.BooleanVar(value=True)
        notif_cb = tk.Checkbutton(card, variable=self._notif_enabled,
                                   bg=COLORS['surface'], fg=COLORS['text'],
                                   selectcolor=COLORS['surface2'],
                                   activebackground=COLORS['surface'],
                                   font=('Segoe UI', 10),
                                   text='Enviar notificaciones de escritorio al terminar análisis')
        row(4, 'Notificaciones:', notif_cb)

        # Botón guardar
        StyledButton(card, '  Guardar configuración  ',
                     self._save_settings, color=COLORS['primary']).grid(
                     row=5, column=0, columnspan=2, sticky='w', pady=(20, 0))

        # ── Auto-refresh scheduler ──────────────────────────────────
        sched_card = tk.Frame(p, bg=COLORS['surface'], padx=30, pady=20)
        sched_card.pack(padx=40, pady=(0, 10), fill=tk.X)

        tk.Label(sched_card, text='Auto-Refresh Diario',
                 bg=COLORS['surface'], fg=COLORS['text'],
                 font=('Segoe UI', 11, 'bold')).grid(
                 row=0, column=0, columnspan=3, sticky='w', pady=(0, 12))

        self._sched_enabled = tk.BooleanVar(value=False)
        tk.Checkbutton(sched_card, variable=self._sched_enabled,
                       text='Activar análisis automático',
                       bg=COLORS['surface'], fg=COLORS['text'],
                       selectcolor=COLORS['surface2'],
                       activebackground=COLORS['surface'],
                       font=('Segoe UI', 10),
                       command=self._schedule_toggle).grid(
                       row=1, column=0, sticky='w', pady=4)

        tk.Label(sched_card, text='Hora (HH:MM):',
                 bg=COLORS['surface'], fg=COLORS['text_dim'],
                 font=('Segoe UI', 10)).grid(row=2, column=0, sticky='w', pady=4)

        self._sched_time = tk.StringVar(value='17:00')
        tk.Entry(sched_card, textvariable=self._sched_time, width=8,
                 bg=COLORS['surface2'], fg=COLORS['text'],
                 insertbackground=COLORS['text'],
                 relief='flat', bd=4, font=('Segoe UI', 10)).grid(
                 row=2, column=1, sticky='w', padx=(8, 16))

        self._sched_label = tk.Label(sched_card, text='Scheduler inactivo',
                                      bg=COLORS['surface'], fg=COLORS['text_muted'],
                                      font=('Segoe UI', 9))
        self._sched_label.grid(row=2, column=2, sticky='w')

        # Info universo
        info = tk.Frame(p, bg=COLORS['surface'], padx=30, pady=20)
        info.pack(padx=40, fill=tk.X)

        tk.Label(info, text='Universo de Acciones (N26-compatible)',
                 bg=COLORS['surface'], fg=COLORS['text'],
                 font=('Segoe UI', 11, 'bold')).pack(anchor='w', pady=(0, 10))

        try:
            from stock_universe import STOCK_UNIVERSE
            for sector, stocks in STOCK_UNIVERSE.items():
                line = f"{sector:<25} ({len(stocks):3} acciones)  {', '.join(stocks[:8])}{'...' if len(stocks) > 8 else ''}"
                tk.Label(info, text=line, bg=COLORS['surface'], fg=COLORS['text_dim'],
                         font=('Consolas', 8)).pack(anchor='w')
            total = sum(len(s) for s in STOCK_UNIVERSE.values())
            tk.Label(info, text=f"\nTotal: {total} acciones",
                     bg=COLORS['surface'], fg=COLORS['primary'],
                     font=('Segoe UI', 10, 'bold')).pack(anchor='w')
        except Exception as e:
            tk.Label(info, text=f'Error cargando universo: {e}',
                     bg=COLORS['surface'], fg=COLORS['danger'],
                     font=('Segoe UI', 9)).pack(anchor='w')

    # ─────────────────────────────────────────────
    #  MEJORA 2: NOTIFICACIONES DE ESCRITORIO
    # ─────────────────────────────────────────────

    def _send_desktop_notification(self, title, msg):
        """
        Envía notificación de escritorio Windows mediante PowerShell + System.Windows.Forms.
        No requiere librerías adicionales — funciona en Windows 10/11.
        """
        if self._notif_enabled is not None and not self._notif_enabled.get():
            return
        try:
            import subprocess
            # Escapar comillas simples para PowerShell
            safe_title = title.replace("'", "''")
            safe_msg   = msg.replace("'", "''")
            ps_script = (
                "Add-Type -AssemblyName System.Windows.Forms; "
                "$n = New-Object System.Windows.Forms.NotifyIcon; "
                "$n.Icon = [System.Drawing.SystemIcons]::Information; "
                f"$n.BalloonTipTitle = '{safe_title}'; "
                f"$n.BalloonTipText  = '{safe_msg}'; "
                "$n.Visible = $true; "
                "$n.ShowBalloonTip(7000); "
                "Start-Sleep -Seconds 8; "
                "$n.Dispose()"
            )
            subprocess.Popen(
                ['powershell', '-WindowStyle', 'Hidden', '-Command', ps_script],
                creationflags=0x08000000,   # CREATE_NO_WINDOW
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
        except Exception:
            pass   # Notificaciones son opcionales, nunca crashean la app

    # ─── STATUS BAR ──────────────────────────

    def _build_statusbar(self):
        bar = tk.Frame(self.root, bg=COLORS['surface'], height=26)
        bar.pack(fill=tk.X, side=tk.BOTTOM)
        bar.pack_propagate(False)

        self.status_label = tk.Label(bar, text='Listo', bg=COLORS['surface'],
                                     fg=COLORS['text_dim'], font=('Segoe UI', 8),
                                     padx=16)
        self.status_label.pack(side=tk.LEFT, pady=4)

        self.time_label = tk.Label(bar, text='', bg=COLORS['surface'],
                                   fg=COLORS['text_muted'], font=('Segoe UI', 8),
                                   padx=16)
        self.time_label.pack(side=tk.RIGHT, pady=4)
        self._tick()

    def _tick(self):
        self.time_label.config(text=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        self.root.after(1000, self._tick)

    # ─── HELPER: CREAR TREEVIEW ──────────────

    def _make_treeview(self, parent, cols, hdrs, col_widths=None, height=None):
        frame = tk.Frame(parent, bg=COLORS['bg'])

        kw = {'columns': cols, 'show': 'headings', 'selectmode': 'browse'}
        if height is not None:
            kw['height'] = height
        tree = ttk.Treeview(frame, **kw)

        vsb = ttk.Scrollbar(frame, orient='vertical', command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient='horizontal', command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        tree.pack(fill=tk.BOTH, expand=True)

        for i, (col, hdr) in enumerate(zip(cols, hdrs)):
            w = col_widths[i] if col_widths and i < len(col_widths) else 100
            tree.heading(col, text=hdr, anchor='center')
            tree.column(col, width=w, anchor='center', minwidth=30)

        # Zebra rows
        tree.tag_configure('odd',  background=COLORS['surface'])
        tree.tag_configure('even', background=COLORS['surface2'])
        tree.tag_configure('buy',  foreground=COLORS['buy'])
        tree.tag_configure('sell', foreground=COLORS['sell'])
        tree.tag_configure('warn', foreground=COLORS['warning'])

        return tree, frame

    # ─── LÓGICA DE ANÁLISIS ──────────────────

    def _start_analysis(self):
        if self.running:
            return
        self.running = True
        self.btn_run.set_enabled(False)
        self.btn_export.set_enabled(False)
        self.btn_rebuild.set_enabled(False)
        self.hdr_status.config(text='● ANALIZANDO', fg=COLORS['warning'])
        self.progress_bar['value'] = 0
        self._log_clear()
        self._log('info', 'Iniciando análisis completo...')
        threading.Thread(target=self._run_analysis, daemon=True).start()

    def _stop_analysis(self):
        self.running = False
        self._log('warn', 'Análisis detenido por el usuario.')
        self.hdr_status.config(text='● DETENIDO', fg=COLORS['danger'])
        self.btn_run.set_enabled(True)

    def _run_analysis(self):
        try:
            self._progress(2, 'Cargando módulos...')

            from database import TradingDatabase
            from stock_analyzer import StockAnalyzer
            from stock_universe import get_all_symbols

            self.db = TradingDatabase()
            self.analyzer = StockAnalyzer(self.db)

            symbols = get_all_symbols()
            total = len(symbols)
            self._progress(5, f'Universo cargado: {total} acciones')
            self._log('ok', f'Universo: {total} acciones N26-compatible')

            # VIX
            self._progress(6, 'Obteniendo VIX...')
            self._vix_data = self.analyzer.get_vix_level()
            if self._vix_data:
                self._log('info', f"VIX: {self._vix_data['vix']} — {self._vix_data['level']} → {self._vix_data['signal']}")
            else:
                self._log('warn', 'VIX no disponible')

            # SPY (referencia de fuerza relativa)
            self._progress(8, 'Obteniendo SPY (referencia RS)...')
            self._spy_data = self.analyzer._get_spy_returns()
            if self._spy_data:
                r60  = self._spy_data.get('ret_60d')
                r252 = self._spy_data.get('ret_252d')
                self._log('info', f"SPY 60d: {r60:+.1f}% | 252d: {r252:+.1f}%")
            else:
                self._log('warn', 'SPY no disponible')

            # ── Descarga de datos (incremental) ──
            if self.cfg['force_download']:
                import yfinance as yf
                from data_downloader import DataDownloader
                from datetime import timedelta
                dl = DataDownloader(self.db)
                today_str = datetime.now().strftime('%Y-%m-%d')
                success = 0
                skipped = 0
                failed  = []

                self._log('info', 'Descargando datos históricos (modo incremental)...')

                for i, symbol in enumerate(symbols, 1):
                    if not self.running:
                        break
                    try:
                        last_date = self.db.get_last_date(symbol)
                        if last_date is not None:
                            if last_date >= today_str:
                                # Datos ya al día — omitir descarga
                                skipped += 1
                            else:
                                # Descarga incremental desde el día siguiente
                                start_dt = (
                                    datetime.strptime(last_date, '%Y-%m-%d') + timedelta(days=1)
                                ).strftime('%Y-%m-%d')
                                df_new = yf.download(
                                    symbol, start=start_dt, interval='1d',
                                    progress=False, auto_adjust=True
                                )
                                if df_new is not None and not df_new.empty:
                                    if isinstance(df_new.columns, pd.MultiIndex):
                                        df_new.columns = [col[0] for col in df_new.columns]
                                    df_new.columns = [c.lower() for c in df_new.columns]
                                    df_new = df_new.dropna()
                                    if not df_new.empty:
                                        self.db.append_precios(df_new, symbol)
                                success += 1
                        else:
                            # Sin datos previos — descarga completa
                            df = dl.get_max_historical_data(symbol, interval='1d', force_download=True)
                            if df is not None and len(df) > 0:
                                success += 1
                            else:
                                failed.append(symbol)
                    except Exception:
                        failed.append(symbol)
                    pct = 5 + int((i / total) * 40)
                    self._progress(pct, f'Actualizando [{i}/{total}] {symbol}')

                self._log('ok', f'Descarga: {success} actualizadas, {skipped} al día, '
                                f'{len(failed)} fallidas')
                if failed:
                    self._log('warn', f'Fallidos: {", ".join(failed[:10])}'
                                      f'{"..." if len(failed) > 10 else ""}')
            else:
                self._log('dim', 'Usando datos existentes en base de datos')

            if not self.running:
                return

            # ── Análisis técnico ──
            self._progress(50, 'Calculando indicadores técnicos...')
            self._log('info', 'Analizando indicadores técnicos + Volumen MA...')

            self.df_results = self._analyze_with_progress(symbols, total)

            if self.df_results is None or self.df_results.empty:
                self._log('err', 'Error: no se pudieron analizar acciones suficientes')
                self.running = False
                self.q.put(('done_error', None))
                self._q_notify()
                return

            # ── Portfolio ──
            self._progress(90, 'Construyendo portfolio diversificado...')
            self._log('info', 'Calculando correlaciones y seleccionando portfolio...')

            n = self.cfg['portfolio_n']
            max_corr = self.cfg['max_corr']
            self.portfolio = self.analyzer.find_diversified_portfolio(
                self.df_results, n=n, max_corr=max_corr)

            self._log('ok', f'Portfolio: {len(self.portfolio)} acciones seleccionadas')

            # ── Actualizar DataController (4.2) ──
            try:
                regime = getattr(self.analyzer, '_current_regime', 'transition')
                self.ctrl.update(
                    self.df_results,
                    portfolio=self.portfolio,
                    vix=self._vix_data,
                    spy=self._spy_data,
                    regime=regime,
                )
            except Exception:
                pass

            # ── Alertas de cambio de señal + alertas inteligentes ──
            self._progress(97, 'Detectando cambios de señal...')
            alerts        = self._detect_signal_changes(self.df_results)
            smart_alerts  = self._detect_smart_alerts(self.df_results)
            self.db.save_signal_snapshot(self.df_results)
            self.db.cleanup_old_snapshots(keep=10)
            all_alerts = smart_alerts + alerts   # criticos primero
            if all_alerts:
                self.q.put(('alerts', all_alerts))
                self._q_notify()

            self._progress(100, 'Análisis completado')
            self.q.put(('done_ok', self.df_results))
            self._q_notify()

        except Exception as e:
            import traceback
            self._log('err', f'Error inesperado: {e}')
            self._log('err', traceback.format_exc())
            self.running = False
            self.q.put(('done_error', None))
            self._q_notify()

    def _analyze_with_progress(self, symbols, total):
        """Analiza acciones en paralelo con ThreadPoolExecutor."""
        from stock_universe import get_sector
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import threading

        # ── 1) Pre-cargar todos los DataFrames desde DB (secuencial, evita
        #        problemas de concurrencia con SQLite) ──────────────────────
        self._progress(50, 'Cargando datos de base de datos...')
        symbol_dfs = {}
        for symbol in symbols:
            df = self.db.get_precios(symbol)
            if not df.empty and len(df) >= 100:
                symbol_dfs[symbol] = df

        valid_symbols = list(symbol_dfs.keys())
        n_valid = len(valid_symbols)
        self._log('info', f'{n_valid} símbolos con datos suficientes — analizando en paralelo...')

        # ── 2) Contador thread-safe para progreso ──────────────────────────
        completed_count = [0]
        lock = threading.Lock()

        # ── 3) Función por símbolo ─────────────────────────────────────────
        def analyze_one(symbol):
            if not self.running:
                return None
            try:
                df = symbol_dfs[symbol]
                last_date, freshness = self.analyzer.check_data_freshness(df)
                analysis = self.analyzer.calculate_technical_score(df, symbol)
                if analysis is None:
                    return None

                vol_signal    = self.analyzer.calculate_volume_ma_signal(df)
                candle_signal = self.analyzer.calculate_candlestick_signal(df)
                adx_signal    = self.analyzer.calculate_adx_signal(df)
                fund_data     = self.analyzer.calculate_fundamentals(symbol)   # I/O red
                tf_data       = self.analyzer.calculate_timeframe_alignment(df)
                rs_data       = self.analyzer.calculate_relative_strength(df, self._spy_data)
                risk_data     = self.analyzer.calculate_risk_metrics(df)
                bt_data       = self.analyzer.calculate_backtest(df)

                # ── Dip metrics + retornos mediano plazo ──────────────────
                try:
                    df_s = df.sort_values('timestamp')
                    closes = df_s['close']
                    last_close = float(closes.iloc[-1])
                    high_15d = float(closes.tail(15).max())
                    dip_from_high_pct = (high_15d - last_close) / high_15d * 100 if high_15d > 0 else None
                    dip_3d = ((last_close - float(closes.iloc[-4])) / float(closes.iloc[-4]) * 100
                              if len(closes) >= 4 else None)
                    daily_rets = closes.tail(22).pct_change(fill_method=None).dropna()
                    avg_daily_vol = float(daily_rets.std() * 100) if len(daily_rets) >= 5 else None
                    # Cuántas veces el ATR diario vale la caída de 3 días
                    dip_atr_multiple = (abs(dip_3d) / avg_daily_vol
                                        if dip_3d is not None and avg_daily_vol and avg_daily_vol > 0
                                        else None)
                    # Retornos mediano-largo plazo
                    close_60d  = float(closes.iloc[-61])  if len(closes) >= 61  else None
                    close_252d = float(closes.iloc[-253]) if len(closes) >= 253 else None
                    change_60d  = (last_close - close_60d)  / close_60d  * 100 if close_60d  else None
                    change_252d = (last_close - close_252d) / close_252d * 100 if close_252d else None
                except Exception:
                    dip_from_high_pct = None
                    dip_3d            = None
                    dip_atr_multiple  = None
                    change_60d        = None
                    change_252d       = None

                # ── Predicción: tendencia + patrones históricos ───────────
                try:
                    cl = df.sort_values('timestamp')['close'].values.astype(float)
                    n_cl = len(cl)

                    # 1. Regresión lineal sobre últimos 30 días → proyección 10d
                    if n_cl >= 30:
                        xr = np.arange(30, dtype=float)
                        yr = cl[-30:]
                        coeffs = np.polyfit(xr, yr, 1)
                        slope_pct_day = coeffs[0] / yr[-1] * 100
                        pred_trend_10d = round(slope_pct_day * 10, 2)
                        fitted = np.polyval(coeffs, xr)
                        ss_res = np.sum((yr - fitted) ** 2)
                        ss_tot = np.sum((yr - yr.mean()) ** 2)
                        pred_r2 = round(1 - ss_res / ss_tot, 3) if ss_tot > 0 else 0.0
                    else:
                        pred_trend_10d = None
                        pred_r2 = None

                    # 2. Reconocimiento de patrones: compara últimos 10 días
                    #    con todas las ventanas históricas; mira qué pasó 10d después
                    WIN = 10
                    FWD = 10
                    MIN_CORR = 0.70
                    if n_cl >= WIN + FWD + 20:
                        cur_rets = np.diff(cl[-WIN:]) / cl[-WIN:-1]
                        cur_norm = cur_rets - cur_rets.mean()
                        cur_std  = cur_rets.std()
                        fwd_rets = []
                        for i in range(n_cl - WIN - FWD):
                            h_rets = np.diff(cl[i:i+WIN]) / cl[i:i+WIN-1]
                            h_norm = h_rets - h_rets.mean()
                            h_std  = h_rets.std()
                            if cur_std > 0 and h_std > 0:
                                corr = float(np.dot(cur_norm, h_norm) /
                                             (len(cur_norm) * cur_std * h_std))
                                if corr >= MIN_CORR:
                                    fwd = (cl[i + WIN + FWD - 1] - cl[i + WIN]) / cl[i + WIN] * 100
                                    fwd_rets.append(fwd)
                        if fwd_rets:
                            arr = np.array(fwd_rets)
                            pred_pat_wr  = round(float((arr > 0).mean() * 100), 1)
                            pred_pat_avg = round(float(arr.mean()), 2)
                            pred_pat_n   = len(arr)
                        else:
                            pred_pat_wr  = None
                            pred_pat_avg = None
                            pred_pat_n   = 0
                    else:
                        pred_pat_wr  = None
                        pred_pat_avg = None
                        pred_pat_n   = 0

                    # 3. Z-score: distancia del precio actual vs media 20d
                    if n_cl >= 20:
                        ma20 = cl[-20:].mean()
                        sd20 = cl[-20:].std()
                        pred_zscore = round(float((cl[-1] - ma20) / sd20), 2) if sd20 > 0 else 0.0
                    else:
                        pred_zscore = None

                    # 4. Señal combinada
                    up, dn = 0, 0
                    if pred_trend_10d is not None:
                        if pred_trend_10d > 1.5 and (pred_r2 or 0) > 0.35: up += 2
                        elif pred_trend_10d > 0.5:                          up += 1
                        elif pred_trend_10d < -1.5 and (pred_r2 or 0) > 0.35: dn += 2
                        elif pred_trend_10d < -0.5:                            dn += 1
                    if pred_pat_wr is not None:
                        if pred_pat_wr >= 65:  up += 2
                        elif pred_pat_wr >= 55: up += 1
                        elif pred_pat_wr <= 35: dn += 2
                        elif pred_pat_wr <= 45: dn += 1
                    if pred_zscore is not None:
                        if pred_zscore <= -1.5:  up += 1   # oversold → rebote esperado
                        elif pred_zscore >= 1.5: dn += 1   # overbought → corrección esperada

                    if   up >= 3 and dn == 0: pred_signal = 'SUBE FUERTE'
                    elif up >= 2 and dn <= 1: pred_signal = 'SUBE'
                    elif dn >= 3 and up == 0: pred_signal = 'BAJA FUERTE'
                    elif dn >= 2 and up <= 1: pred_signal = 'BAJA'
                    else:                     pred_signal = 'NEUTRO'

                except Exception:
                    pred_trend_10d = None
                    pred_r2        = None
                    pred_pat_wr    = None
                    pred_pat_avg   = None
                    pred_pat_n     = 0
                    pred_zscore    = None
                    pred_signal    = 'NEUTRO'

                result = {
                    'symbol': symbol,
                    'sector': get_sector(symbol),
                    'total_score': analysis['total_score'],
                    'price': analysis['price'],
                    'change_5d': analysis['change_5d'],
                    'change_20d': analysis['change_20d'],
                    'rsi': analysis['rsi'],
                    'support_52w': analysis['support_52w'],
                    'resistance_52w': analysis['resistance_52w'],
                    'last_update': last_date.strftime('%Y-%m-%d') if last_date else 'N/A',
                    'data_freshness': freshness,
                    # Volumen MA
                    'vol_signal':    vol_signal['signal']           if vol_signal else 'N/A',
                    'vol_ratio':     vol_signal['ratio']            if vol_signal else None,
                    'vol_strength':  vol_signal['signal_strength']  if vol_signal else None,
                    'vol_ma7':       vol_signal['ma7']              if vol_signal else None,
                    'vol_ma60':      vol_signal['ma60']             if vol_signal else None,
                    'vol_ma7_trend': vol_signal['ma7_trend_pct']    if vol_signal else None,
                    # Velas
                    'candle_pattern':  candle_signal['pattern']  if candle_signal else 'Sin patrón',
                    'candle_signal':   candle_signal['signal']   if candle_signal else 'NEUTRO',
                    'candle_strength': candle_signal['strength'] if candle_signal else '—',
                    # ADX
                    'adx_value':     adx_signal['adx']       if adx_signal else None,
                    'adx_strength':  adx_signal['strength']  if adx_signal else '—',
                    'adx_direction': adx_signal['direction'] if adx_signal else '—',
                    'adx_plus_di':   adx_signal['plus_di']   if adx_signal else None,
                    'adx_minus_di':  adx_signal['minus_di']  if adx_signal else None,
                    # Fundamentales
                    'fund_score':   fund_data['fund_score']  if fund_data else 50,
                    'fund_signal':  fund_data['fund_signal'] if fund_data else 'N/A',
                    'fund_pe':      fund_data['pe']          if fund_data else None,
                    'fund_peg':     fund_data['peg']         if fund_data else None,
                    'fund_debt':    fund_data['debt_equity'] if fund_data else None,
                    'fund_growth':  fund_data['rev_growth']  if fund_data else None,
                    'fund_margins': fund_data['margins']     if fund_data else None,
                    # Multi-Timeframe
                    'tf_alignment':    tf_data['alignment']    if tf_data else 'N/A',
                    'tf_signal':       tf_data['signal']       if tf_data else 'NEUTRO',
                    'tf_conf_pts':     tf_data['conf_pts']     if tf_data else 0,
                    'tf_weekly_rsi':   tf_data['weekly_rsi']   if tf_data else None,
                    'tf_weekly_trend': tf_data['weekly_trend'] if tf_data else '—',
                    'tf_daily_trend':  tf_data['daily_trend']  if tf_data else '—',
                    # Fuerza Relativa vs SPY
                    'rs_score':  rs_data['rs_score']  if rs_data else None,
                    'rs_signal': rs_data['rs_signal'] if rs_data else 'N/A',
                    'rs_20d':    rs_data['rs_20d']    if rs_data else None,
                    'rs_60d':    rs_data['rs_60d']    if rs_data else None,
                    'rs_252d':   rs_data['rs_252d']   if rs_data else None,
                    # Risk Management
                    'sharpe':       risk_data['sharpe']      if risk_data else None,
                    'max_dd':       risk_data['max_dd']      if risk_data else None,
                    'kelly_pct':    risk_data['kelly_pct']   if risk_data else None,
                    'win_rate':     risk_data['win_rate']    if risk_data else None,
                    'ann_return':   risk_data['ann_return']  if risk_data else None,
                    'ann_vol':      risk_data['ann_vol']     if risk_data else None,
                    'risk_score':   risk_data['risk_score']  if risk_data else None,
                    'risk_signal':  risk_data['risk_signal'] if risk_data else 'N/A',
                    # Backtesting
                    'bt_win_rate_tech':      bt_data['bt_win_rate_tech']      if bt_data else None,
                    'bt_avg_ret_tech':       bt_data['bt_avg_ret_tech']       if bt_data else None,
                    'bt_expectancy_tech':    bt_data['bt_expectancy_tech']    if bt_data else None,
                    'bt_n_tech':             bt_data['bt_n_tech']             if bt_data else 0,
                    'bt_best_tech':          bt_data['bt_best_tech']          if bt_data else None,
                    'bt_worst_tech':         bt_data['bt_worst_tech']         if bt_data else None,
                    'bt_sortino_tech':       bt_data.get('bt_sortino_tech')   if bt_data else None,
                    'bt_calmar_tech':        bt_data.get('bt_calmar_tech')    if bt_data else None,
                    'bt_max_dd_tech':        bt_data.get('bt_max_dd_tech')    if bt_data else None,
                    'bt_exposure_pct_tech':  bt_data.get('bt_exposure_pct_tech')  if bt_data else None,
                    'bt_avg_hold_days_tech': bt_data.get('bt_avg_hold_days_tech') if bt_data else None,
                    'bt_win_rate_vol':       bt_data['bt_win_rate_vol']       if bt_data else None,
                    'bt_avg_ret_vol':        bt_data['bt_avg_ret_vol']        if bt_data else None,
                    'bt_expectancy_vol':     bt_data['bt_expectancy_vol']     if bt_data else None,
                    'bt_n_vol':              bt_data['bt_n_vol']              if bt_data else 0,
                    'bt_sortino_vol':        bt_data.get('bt_sortino_vol')    if bt_data else None,
                    'bt_calmar_vol':         bt_data.get('bt_calmar_vol')     if bt_data else None,
                    'bt_max_dd_vol':         bt_data.get('bt_max_dd_vol')     if bt_data else None,
                    'bt_exposure_pct_vol':   bt_data.get('bt_exposure_pct_vol')   if bt_data else None,
                    'bt_avg_hold_days_vol':  bt_data.get('bt_avg_hold_days_vol')  if bt_data else None,
                    'bt_buy_hold':           bt_data['bt_buy_hold']          if bt_data else None,
                    # Dip metrics
                    'dip_from_high_pct': dip_from_high_pct,
                    'dip_3d':            dip_3d,
                    'dip_atr_multiple':  dip_atr_multiple,
                    # Retornos mediano-largo plazo
                    'change_60d':  change_60d,
                    'change_252d': change_252d,
                    # Predicción
                    'pred_trend_10d': pred_trend_10d,
                    'pred_r2':        pred_r2,
                    'pred_pat_wr':    pred_pat_wr,
                    'pred_pat_avg':   pred_pat_avg,
                    'pred_pat_n':     pred_pat_n,
                    'pred_zscore':    pred_zscore,
                    'pred_signal':    pred_signal,
                    **analysis['scores'],
                }

                conf = self.analyzer.calculate_signal_confidence(result, self._vix_data)
                result['confidence']       = conf['confidence']
                result['confidence_score'] = conf['score']

                # Progreso thread-safe
                with lock:
                    completed_count[0] += 1
                    done = completed_count[0]
                pct = 50 + int((done / n_valid) * 38)
                self._progress(pct, f'Analizando [{done}/{n_valid}] {symbol}')

                return result

            except Exception as e:
                self._log('warn', f'Error en {symbol}: {e}')
                return None

        # ── 4) Ejecutar en paralelo ────────────────────────────────────────
        # max_workers 10: equilibrio entre velocidad y límites de yfinance
        results = []
        max_workers = min(10, n_valid)
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(analyze_one, sym): sym for sym in valid_symbols}
            for future in as_completed(futures):
                if not self.running:
                    executor.shutdown(wait=False, cancel_futures=True)
                    break
                res = future.result()
                if res is not None:
                    results.append(res)

        return pd.DataFrame(results) if results else pd.DataFrame()

    # ─── DETALLE POR ACCIÓN ──────────────────

    def _detail_from_tree(self, tree, sym_col):
        sel = tree.selection()
        if not sel:
            return
        vals = tree.item(sel[0], 'values')
        if vals and len(vals) > sym_col:
            self._show_detail(str(vals[sym_col]).strip())

    def _show_detail(self, symbol):
        if self.df_results is None or self.df_results.empty:
            return
        rows = self.df_results[self.df_results['symbol'] == symbol]
        if rows.empty:
            return
        row = rows.iloc[0]

        # ── Ventana ─────────────────────────────────────────────
        win = tk.Toplevel(self.root)
        win.title(f'Detalle — {symbol}')
        win.configure(bg=COLORS['bg'])
        win.geometry('700x760')
        win.resizable(True, True)
        win.transient(self.root)

        # Canvas scrollable
        canvas = tk.Canvas(win, bg=COLORS['bg'], highlightthickness=0)
        vsb = ttk.Scrollbar(win, orient='vertical', command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(fill=tk.BOTH, expand=True)

        content = tk.Frame(canvas, bg=COLORS['bg'])
        cw = canvas.create_window((0, 0), window=content, anchor='nw')
        content.bind('<Configure>', lambda e: canvas.configure(
            scrollregion=canvas.bbox('all')))
        canvas.bind('<Configure>', lambda e: canvas.itemconfig(cw, width=e.width))

        def _scroll(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), 'units')
        canvas.bind_all('<MouseWheel>', _scroll)
        win.bind('<Destroy>', lambda e: canvas.unbind_all('<MouseWheel>'))

        # ── Colores helpers ──────────────────────────────────────
        C = COLORS
        BUY, SELL, WARN = C['buy'], C['sell'], C['warning']

        conf_colors = {'MUY ALTA': BUY, 'ALTA': '#4ade80',
                       'MEDIA': WARN, 'BAJA': SELL, 'MUY BAJA': '#ef4444'}

        def _col_signed(v, good_above=0):
            if v is None: return C['text_dim']
            return BUY if v > good_above else SELL

        # ── Header ──────────────────────────────────────────────
        hdr = tk.Frame(content, bg=C['surface'], pady=14)
        hdr.pack(fill=tk.X, pady=(0, 6))

        from stock_universe import get_stock_info
        sinfo = get_stock_info(symbol)
        tk.Label(hdr, text=symbol, font=('Segoe UI', 22, 'bold'),
                 bg=C['surface'], fg=C['text']).pack(anchor='w', padx=22)
        tk.Label(hdr, text=f"{sinfo['name']}   ·   ISIN: {sinfo['isin']}",
                 font=('Consolas', 9), bg=C['surface'],
                 fg=C['text_dim']).pack(anchor='w', padx=22)

        conf = row.get('confidence', '—')
        price = row.get('price', 0)
        c5d = row.get('change_5d', 0)
        c20d = row.get('change_20d', 0)
        sector = row.get('sector', '')

        info_f = tk.Frame(hdr, bg=C['surface'])
        info_f.pack(fill=tk.X, padx=22, pady=(4, 0))
        tk.Label(info_f,
                 text=f"${price:.2f}   {c5d:+.1f}% (5d)   {c20d:+.1f}% (20d)   |   {sector}",
                 font=('Consolas', 9), bg=C['surface'], fg=C['text_dim']).pack(side=tk.LEFT)
        tk.Label(info_f, text=f'  Confianza: {conf}',
                 font=('Segoe UI', 10, 'bold'),
                 bg=C['surface'],
                 fg=conf_colors.get(conf, C['text_dim'])).pack(side=tk.RIGHT)

        tk.Button(info_f, text='📈 Ver Gráfico',
                  command=lambda s=symbol: self._show_chart(s),
                  bg=C['primary'], fg='white',
                  font=('Segoe UI', 9), relief='flat',
                  padx=10, pady=3, cursor='hand2').pack(side=tk.RIGHT, padx=(0, 12))

        # ── Helpers de layout ────────────────────────────────────
        def section(title, bg=C['surface2']):
            f = tk.Frame(content, bg=bg, pady=6)
            f.pack(fill=tk.X, padx=14, pady=(4, 0))
            tk.Label(f, text=f'  {title}', font=('Segoe UI', 9, 'bold'),
                     bg=bg, fg=C['primary']).pack(anchor='w', padx=8)
            return f

        def item(parent, label, value, color=None):
            f = tk.Frame(parent, bg=parent['bg'])
            f.pack(fill=tk.X, padx=14, pady=1)
            tk.Label(f, text=f'{label}:', width=22, anchor='w',
                     font=('Consolas', 9), bg=parent['bg'],
                     fg=C['text_dim']).pack(side=tk.LEFT)
            tk.Label(f, text=str(value), anchor='w',
                     font=('Consolas', 9), bg=parent['bg'],
                     fg=color or C['text']).pack(side=tk.LEFT)

        # ── Score técnico ────────────────────────────────────────
        s = section('Score Técnico')
        score = row.get('total_score', 0)
        item(s, 'Score', f"{score:.1f} / 100",
             BUY if score >= 65 else (WARN if score >= 45 else SELL))
        item(s, 'RSI', f"{row.get('rsi', 0):.1f}")
        adx_v = row.get('adx_value')
        item(s, 'ADX',
             f"{adx_v:.1f} ({row.get('adx_strength','')}) — {row.get('adx_direction','')}"
             if adx_v else '—')

        # ── Volumen MA ───────────────────────────────────────────
        s = section('Volumen MA')
        vs = row.get('vol_signal', 'N/A')
        item(s, 'Señal', vs,
             BUY if vs == 'COMPRAR' else (SELL if vs == 'VENDER' else C['text_dim']))
        vr = row.get('vol_ratio')
        item(s, 'Ratio MA7/MA60', f"{vr:.3f}" if vr else '—')
        m7 = row.get('vol_ma7'); m60 = row.get('vol_ma60')
        item(s, 'MA7  (dinero 7d)',  f"${m7/1e6:.1f}M"  if m7  else '—')
        item(s, 'MA60 (dinero 60d)', f"${m60/1e6:.1f}M" if m60 else '—')

        # ── Multi-TF ─────────────────────────────────────────────
        s = section('Multi-Timeframe')
        tfa = row.get('tf_alignment', 'N/A')
        tf_pts = row.get('tf_conf_pts', 0)
        tf_sig = row.get('tf_signal', '')
        item(s, 'Alineacion',
             f"{tfa}  ({int(tf_pts):+d} pts)",
             BUY if tf_sig == 'ALCISTA' else (SELL if tf_sig == 'BAJISTA' else C['text_dim']))
        item(s, 'Tendencia Semanal', row.get('tf_weekly_trend', '—'))
        item(s, 'Tendencia Diaria',  row.get('tf_daily_trend',  '—'))
        wrsi = row.get('tf_weekly_rsi')
        item(s, 'RSI Semanal', f"{wrsi:.1f}" if wrsi else '—')

        # ── Fuerza Relativa ──────────────────────────────────────
        s = section('Fuerza Relativa vs S&P 500')
        rs_sig = row.get('rs_signal', 'N/A')
        item(s, 'Señal',
             f"{rs_sig}  (score {row.get('rs_score', 0):.0f}/100)" if row.get('rs_score') else rs_sig,
             BUY if rs_sig == 'LIDER' else (SELL if rs_sig == 'REZAGADO' else C['text_dim']))
        for lbl, key in [('RS 20 dias', 'rs_20d'), ('RS 60 dias', 'rs_60d'), ('RS 252 dias', 'rs_252d')]:
            v = row.get(key)
            item(s, lbl, f"{v:+.1f}%" if v is not None else '—', _col_signed(v))

        # ── Riesgo ───────────────────────────────────────────────
        s = section('Gestion de Riesgo')
        rsig = row.get('risk_signal', 'N/A')
        item(s, 'Perfil',
             rsig,
             BUY if rsig == 'BAJO RIESGO' else (SELL if rsig == 'ALTO RIESGO' else WARN))
        sh = row.get('sharpe')
        item(s, 'Sharpe Ratio',
             f"{sh:.2f}" if sh is not None else '—',
             BUY if sh and sh >= 1.0 else (SELL if sh and sh < 0 else C['text']))
        md = row.get('max_dd')
        item(s, 'Max Drawdown',
             f"{md:.1f}%" if md is not None else '—',
             BUY if md and md >= -15 else (SELL if md and md < -30 else WARN))
        kl = row.get('kelly_pct')
        item(s, 'Kelly % sugerido',
             f"{kl:.1f}%  →  de €10 000 invertir €{kl*100:.0f}" if kl else '—',
             C['primary'])
        wr_k = row.get('win_rate')
        item(s, 'Tasa de acierto', f"{wr_k:.1f}%" if wr_k else '—')
        ar = row.get('ann_return'); av = row.get('ann_vol')
        item(s, 'Retorno anualizado', f"{ar:+.1f}%" if ar is not None else '—', _col_signed(ar))
        item(s, 'Volatilidad anual',  f"{av:.1f}%"  if av is not None else '—')

        # ── Backtest ─────────────────────────────────────────────
        s = section('Backtesting (ventanas 20 dias)')
        bt_wr = row.get('bt_win_rate_vol')
        bt_exp = row.get('bt_expectancy_vol')
        bt_n = row.get('bt_n_vol', 0)
        item(s, 'Win Rate Tec+Vol',
             f"{bt_wr:.1f}%" if bt_wr else '—',
             BUY if bt_wr and bt_wr >= 65 else (SELL if bt_wr and bt_wr < 40 else C['text']))
        item(s, 'Win Rate Tecnico',
             f"{row.get('bt_win_rate_tech', 0):.1f}%" if row.get('bt_win_rate_tech') else '—')
        item(s, 'Expectativa',
             f"{bt_exp:+.2f}% por trade" if bt_exp is not None else '—',
             _col_signed(bt_exp))
        item(s, 'Num. señales hist.', str(int(bt_n)) if bt_n else '—')
        bh = row.get('bt_buy_hold')
        item(s, 'Buy & Hold periodo', f"{bh:+.1f}%" if bh is not None else '—', _col_signed(bh))

        # ── Fundamentales ────────────────────────────────────────
        s = section('Fundamentales')
        fs = row.get('fund_signal', 'N/A')
        item(s, 'Señal',
             f"{fs}  (score {row.get('fund_score', 50):.0f}/100)",
             BUY if fs == 'FAVORABLE' else (SELL if fs == 'DESFAVORABLE' else C['text_dim']))
        from stock_universe import is_etf as _is_etf
        if _is_etf(symbol):
            item(s, 'Tipo', 'ETF')
            gr = row.get('fund_growth'); mg = row.get('fund_margins')
            item(s, 'Retorno 3y', f"{gr:+.1f}%" if gr is not None else '—', _col_signed(gr))
            item(s, 'Retorno 5y', f"{mg:+.1f}%" if mg is not None else '—', _col_signed(mg))
        else:
            pe = row.get('fund_pe'); peg = row.get('fund_peg')
            debt = row.get('fund_debt'); gr = row.get('fund_growth')
            mg = row.get('fund_margins')
            item(s, 'P/E',           f"{pe:.1f}x" if pe is not None else '—')
            item(s, 'PEG',           f"{peg:.2f}" if peg is not None else '—')
            item(s, 'Deuda/Capital', f"{debt:.1f}%" if debt is not None else '—')
            item(s, 'Crec. Ingresos',f"{gr:+.1f}%" if gr is not None else '—', _col_signed(gr))
            item(s, 'Margen Neto',   f"{mg:.1f}%"  if mg is not None else '—', _col_signed(mg, 0))

        # ── Vela ─────────────────────────────────────────────────
        s = section('Patron de Vela')
        cs = row.get('candle_signal', 'NEUTRO')
        item(s, 'Patron', row.get('candle_pattern', 'Sin patron'))
        item(s, 'Señal', cs,
             BUY if cs == 'COMPRA' else (SELL if cs == 'VENTA' else C['text_dim']))
        item(s, 'Fuerza', row.get('candle_strength', '—'))

        # ── Resumen confianza ────────────────────────────────────
        conf_pts = row.get('confidence_score', 0)
        s = section(f'Resumen de Confianza  (puntos = {int(conf_pts):+d})')
        try:
            cd = self.analyzer.calculate_signal_confidence(row.to_dict(), self._vix_data)
            for a in cd.get('aligned', []):
                item(s, '+', a, BUY)
            for ag in cd.get('against', []):
                item(s, '-', ag, SELL)
        except Exception:
            pass

        tk.Frame(content, bg=C['bg'], height=20).pack()

    # ─── ACTUALIZAR TABLAS ───────────────────

    def _populate_all_tables(self, df):
        self._populate_topbuys()
        self._populate_dip_tab()
        self._populate_prediction_tab()
        self._populate_rankings(df)
        self._populate_volume(df)

        self._apply_candle_filter()
        self._apply_tf_filter()
        self._apply_rs_filter()
        self._apply_risk_filter()
        self._apply_bt_sort()
        self._apply_fund_filter()
        self._populate_portfolio(df)
        self._populate_sector_rotation()
        self._populate_journal()
        self._populate_portfolio_risk(df)
        self._update_metrics(df)

    def _populate_rankings(self, df):
        if df is None or df.empty:
            return

        # Filtro sector
        sectors = ['Todos'] + sorted(df['sector'].unique().tolist())
        self.filter_sector['values'] = sectors
        self.filter_sector.set('Todos')

        self._fill_rank_tree(df)

    def _apply_ranking_filter(self, event=None):
        if self.df_results is None or self.df_results.empty:
            return
        df = self.df_results.copy()

        sector = self.filter_sector.get()
        if sector and sector != 'Todos':
            df = df[df['sector'] == sector]

        search = self.search_var.get().strip().upper()
        if search:
            df = df[df['symbol'].str.upper().str.contains(search, na=False)]

        sort_map = {
            'Confianza': ('confidence_score', False),
            'Score': ('total_score', False),
            'RSI': ('rsi', False),
            'Cambio 5d': ('change_5d', False),
            'Cambio 20d': ('change_20d', False),
            'Señal Vol': ('vol_signal', True),
            'Ratio Vol': ('vol_ratio', False),
        }
        sort_key, asc = sort_map.get(self.sort_col.get(), ('confidence_score', False))
        if sort_key in df.columns:
            df = df.sort_values(sort_key, ascending=asc)

        self._fill_rank_tree(df)

    def _fill_rank_tree(self, df):
        tree = self.rank_tree
        tree.delete(*tree.get_children())

        sort_col = 'confidence_score' if 'confidence_score' in df.columns else 'total_score'
        df_sorted = df.sort_values(sort_col, ascending=False).reset_index(drop=True)

        for i, (_, row) in enumerate(df_sorted.iterrows()):
            tag = 'odd' if i % 2 == 0 else 'even'

            vs = row.get('vol_signal', 'N/A')
            vr = row.get('vol_ratio')

            # Recomendación simple según score
            score = row['total_score']
            if score >= 75:
                rec = 'COMPRA FUERTE'
            elif score >= 60:
                rec = 'COMPRA'
            elif score >= 50:
                rec = 'COMPRA MODERADA'
            elif score >= 40:
                rec = 'ESPERAR'
            else:
                rec = 'EVITAR'

            cp = row.get('candle_pattern', 'Sin patrón')
            ck = row.get('candle_strength', '—')
            candle_display = f"{cp} ({ck})" if cp != 'Sin patrón' else '—'

            adx_v = row.get('adx_value')
            adx_s = row.get('adx_strength', '—')
            adx_display = f"{adx_v:.0f} ({adx_s})" if adx_v else '—'

            fs = row.get('fund_signal', 'N/A')
            fp = row.get('fund_pe')
            fund_display = f"{fs} (P/E {fp:.0f})" if fp else fs

            conf = row.get('confidence', '—')
            tf_align = row.get('tf_alignment', '—')
            tf_pts = row.get('tf_conf_pts', 0)
            tf_display = f"{tf_align} ({int(tf_pts):+d})" if tf_align not in ('N/A', '—') else '—'

            tree.insert('', 'end', tags=(tag,), values=(
                i + 1,
                row['symbol'],
                row['sector'],
                f"{score:.1f}",
                f"${row['price']:.2f}",
                f"{row['change_5d']:+.1f}%",
                f"{row['change_20d']:+.1f}%",
                f"{row['rsi']:.0f}",
                adx_display,
                vs,
                tf_display,
                candle_display,
                fund_display,
                conf,
            ))

    def _populate_volume(self, df):
        self._apply_volume_filter()

    def _apply_volume_filter(self, event=None):
        if self.df_results is None or self.df_results.empty:
            return
        df = self.df_results.copy()

        flt = self.vol_filter.get()
        if flt == 'Solo COMPRAR':
            df = df[df['vol_signal'] == 'COMPRAR']
        elif flt == 'Solo VENDER':
            df = df[df['vol_signal'] == 'VENDER']

        # Ordenar por ratio (compras primero con ratio mayor)
        if 'vol_ratio' in df.columns:
            df = df.sort_values('vol_ratio', ascending=False)

        tree = self.vol_tree
        tree.delete(*tree.get_children())

        for i, (_, row) in enumerate(df.iterrows()):
            tag = 'odd' if i % 2 == 0 else 'even'
            vs = row.get('vol_signal', 'N/A')
            vr = row.get('vol_ratio')
            vs_pct = row.get('vol_strength')
            ma7 = row.get('vol_ma7')
            ma60 = row.get('vol_ma60')
            trend = row.get('vol_ma7_trend')

            def fmt_m(v):
                if v is None or pd.isna(v):
                    return '—'
                return f"${v/1e6:.1f}M"

            row_tag = tag
            if vs == 'COMPRAR':
                row_tag = (tag, 'buy')
            elif vs == 'VENDER':
                row_tag = (tag, 'sell')

            tree.insert('', 'end', tags=row_tag, values=(
                row['symbol'],
                row['sector'],
                f"${row['price']:.2f}",
                vs,
                f"{vr:.3f}" if vr is not None else '—',
                f"{vs_pct:.1f}%" if vs_pct is not None else '—',
                fmt_m(ma7),
                fmt_m(ma60),
                f"{trend:+.1f}%" if trend is not None else '—',
                f"{row['total_score']:.1f}",
            ))

    def _populate_portfolio(self, df):
        if df is None or df.empty:
            return
        self._fill_port_tree()

    def _fill_port_tree(self):
        if self.df_results is None or not self.portfolio:
            return

        tree = self.port_tree
        tree.delete(*tree.get_children())

        port_df = self.df_results[
            self.df_results['symbol'].isin(self.portfolio)
        ].sort_values('total_score', ascending=False).reset_index(drop=True)

        for i, (_, row) in enumerate(port_df.iterrows()):
            tag = 'odd' if i % 2 == 0 else 'even'
            score = row['total_score']

            if score >= 75:
                pos = 'Grande (5-10%)'
            elif score >= 60:
                pos = 'Media (3-5%)'
            else:
                pos = 'Pequeña (1-3%)'

            vs = row.get('vol_signal', 'N/A')
            vr = row.get('vol_ratio')

            tree.insert('', 'end', tags=(tag,), values=(
                i + 1,
                row['symbol'],
                row['sector'],
                f"{score:.1f}",
                f"${row['price']:.2f}",
                vs,
                f"{vr:.3f}" if vr is not None else '—',
                f"{row['rsi']:.0f}",
                f"{row['change_5d']:+.1f}%",
                pos,
            ))

    def _update_metrics(self, df):
        if df is None or df.empty:
            return
        total     = len(df)
        buys      = len(df[df['vol_signal'] == 'COMPRAR'])
        sells     = len(df[df['vol_signal'] == 'VENDER'])
        top_score = df['total_score'].max()
        high_conf = len(df[df.get('confidence', pd.Series(dtype=str)).isin(['ALTA', 'MUY ALTA'])]) \
                    if 'confidence' in df.columns else 0

        self.m_total.update_value(str(total))
        self.m_buy.update_value(str(buys), COLORS['buy'])
        self.m_sell.update_value(str(sells), COLORS['sell'])
        self.m_topscore.update_value(f"{top_score:.1f}", COLORS['primary'])
        self.m_confident.update_value(str(high_conf), COLORS['accent'])

        # VIX
        if self._vix_data:
            vix_color = {
                'buy': COLORS['buy'], 'sell': COLORS['sell'],
                'warn': COLORS['warning'], 'neutral': COLORS['text_dim']
            }.get(self._vix_data.get('color', 'neutral'), COLORS['text_dim'])
            self.m_vix.update_value(
                f"{self._vix_data['vix']:.1f}", vix_color)
        else:
            self.m_vix.update_value('N/A')

        # Benchmark: retorno anualizado promedio del Top 10 Score Compuesto
        if 'ann_return' in df.columns:
            df_buy = df[df['vol_signal'] == 'COMPRAR'].copy()
            if df_buy.empty:
                df_buy = df.copy()
            df_buy['_buy_score'] = df_buy.apply(self._compute_buy_score, axis=1)
            top10 = df_buy.sort_values('_buy_score', ascending=False).head(10)
            avg_ann = top10['ann_return'].dropna().mean()
            if not pd.isna(avg_ann):
                ann_color = (COLORS['buy']     if avg_ann > 15 else
                             COLORS['warning'] if avg_ann > 10 else COLORS['sell'])
                self.m_target.update_value(f"{avg_ann:.1f}%", ann_color)
            else:
                self.m_target.update_value('N/A')
        else:
            self.m_target.update_value('N/A')

    def _rebuild_portfolio(self):
        if self.df_results is None or self.df_results.empty:
            return
        try:
            n = self.cfg['portfolio_n']
            max_corr = self.cfg['max_corr']
            self.portfolio = self.analyzer.find_diversified_portfolio(
                self.df_results, n=n, max_corr=max_corr)
            self._fill_port_tree()
            self._log('ok', f'Portfolio recalculado: {len(self.portfolio)} acciones')
        except Exception as e:
            self._log('err', f'Error recalculando portfolio: {e}')

    # ─── GRÁFICO DE PRECIO ───────────────────

    def _show_chart(self, symbol):
        """Navega al tab Gráfico y muestra el símbolo indicado."""
        self.nb.select(self.tab_chart)
        self._update_chart(symbol=symbol)

    def _show_chart_popup(self, symbol):
        """(Legacy) Abre ventana popup con gráfico de precio + EMAs + volumen MA."""
        try:
            import matplotlib
            matplotlib.use('TkAgg')
            import matplotlib.pyplot as plt
            import matplotlib.dates as mdates
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        except ImportError:
            messagebox.showerror('Error', 'matplotlib no instalado.\npip install matplotlib')
            return

        df = self.db.get_precios(symbol)
        if df.empty or len(df) < 60:
            messagebox.showwarning('Sin datos', f'No hay datos suficientes para {symbol}')
            return

        df = df.sort_values('timestamp', ascending=True).reset_index(drop=True)
        df['timestamp'] = pd.to_datetime(df['timestamp'])
        df = df.tail(365).copy()   # último año

        # ── Calcular indicadores ──
        close = df['close']
        df['ema9']  = close.ewm(span=9,  adjust=False).mean()
        df['ema21'] = close.ewm(span=21, adjust=False).mean()
        df['ema50'] = close.ewm(span=50, adjust=False).mean()
        df['vol_traded'] = close * df['volume']
        df['ma7']  = df['vol_traded'].rolling(7,  min_periods=1).mean()
        df['ma60'] = df['vol_traded'].rolling(60, min_periods=1).mean()

        # Señales Vol MA: cruces
        prev_above = df['ma7'].shift(1) > df['ma60'].shift(1)
        curr_above = df['ma7'] > df['ma60']
        buy_signals  = (~prev_above) &  curr_above
        sell_signals =  prev_above  & (~curr_above)

        # ── Figura matplotlib con dark theme ──
        bg   = '#0d1117'
        surf = '#161b22'
        grid = '#30363d'
        txt  = '#e6edf3'

        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 7),
                                        gridspec_kw={'height_ratios': [3, 1]},
                                        facecolor=bg)
        fig.subplots_adjust(hspace=0.06)

        dates = df['timestamp']

        # ── Subplot 1: Precio + EMAs ──
        ax1.set_facecolor(surf)
        ax1.plot(dates, close,       color='#79c0ff', lw=1.4, label='Precio')
        ax1.plot(dates, df['ema9'],  color='#3fb950', lw=1.0, label='EMA 9',  alpha=0.9)
        ax1.plot(dates, df['ema21'], color='#d29922', lw=1.0, label='EMA 21', alpha=0.9)
        ax1.plot(dates, df['ema50'], color='#f85149', lw=1.0, label='EMA 50', alpha=0.9)

        # Señales de compra/venta
        ax1.scatter(dates[buy_signals],  close[buy_signals],
                    marker='^', color='#3fb950', s=80, zorder=5, label='Compra Vol MA')
        ax1.scatter(dates[sell_signals], close[sell_signals],
                    marker='v', color='#f85149', s=80, zorder=5, label='Venta Vol MA')

        # ATR stop-loss band (si hay datos del análisis)
        if self.df_results is not None:
            r = self.df_results[self.df_results['symbol'] == symbol]
            if not r.empty:
                atr_pct = r.iloc[0].get('atr_pct')
                cur_p   = r.iloc[0].get('price')
                if atr_pct and cur_p:
                    stop = cur_p - 2 * (cur_p * atr_pct / 100)
                    ax1.axhline(stop, color='#f85149', lw=0.8, ls='--', alpha=0.6,
                                label=f'Stop ATR ${stop:.2f}')

        ax1.set_title(f'{symbol} — Precio, EMAs y Señales Vol MA (último año)',
                      color=txt, fontsize=11, pad=10)
        ax1.tick_params(colors=txt, labelsize=8)
        ax1.set_facecolor(surf)
        for spine in ax1.spines.values():
            spine.set_color(grid)
        ax1.yaxis.grid(True, color=grid, lw=0.4)
        ax1.xaxis.set_visible(False)
        legend = ax1.legend(fontsize=8, facecolor=surf, edgecolor=grid,
                             labelcolor=txt, loc='upper left')

        # ── Subplot 2: Volumen con color por señal Vol MA ──
        ax2.set_facecolor(surf)
        colors_vol = ['#3fb950' if m > c else '#f85149'
                      for m, c in zip(df['ma7'], df['ma60'])]
        ax2.bar(dates, df['volume'], color=colors_vol, alpha=0.6, width=1)
        ax2.set_ylabel('Volumen', color=txt, fontsize=8)
        ax2.tick_params(colors=txt, labelsize=7)
        ax2.set_facecolor(surf)
        for spine in ax2.spines.values():
            spine.set_color(grid)
        ax2.yaxis.grid(True, color=grid, lw=0.3)
        ax2.xaxis.set_major_formatter(mdates.DateFormatter('%b %Y'))
        ax2.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
        plt.setp(ax2.xaxis.get_majorticklabels(), rotation=30, ha='right', color=txt)

        # ── Embed en tkinter ──
        win = tk.Toplevel(self.root)
        win.title(f'Gráfico — {symbol}')
        win.configure(bg=bg)
        win.geometry('1000x600')

        canvas_fig = FigureCanvasTkAgg(fig, master=win)
        canvas_fig.draw()
        canvas_fig.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        tk.Button(win, text='Cerrar', command=lambda: (plt.close(fig), win.destroy()),
                  bg=COLORS['surface2'], fg=COLORS['text'],
                  font=('Segoe UI', 9), relief='flat', padx=16, pady=4).pack(pady=6)

        win.protocol('WM_DELETE_WINDOW', lambda: (plt.close(fig), win.destroy()))

    # ─── TAB GRÁFICO ─────────────────────────

    def _build_chart_tab(self):
        p = self.tab_chart

        # ── Barra de controles ──────────────────────────────────────────
        ctrl = tk.Frame(p, bg=COLORS['surface'], height=52)
        ctrl.pack(fill=tk.X)
        ctrl.pack_propagate(False)

        tk.Label(ctrl, text='Símbolo:', bg=COLORS['surface'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=(16, 4))

        from stock_universe import get_all_symbols
        all_syms = sorted(get_all_symbols())
        self._chart_sym_var = tk.StringVar(value=all_syms[0])
        sym_cb = ttk.Combobox(ctrl, textvariable=self._chart_sym_var,
                              values=all_syms, width=10, state='readonly',
                              font=('Segoe UI', 9))
        sym_cb.pack(side=tk.LEFT, padx=(0, 20))
        sym_cb.bind('<<ComboboxSelected>>', lambda e: self._update_chart())

        tk.Label(ctrl, text='Período:', bg=COLORS['surface'],
                 fg=COLORS['text_dim'], font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=(0, 6))

        self._chart_period = tk.StringVar(value='1Y')
        self._period_btns = []
        for label in ('1M', '3M', '6M', '1Y', '2Y', 'MAX'):
            is_active = label == '1Y'
            btn = tk.Button(
                ctrl, text=label,
                bg=COLORS['primary'] if is_active else COLORS['surface2'],
                fg='white' if is_active else COLORS['text_dim'],
                font=('Segoe UI', 8, 'bold' if is_active else 'normal'),
                relief='flat', padx=10, pady=5, cursor='hand2',
                command=lambda lbl=label: self._set_chart_period(lbl))
            btn.pack(side=tk.LEFT, padx=2)
            self._period_btns.append((btn, label))

        # ── Área del gráfico ────────────────────────────────────────────
        self._chart_area = tk.Frame(p, bg=COLORS['bg'])
        self._chart_area.pack(fill=tk.BOTH, expand=True)

        tk.Label(self._chart_area,
                 text='Haz clic en cualquier acción de las tablas\no selecciona un símbolo arriba',
                 bg=COLORS['bg'], fg=COLORS['text_dim'],
                 font=('Segoe UI', 12)).place(relx=0.5, rely=0.5, anchor='center')

    def _set_chart_period(self, period):
        self._chart_period.set(period)
        for btn, lbl in self._period_btns:
            active = lbl == period
            btn.config(
                bg=COLORS['primary'] if active else COLORS['surface2'],
                fg='white' if active else COLORS['text_dim'],
                font=('Segoe UI', 8, 'bold' if active else 'normal'))
        self._update_chart()

    def _update_chart(self, symbol=None):
        """Renderiza el gráfico en el tab con precio, RSI y volumen."""
        import matplotlib
        matplotlib.use('TkAgg')
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

        if symbol:
            self._chart_sym_var.set(symbol)

        sym = self._chart_sym_var.get()
        per = self._chart_period.get()

        # Destruir canvas anterior
        if self._chart_canvas is not None:
            plt.close(self._chart_fig)
            self._chart_canvas.get_tk_widget().destroy()
            self._chart_canvas = None
            self._chart_fig = None
        for w in self._chart_area.winfo_children():
            w.destroy()

        # Cargar datos
        try:
            db = self.db if self.db else __import__('database').TradingDatabase()
            df = db.get_precios(sym)
        except Exception as e:
            tk.Label(self._chart_area, text=f'Error cargando datos: {e}',
                     bg=COLORS['bg'], fg=COLORS['danger'],
                     font=('Segoe UI', 10)).place(relx=0.5, rely=0.5, anchor='center')
            return

        if df.empty or len(df) < 30:
            tk.Label(self._chart_area, text=f'Sin datos suficientes para {sym}',
                     bg=COLORS['bg'], fg=COLORS['text_dim'],
                     font=('Segoe UI', 11)).place(relx=0.5, rely=0.5, anchor='center')
            return

        df = df.sort_values('timestamp', ascending=True).reset_index(drop=True)
        df['timestamp'] = pd.to_datetime(df['timestamp'])

        period_days = {'1M': 30, '3M': 90, '6M': 180, '1Y': 365, '2Y': 730, 'MAX': len(df)}
        days = period_days.get(per, 365)
        df = df.tail(days).copy()

        close = df['close']
        # Usar indicadores precalculados si están disponibles, sino calcular
        df['ema9']  = df['ema_9']  if 'ema_9'  in df.columns else close.ewm(span=9,  adjust=False).mean()
        df['ema21'] = df['ema_21'] if 'ema_21' in df.columns else close.ewm(span=21, adjust=False).mean()
        df['ema50'] = close.ewm(span=50,  adjust=False).mean()
        df['ema200']= close.ewm(span=200, adjust=False).mean()

        if 'rsi' in df.columns and df['rsi'].notna().sum() > 10:
            rsi_col = df['rsi']
        else:
            delta = close.diff()
            gain = delta.clip(lower=0).rolling(14).mean()
            loss = (-delta.clip(upper=0)).rolling(14).mean()
            rsi_col = 100 - (100 / (1 + gain / loss.replace(0, float('nan'))))

        df['vol_ma20'] = df['volume'].rolling(20, min_periods=1).mean()

        bg_c  = '#0d1117'
        surf  = '#161b22'
        grid  = '#30363d'
        txt   = '#e6edf3'

        fig, axes = plt.subplots(3, 1, figsize=(13, 8),
                                 gridspec_kw={'height_ratios': [4, 1.2, 1]},
                                 facecolor=bg_c)
        fig.subplots_adjust(hspace=0.04, top=0.92, bottom=0.07, left=0.04, right=0.96)
        ax1, ax2, ax3 = axes

        dates = df['timestamp']

        # ── Precio + EMAs ──────────────────────────────────────────────
        ax1.set_facecolor(surf)
        ax1.plot(dates, close,        color='#79c0ff', lw=1.5, label=sym, zorder=3)
        ax1.plot(dates, df['ema9'],   color='#3fb950', lw=0.9, label='EMA9',   alpha=0.85)
        ax1.plot(dates, df['ema21'],  color='#d29922', lw=0.9, label='EMA21',  alpha=0.85)
        ax1.plot(dates, df['ema50'],  color='#f85149', lw=0.9, label='EMA50',  alpha=0.85)
        if days >= 200:
            ax1.plot(dates, df['ema200'], color='#8957e5', lw=1.0, label='EMA200', alpha=0.75)

        # Bandas Bollinger si existen
        if 'bb_upper' in df.columns and df['bb_upper'].notna().sum() > 5:
            ax1.fill_between(dates, df['bb_lower'], df['bb_upper'],
                             alpha=0.06, color='#79c0ff', zorder=1)

        last_price = float(close.iloc[-1])
        chg_pct = (last_price / float(close.iloc[0]) - 1) * 100
        title = f'{sym}   ${last_price:.2f}   {chg_pct:+.1f}% ({per})'
        ax1.set_title(title, color=txt, fontsize=12, pad=8, fontweight='bold', loc='left')
        ax1.legend(fontsize=7, facecolor=surf, edgecolor=grid, labelcolor=txt,
                   loc='upper left', ncol=5, framealpha=0.8)
        ax1.tick_params(colors=txt, labelsize=7)
        for spine in ax1.spines.values(): spine.set_color(grid)
        ax1.yaxis.grid(True, color=grid, lw=0.3, alpha=0.6)
        ax1.xaxis.set_visible(False)
        ax1.yaxis.set_label_position('right')
        ax1.yaxis.tick_right()

        # ── RSI ────────────────────────────────────────────────────────
        ax2.set_facecolor(surf)
        ax2.plot(dates, rsi_col, color='#c9d1d9', lw=1.0)
        ax2.axhline(70, color='#f85149', lw=0.7, ls='--', alpha=0.8)
        ax2.axhline(50, color=grid,      lw=0.5, ls=':',  alpha=0.6)
        ax2.axhline(30, color='#3fb950', lw=0.7, ls='--', alpha=0.8)
        ax2.fill_between(dates, rsi_col, 70, where=(rsi_col >= 70),
                         color='#f85149', alpha=0.18, interpolate=True)
        ax2.fill_between(dates, rsi_col, 30, where=(rsi_col <= 30),
                         color='#3fb950', alpha=0.18, interpolate=True)
        ax2.set_ylim(0, 100)
        ax2.set_ylabel('RSI', color=txt, fontsize=7, labelpad=2)
        ax2.tick_params(colors=txt, labelsize=6)
        for spine in ax2.spines.values(): spine.set_color(grid)
        ax2.yaxis.grid(True, color=grid, lw=0.3, alpha=0.4)
        ax2.xaxis.set_visible(False)
        ax2.yaxis.set_label_position('right')
        ax2.yaxis.tick_right()

        # ── Volumen ────────────────────────────────────────────────────
        ax3.set_facecolor(surf)
        if 'open' in df.columns:
            vol_colors = ['#3fb950' if c >= o else '#f85149'
                          for c, o in zip(df['close'], df['open'])]
        else:
            vol_colors = '#58a6ff'
        ax3.bar(dates, df['volume'], color=vol_colors, alpha=0.55, width=1)
        ax3.plot(dates, df['vol_ma20'], color='#d29922', lw=0.8, alpha=0.85)
        ax3.set_ylabel('Vol', color=txt, fontsize=7, labelpad=2)
        ax3.tick_params(colors=txt, labelsize=6)
        for spine in ax3.spines.values(): spine.set_color(grid)
        ax3.yaxis.grid(True, color=grid, lw=0.3, alpha=0.3)
        interval = max(1, days // 180)
        ax3.xaxis.set_major_formatter(
            mdates.DateFormatter('%b %Y' if days > 90 else '%d %b'))
        ax3.xaxis.set_major_locator(mdates.MonthLocator(interval=interval))
        plt.setp(ax3.xaxis.get_majorticklabels(),
                 rotation=30, ha='right', color=txt, fontsize=6)
        ax3.yaxis.set_label_position('right')
        ax3.yaxis.tick_right()

        # ── Embed en tkinter ──────────────────────────────────────────
        self._chart_fig = fig
        canvas = FigureCanvasTkAgg(fig, master=self._chart_area)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        self._chart_canvas = canvas

    # ─── ACTUALIZACIÓN DE DATOS ──────────────

    def _check_data_freshness(self):
        """Comprueba si la BD tiene datos de hoy y actualiza el label del header."""
        try:
            import sqlite3, datetime
            db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'trading_bot.db')
            conn = sqlite3.connect(db_path)
            cur  = conn.cursor()
            today = datetime.date.today().isoformat()
            cur.execute('SELECT MAX(timestamp) FROM precios')
            last_row = cur.fetchone()
            last_date = (last_row[0] or '')[:10] if last_row else ''
            cur.execute('SELECT COUNT(DISTINCT symbol) FROM precios WHERE timestamp >= ?', (today,))
            n_today = cur.fetchone()[0]
            conn.close()

            if n_today > 0:
                self.hdr_db_fresh.config(text=f'BD: hoy', fg=COLORS['success'])
            elif last_date:
                self.hdr_db_fresh.config(text=f'BD: {last_date} ⚠', fg=COLORS['warning'])
                # Aviso en el log
                self._log('warn', f'Datos desactualizados: último registro {last_date}. '
                          'Pulsa "Actualizar Datos" para traer datos de hoy.')
            else:
                self.hdr_db_fresh.config(text='BD: sin datos', fg=COLORS['danger'])
        except Exception as e:
            self.hdr_db_fresh.config(text='BD: ?', fg=COLORS['text_dim'])

    def _run_data_update(self):
        """Descarga los últimos 7 días para todos los símbolos en background."""
        if self._data_updating:
            return
        self._data_updating = True
        self.btn_update_data.config(text='⏳ Actualizando...', state='disabled',
                                    fg=COLORS['text_dim'])
        self._log('info', 'Iniciando actualización incremental de datos...')

        def _worker():
            try:
                from data_downloader import DataDownloader
                from database import TradingDatabase
                from stock_universe import get_all_symbols
                db = TradingDatabase()
                dl = DataDownloader(db)
                symbols = get_all_symbols()
                ok = 0
                for sym in symbols:
                    try:
                        dl.update_recent_data(sym, days=7)
                        ok += 1
                    except Exception:
                        pass
                self.q.put(('data_update_done', ok))
                self._q_notify()
            except Exception as e:
                self.q.put(('data_update_error', str(e)))
                self._q_notify()

        threading.Thread(target=_worker, daemon=True).start()

    # ─── EXPORT ──────────────────────────────

    def _export_report(self):
        if self.df_results is None or self.df_results.empty:
            messagebox.showwarning('Sin datos', 'Primero ejecuta el análisis.')
            return

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_name = f'REPORTE_INVERSIONES_{timestamp}'

        file_path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel', '*.xlsx'), ('Markdown', '*.md'), ('CSV', '*.csv'), ('Todos', '*.*')],
            initialfile=default_name,
            title='Guardar reporte',
        )
        if not file_path:
            return

        try:
            if file_path.endswith('.csv'):
                self.df_results.to_csv(file_path, index=False)
                self._log('ok', f'CSV guardado: {file_path}')
            elif file_path.endswith('.md'):
                self._export_markdown(file_path)
                self._log('ok', f'Reporte Markdown guardado: {file_path}')
            else:
                self._export_excel(file_path)
                self._log('ok', f'Excel guardado: {file_path}')

            messagebox.showinfo('Exportado', f'Guardado en:\n{file_path}')
        except Exception as e:
            self._log('err', f'Error exportando: {e}')
            messagebox.showerror('Error', str(e))

    def _export_markdown(self, path):
        from generate_investment_report import generate_markdown_report
        report = generate_markdown_report(
            self.df_results, self.portfolio, self.analyzer, self.db)
        with open(path, 'w', encoding='utf-8') as f:
            f.write(report)

    def _export_excel(self, path):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        df = self.df_results.copy()

        # ── Paleta de colores (RRGGBB sin #) ──
        C_HEADER_BG  = '1F2937'   # gris oscuro
        C_HEADER_FG  = 'F9FAFB'   # blanco
        C_BUY        = '14532D'   # verde oscuro
        C_SELL       = '7F1D1D'   # rojo oscuro
        C_NEUTRAL    = '1E293B'   # azul gris
        C_ALT        = '111827'   # fila alternada
        C_FG_LIGHT   = 'E2E8F0'   # texto claro

        thin = Side(style='thin', color='374151')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def _hdr_font():
            return Font(name='Calibri', bold=True, color=C_HEADER_FG, size=10)

        def _hdr_fill():
            return PatternFill('solid', fgColor=C_HEADER_BG)

        def _cell_font(color=C_FG_LIGHT):
            return Font(name='Calibri', color=color, size=10)

        def _row_fill(signal_val, idx):
            sig = str(signal_val).upper() if signal_val else ''
            if any(k in sig for k in ('COMPRAR', 'ALCISTA', 'LIDER', 'BAJO RIESGO',
                                       'MUY ALTA', 'ALTA')):
                bg = C_BUY
            elif any(k in sig for k in ('VENDER', 'BAJISTA', 'REZAGADO', 'ALTO RIESGO',
                                         'MUY BAJA', 'BAJA')):
                bg = C_SELL
            else:
                bg = C_ALT if idx % 2 == 0 else C_NEUTRAL
            return PatternFill('solid', fgColor=bg)

        def _write_sheet(ws, headers, rows, signal_col_idx=None):
            """Write header + data rows; auto-fit column widths."""
            ws.sheet_view.showGridLines = False
            # Header row
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font   = _hdr_font()
                cell.fill   = _hdr_fill()
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           wrap_text=True)
            ws.row_dimensions[1].height = 30

            col_widths = [len(str(h)) + 2 for h in headers]

            for ri, row_vals in enumerate(rows, 2):
                sig_val = row_vals[signal_col_idx] if signal_col_idx is not None else None
                fill    = _row_fill(sig_val, ri)
                for ci, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.font      = _cell_font()
                    cell.fill      = fill
                    cell.border    = border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    col_widths[ci - 1] = max(col_widths[ci - 1], len(str(val)) + 2)

            for ci, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = min(w, 35)

        # ── helpers ──
        def pct(v):
            return f'{v:.1f}%' if v is not None else '—'
        def flt2(v):
            return f'{v:.2f}' if v is not None else '—'
        def flt1(v):
            return f'{v:.1f}' if v is not None else '—'
        def price(v):
            return f'${v:.2f}' if v is not None else '—'

        # helper: sort multi-columna con na_position='last'
        def msort(frame, cols, asc=None):
            asc = asc or [False] * len(cols)
            existing = [c for c in cols if c in frame.columns]
            asc_ex   = [asc[i] for i, c in enumerate(cols) if c in frame.columns]
            if not existing:
                return frame
            return frame.sort_values(existing, ascending=asc_ex, na_position='last')

        # ══════════════════════════════════════
        # Hoja 1 — Rankings
        # Orden: mayor confianza → mayor score técnico → mejor rendimiento 20d
        # ══════════════════════════════════════
        ws1 = wb.active
        ws1.title = 'Rankings'
        df_r = msort(df, ['confidence_score', 'total_score', 'change_20d'])
        hdrs1 = ['#', 'Símbolo', 'Sector', 'Score', 'Precio', 'Δ5d%', 'Δ20d%',
                 'RSI', 'ADX', 'Vol MA', 'Multi-TF', 'Vela', 'Confianza']
        rows1 = []
        for i, (_, row) in enumerate(df_r.iterrows(), 1):
            rows1.append([
                i,
                row.get('symbol', ''),
                row.get('sector', ''),
                flt1(row.get('total_score')),
                price(row.get('price')),
                pct(row.get('change_5d')),
                pct(row.get('change_20d')),
                flt1(row.get('rsi')),
                flt1(row.get('adx')),
                row.get('vol_signal', '—'),
                row.get('tf_alignment', '—'),
                row.get('candle_pattern', '—'),
                row.get('confidence', '—'),
            ])
        _write_sheet(ws1, hdrs1, rows1, signal_col_idx=12)

        # ══════════════════════════════════════
        # Hoja 2 — Volumen MA
        # Orden: mayor ratio MA7/MA60 (>1=COMPRAR, <1=VENDER) → confianza → score
        # ══════════════════════════════════════
        ws2 = wb.create_sheet('Volumen MA')
        df_v = msort(df, ['vol_ratio', 'confidence_score', 'total_score'])
        hdrs2 = ['Símbolo', 'Sector', 'Precio', 'Señal Vol MA', 'Ratio MA7/MA60',
                 'MA7 (€M)', 'MA60 (€M)', 'Tendencia MA7', 'Score Téc.', 'Confianza']
        rows2 = []
        for _, row in df_v.iterrows():
            ma7  = row.get('vol_ma7')
            ma60 = row.get('vol_ma60')
            rows2.append([
                row.get('symbol', ''),
                row.get('sector', ''),
                price(row.get('price')),
                row.get('vol_signal', '—'),
                flt2(row.get('vol_ratio')),
                f'{ma7/1e6:.1f}M' if ma7 is not None else '—',
                f'{ma60/1e6:.1f}M' if ma60 is not None else '—',
                row.get('vol_ma7_trend', '—'),
                flt1(row.get('total_score')),
                row.get('confidence', '—'),
            ])
        _write_sheet(ws2, hdrs2, rows2, signal_col_idx=3)

        # ══════════════════════════════════════
        # Hoja 3 — Multi-Timeframe
        # Orden: mayor pts confianza TF (+2=ALINEADO ALCISTA) → confianza → score
        # ══════════════════════════════════════
        ws3 = wb.create_sheet('Multi-Timeframe')
        df_tf = msort(df, ['tf_conf_pts', 'confidence_score', 'total_score'])
        hdrs3 = ['Símbolo', 'Sector', 'Precio', 'Alineación TF', 'Señal TF',
                 'Tend. Semanal', 'Tend. Diaria', 'RSI Semanal', 'Pts Conf.', 'Score Téc.', 'Confianza']
        rows3 = []
        for _, row in df_tf.iterrows():
            rows3.append([
                row.get('symbol', ''),
                row.get('sector', ''),
                price(row.get('price')),
                row.get('tf_alignment', '—'),
                row.get('tf_signal', '—'),
                row.get('tf_weekly_trend', '—'),
                row.get('tf_daily_trend', '—'),
                flt1(row.get('tf_weekly_rsi')),
                row.get('tf_conf_pts', '—'),
                flt1(row.get('total_score')),
                row.get('confidence', '—'),
            ])
        _write_sheet(ws3, hdrs3, rows3, signal_col_idx=4)

        # ══════════════════════════════════════
        # Hoja 4 — Fuerza Relativa
        # Orden: mayor score RS (100=LIDER) → confianza → score
        # ══════════════════════════════════════
        ws4 = wb.create_sheet('Fuerza Relativa')
        df_rs = msort(df, ['rs_score', 'confidence_score', 'total_score'])
        hdrs4 = ['Símbolo', 'Sector', 'Precio', 'Señal RS', 'Score RS',
                 'RS 20d vs SPY', 'RS 60d vs SPY', 'RS 252d vs SPY', 'Score Téc.', 'Confianza']
        rows4 = []
        for _, row in df_rs.iterrows():
            rows4.append([
                row.get('symbol', ''),
                row.get('sector', ''),
                price(row.get('price')),
                row.get('rs_signal', '—'),
                flt1(row.get('rs_score')),
                pct(row.get('rs_20d')),
                pct(row.get('rs_60d')),
                pct(row.get('rs_252d')),
                flt1(row.get('total_score')),
                row.get('confidence', '—'),
            ])
        _write_sheet(ws4, hdrs4, rows4, signal_col_idx=3)

        # ══════════════════════════════════════
        # Hoja 5 — Riesgo
        # Orden: mejor risk_score (BAJO RIESGO=alto score) → Sharpe → confianza
        # ══════════════════════════════════════
        ws5 = wb.create_sheet('Riesgo')
        df_rk = msort(df, ['risk_score', 'sharpe', 'confidence_score'])
        hdrs5 = ['Símbolo', 'Sector', 'Precio', 'Señal Riesgo', 'Score Riesgo',
                 'Sharpe', 'Max Drawdown', 'Kelly %', 'Ret Anual', 'Vol Anual', 'Score Téc.', 'Confianza']
        rows5 = []
        for _, row in df_rk.iterrows():
            rows5.append([
                row.get('symbol', ''),
                row.get('sector', ''),
                price(row.get('price')),
                row.get('risk_signal', '—'),
                flt1(row.get('risk_score')),
                flt2(row.get('sharpe')),
                pct(row.get('max_dd')),
                pct(row.get('kelly_pct')),
                pct(row.get('ann_return')),
                pct(row.get('ann_vol')),
                flt1(row.get('total_score')),
                row.get('confidence', '—'),
            ])
        _write_sheet(ws5, hdrs5, rows5, signal_col_idx=3)

        # ══════════════════════════════════════
        # Hoja 6 — Backtesting
        # Orden: mayor win rate Téc+Vol → mayor expectancy → confianza
        # ══════════════════════════════════════
        ws6 = wb.create_sheet('Backtesting')
        df_bt = msort(df, ['bt_win_rate_vol', 'bt_expectancy_vol', 'confidence_score'])
        hdrs6 = ['Símbolo', 'Sector', 'Precio',
                 'Win% Téc+Vol', 'Ret Med Téc+Vol', 'Expectancy Téc+Vol', 'Nº Señales Vol',
                 'Win% Técnica', 'Ret Med Técnica', 'Nº Señales Téc',
                 'Buy&Hold', 'Días Fwd', 'Confianza']
        rows6 = []
        for _, row in df_bt.iterrows():
            rows6.append([
                row.get('symbol', ''),
                row.get('sector', ''),
                price(row.get('price')),
                pct(row.get('bt_win_rate_vol')),
                pct(row.get('bt_avg_ret_vol')),
                pct(row.get('bt_expectancy_vol')),
                row.get('bt_n_vol', '—'),
                pct(row.get('bt_win_rate_tech')),
                pct(row.get('bt_avg_ret_tech')),
                row.get('bt_n_tech', '—'),
                pct(row.get('bt_buy_hold')),
                row.get('bt_forward_days', '—'),
                row.get('confidence', '—'),
            ])
        _write_sheet(ws6, hdrs6, rows6, signal_col_idx=None)

        # ══════════════════════════════════════
        # Hoja 7 — Fundamentales
        # Orden: mayor score fundamentales → confianza → score técnico
        # ══════════════════════════════════════
        ws7 = wb.create_sheet('Fundamentales')
        df_f = msort(df, ['fund_score', 'confidence_score', 'total_score'])
        hdrs7 = ['Símbolo', 'Sector', 'Precio', 'Señal Fund.',
                 'Score Fund.', 'P/E', 'PEG', 'D/E', 'Rev. Growth', 'Score Téc.', 'Confianza']
        rows7 = []
        for _, row in df_f.iterrows():
            rows7.append([
                row.get('symbol', ''),
                row.get('sector', ''),
                price(row.get('price')),
                row.get('fund_signal', '—'),
                flt1(row.get('fund_score')),
                flt2(row.get('fund_pe')),
                flt2(row.get('fund_peg')),
                flt2(row.get('fund_debt')),
                pct(row.get('fund_growth')),
                flt1(row.get('total_score')),
                row.get('confidence', '—'),
            ])
        _write_sheet(ws7, hdrs7, rows7, signal_col_idx=3)

        wb.save(path)

    # ─── CONFIGURACIÓN ───────────────────────

    def _schedule_toggle(self):
        if self._sched_enabled.get():
            self._schedule_next()
        else:
            if self._scheduler_timer:
                self._scheduler_timer.cancel()
                self._scheduler_timer = None
            self._sched_label.config(text='Scheduler inactivo', fg=COLORS['text_muted'])

    def _schedule_next(self):
        """Calcula segundos hasta la próxima ejecución programada y lanza timer."""
        import threading as _th
        from datetime import datetime as _dt, timedelta as _td
        try:
            h, m = map(int, self._sched_time.get().strip().split(':'))
        except ValueError:
            self._sched_label.config(text='Hora inválida (usa HH:MM)', fg=COLORS['danger'])
            return

        now  = _dt.now()
        next_run = now.replace(hour=h, minute=m, second=0, microsecond=0)
        if next_run <= now:
            next_run += _td(days=1)

        secs = (next_run - now).total_seconds()
        self._sched_label.config(
            text=f'Próximo análisis: {next_run.strftime("%d/%m/%Y %H:%M")}',
            fg=COLORS['success'])

        if self._scheduler_timer:
            self._scheduler_timer.cancel()

        def _fire():
            self.root.after(0, self._scheduler_run)

        self._scheduler_timer = _th.Timer(secs, _fire)
        self._scheduler_timer.daemon = True
        self._scheduler_timer.start()

    def _scheduler_run(self):
        """Ejecuta el análisis automático y re-programa el siguiente."""
        if not self.running:
            self._log('info', f'Auto-refresh programado — iniciando análisis...')
            self._start_analysis()
        # Re-programar para mañana a la misma hora
        if self._sched_enabled.get():
            self._schedule_next()

    def _save_settings(self):
        try:
            self.cfg['portfolio_n'] = int(self._cfg_n.get())
            self.cfg['max_corr']    = float(self._cfg_corr.get())
            self.cfg['force_download'] = self._cfg_force.get()
            messagebox.showinfo('Configuración', 'Configuración guardada.')
            self._log('ok', f"Config guardada: portfolio_n={self.cfg['portfolio_n']}, "
                            f"max_corr={self.cfg['max_corr']}, "
                            f"force_download={self.cfg['force_download']}")
        except ValueError as e:
            messagebox.showerror('Error', f'Valor inválido: {e}')

    # ─── COLA DE MENSAJES ────────────────────

    def _poll_queue(self):
        try:
            while True:
                msg_type, data = self.q.get_nowait()

                if msg_type == 'log':
                    tag, text = data
                    self._log_write(tag, text)

                elif msg_type == 'progress':
                    pct, label = data
                    self.progress_bar['value'] = pct
                    self.progress_label.config(text=label)
                    self.status_label.config(text=label)

                elif msg_type == 'alerts':
                    self.root.after(800, lambda a=data: self._show_alerts_popup(a))
                    # También notificación de escritorio para alertas
                    n_al = len(data)
                    pos  = [a['symbol'] for a in data if a.get('type') == 'positive'][:3]
                    al_msg = (f"{n_al} cambios de señal detectados."
                              + (f" Mejoras: {', '.join(pos)}" if pos else ""))
                    threading.Thread(
                        target=self._send_desktop_notification,
                        args=('Investment Report — Alertas', al_msg),
                        daemon=True
                    ).start()

                elif msg_type == 'earnings_done':
                    self._log('ok', f'Earnings cargados: {data} símbolos con fecha próxima')
                    self._populate_topbuys()   # refrescar tabla con datos de earnings

                elif msg_type == 'done_ok':
                    self.running = False
                    self.hdr_status.config(text='● COMPLETADO', fg=COLORS['success'])
                    self.btn_run.set_enabled(True)
                    self.btn_export.set_enabled(True)
                    self.btn_rebuild.set_enabled(True)
                    self._log('ok', '¡Análisis completado exitosamente!')
                    self._populate_all_tables(self.df_results)
                    self._save_last_results()
                    # Cambiar a pestaña Top Compras al terminar
                    self.nb.select(self.tab_topbuys)
                    # Notificación de escritorio al completar análisis
                    df_done = self.df_results
                    n_buy   = int((df_done['vol_signal'] == 'COMPRAR').sum()) if df_done is not None else 0
                    top3    = (df_done[df_done['vol_signal'] == 'COMPRAR']
                               .assign(_s=lambda d: d.apply(self._compute_buy_score, axis=1))
                               .nlargest(3, '_s')['symbol'].tolist()
                               if df_done is not None and n_buy > 0 else [])
                    notif_msg = (f"{n_buy} acciones con señal COMPRAR."
                                 + (f" Top: {', '.join(top3)}" if top3 else ""))
                    threading.Thread(
                        target=self._send_desktop_notification,
                        args=('Investment Report — Análisis listo', notif_msg),
                        daemon=True
                    ).start()

                elif msg_type == 'done_error':
                    self.running = False
                    self.hdr_status.config(text='● ERROR', fg=COLORS['danger'])
                    self.btn_run.set_enabled(True)
                    self._log('err', 'El análisis terminó con errores.')

                elif msg_type == 'ia_result':
                    self._ia_running = False
                    self._ia_results = data
                    self.ia_btn.config(state='normal',
                                       text='🤖 Analizar con Claude IA')
                    self._populate_ia_table(data)
                    self._log('ok', f'Análisis IA completado: {len(data)} acciones evaluadas')
                    self.nb.select(self.tab_ia)

                elif msg_type == 'ia_error':
                    self._ia_running = False
                    self.ia_btn.config(state='normal',
                                       text='🤖 Analizar con Claude IA')
                    self.ia_status_lbl.config(
                        text=f'Error: {str(data)[:80]}',
                        fg=COLORS['danger']
                    )
                    self._log('err', f'Error en análisis IA: {data}')

                elif msg_type == 'data_update_done':
                    self._data_updating = False
                    self.btn_update_data.config(
                        text='⬇ Actualizar Datos', state='normal',
                        fg=COLORS['primary'])
                    self._log('ok', f'Datos actualizados: {data} símbolos procesados.')
                    self._check_data_freshness()

                elif msg_type == 'data_update_error':
                    self._data_updating = False
                    self.btn_update_data.config(
                        text='⬇ Actualizar Datos', state='normal',
                        fg=COLORS['primary'])
                    self._log('err', f'Error actualizando datos: {data}')

                elif msg_type == 'portfolio_risk_done':
                    self._render_portfolio_risk(data)

                elif msg_type == 'portfolio_risk_error':
                    self._pr_status.config(text=f'Error: {data}', fg=COLORS['danger'])
                    self._log('err', f'Portfolio risk error: {data}')

        except queue.Empty:
            pass
        # No re-schedule aquí — el hilo principal se activa por:
        #   a) event_generate('<<QueueUpdate>>') desde _q_notify()
        #   b) _start_poll_fallback() cada 500 ms

    # ─── LOG ─────────────────────────────────

    def _progress(self, pct, label):
        self.q.put(('progress', (pct, label)))
        self._q_notify()

    def _log(self, tag, text):
        self.q.put(('log', (tag, text)))
        self._q_notify()

    def _log_write(self, tag, text):
        self.log_box.config(state='normal')
        ts = datetime.now().strftime('%H:%M:%S')
        self.log_box.insert('end', f'[{ts}] ', 'dim')
        self.log_box.insert('end', text + '\n', tag)
        self.log_box.see('end')
        self.log_box.config(state='disabled')

    def _log_clear(self):
        self.log_box.config(state='normal')
        self.log_box.delete('1.0', 'end')
        self.log_box.config(state='disabled')


# ─────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────

def main():
    root = tk.Tk()

    # Icono (ignorar si no existe)
    try:
        root.iconbitmap('icon.ico')
    except Exception:
        pass

    app = InvestmentReportGUI(root)
    root.mainloop()


if __name__ == '__main__':
    main()
